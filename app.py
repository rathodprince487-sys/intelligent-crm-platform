import streamlit as st
import pandas as pd
import requests
import plotly.express as px
import time
from datetime import datetime
import subprocess
import sys
import os
import socket
import platform
import json
import streamlit.components.v1 as components
from components.auth import AuthManager, render_login_page, render_admin_panel, Role
from components.sidebar import render_sidebar_toggle
from components.profile_page import render_profile_page
from components.analytics import render_analytics_dashboard
from components.tasks import render_tasks_module


# --- PLAYWRIGHT SETUP FOR CLOUD ---
# Ensure browsers are installed (essential for Streamlit Cloud)
if not os.path.exists(".browser_installed"):
    try:
        print("🔧 Installing Playwright Browsers (First Run)...")
        subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=True)
        with open(".browser_installed", "w") as f: 
            f.write("done")
        print("✅ Playwright Browsers Installed.")
    except Exception as e:
        print(f"⚠️ Failed to install browsers: {e}")

# ================== BACKEND AUTO-START ==================
@st.cache_resource
def start_backend_server():
    """
    Checks if the backend is running on port 3000.
    If not, starts 'node server.js' as a background process.
    """
    def is_port_in_use(port):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            return s.connect_ex(('localhost', port)) == 0

    if not is_port_in_use(3000):
        # Start server
        print("🚀 Starting Backend & Dependencies...")
        
        # 1. Install Node modules if missing
        if not os.path.exists("node_modules"):
            print("📦 Installing Node modules...")
            try:
                subprocess.run(["npm", "install"], check=True)
                print("✅ Node dependencies installed.")
            except Exception as e:
                print(f"❌ npm install failed: {e}")

        # 2. Install Playwright Browsers (CRITICAL for Scraper)
        # We check a flag or just run it (it's fast if already installed)
        print("🎭 Ensuring Playwright browsers...")
        try:
             subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=True)
             print("✅ Playwright chromium installed.")
        except Exception as e:
             print(f"⚠️ Playwright install warning: {e}")

        # 3. Start Server
        try:
            # Run node server.js in background
            subprocess.Popen(
                ["node", "server.js"], 
                cwd=os.getcwd(),
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL
            )
            time.sleep(5)  # Give it 5 seconds to boot
            print("✅ Backend launch command sent.")
        except Exception as e:
            print(f"❌ Failed to start backend: {e}")
    else:
        print("✅ Backend already running.")

# Initialize Backend
start_backend_server()

# ================== CONFIG ==================
BACKEND_BASE = os.getenv("BACKEND_URL", "http://localhost:3000")
EXECUTIONS_API = f"{BACKEND_BASE}/executions"
LEADS_API = f"{BACKEND_BASE}/leads"
STATS_API = f"{BACKEND_BASE}/stats"
LEAD_GEN_API = f"{BACKEND_BASE}/lead-gen"

# Detect Environment
IS_LIVE_ENV = platform.system() == "Linux"

st.set_page_config(
    page_title="n8n CRM & Sales Engine",
    layout="wide",
    initial_sidebar_state="auto"
)

# --- AUTHENTICATION CHECK ---
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

import extra_streamlit_components as stx

# --- AUTHENTICATION & SESSION PERSISTENCE ---
# Initialize Cookie Manager
if 'cookie_manager_stx' not in st.session_state:
    st.session_state.cookie_manager_stx = stx.CookieManager(key="stx_cookie_manager_auth")
cookie_manager = st.session_state.cookie_manager_stx

# --- LOGOUT HANDLING ---
if st.session_state.get('logout_clicked', False):
    print("DEBUG: Processing Logout...")
    try:
        cookie_manager.set('crm_user', 'logged_out')
        time.sleep(0.2)
        cookie_manager.delete('crm_user')
    except Exception as e:
        print(f"Logout error: {e}")
    
    st.session_state["authenticated"] = False
    st.session_state["user"] = None
    st.session_state["just_logged_out"] = True
    st.session_state['logout_clicked'] = False
    time.sleep(1) # Final sync delay
    st.rerun()

# Check for existing session cookie if not authenticated
if not st.session_state.get("authenticated", False):
    # Prevent immediate re-login if user just logged out
    if st.session_state.get("just_logged_out", False):
        st.session_state["just_logged_out"] = False 
        cookies = {}
    else:
        # Try to get cookie (Note: get() is async-like in Streamlit components, might need rerun)
        # But usually works if component renders.
        # We render component implicitly via init? No, stx.CookieManager puts a component on page.
        # We must call get() which returns value.
        cookies = cookie_manager.get_all()
    print(f"DEBUG COOKIES: {cookies}")
    cookie_user = cookies.get('crm_user') if cookies else None
    
    if cookie_user:
        # Validate User
        temp_auth = AuthManager()
        user_obj = temp_auth.users.get(cookie_user)
        if user_obj and user_obj.status != "Inactive":
            st.session_state["user"] = user_obj
            st.session_state["authenticated"] = True
            st.session_state["active_crm_selection"] = "My CRM"
            st.rerun()

    # Still not authenticated? Show Login
    if not st.session_state.get("authenticated", False):
        render_login_page(cookie_manager)
        st.stop()

# Initialize Auth Manager
auth_manager = AuthManager()
current_user = st.session_state["user"]

# Render the collapsible sidebar toggle (Must be called early)
render_sidebar_toggle()


# ================== STYLES ==================

# 1. COMMON STYLES (Layout, Fonts, Transitions - Always Applied)
COMMON_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}


header[data-testid="stHeader"] {
    background: transparent !important;
}

div[data-testid="stDecoration"] {
    display: none;
}





.stMarkdown h1 a, .stMarkdown h2 a, .stMarkdown h3 a, 
.stMarkdown h4 a, .stMarkdown h5 a, .stMarkdown h6 a {
    display: none !important;
    opacity: 0 !important;
    pointer-events: none !important;
}
a.anchor-link, a[class*="anchor"] {
    display: none !important;
}

h1 a svg, h2 a svg, h3 a svg {
    display: none !important;
}





section[data-testid="stSidebar"] > div:first-child {
    padding-left: 0.5rem !important;
    padding-right: 0.5rem !important;
}
section[data-testid="stSidebar"] {
    padding-left: 0 !important;
    padding-right: 0 !important;
}






@media only screen and (max-width: 768px) {
    .block-container {
        padding-left: 1rem !important; 
        padding-right: 1rem !important;
        padding-top: 1rem !important; 
        padding-bottom: 2rem !important;
        max-width: 100vw !important;
    }
    div[data-testid="column"] {
        width: 100% !important;
        flex: 1 1 auto !important;
        min-width: 100% !important;
        margin-bottom: 1rem !important;
    }
    h1 { font-size: 1.75rem !important; line-height: 1.2 !important; }
    h2 { font-size: 1.5rem !important; }
    h3 { font-size: 1.25rem !important; }
    p, li, label, .stMarkdown { font-size: 1rem !important; }
    .metric-card {
        width: 100% !important;
        margin-bottom: 12px !important;
        padding: 16px !important; 
    }
    .meeting-card {
        width: 100% !important;
        flex-direction: row !important; 
        align-items: flex-start !important;
        padding: 10px !important;
    }
    .mc-details {
        width: 100% !important;
        overflow: hidden;
    }
    div[data-testid="stDataFrame"], 
    div[data-testid="stTable"] {
        width: 100% !important;
        overflow-x: auto !important; 
        display: block !important;
    }
    .stTextInput input, .stSelectbox div[data-testid="stMarkdownContainer"], .stButton button {
        min-height: 48px !important; 
        font-size: 16px !important; 
    }
    div.stButton > button {
        width: 100% !important;
    }
    footer { display: none !important; }
    div[data-testid="stDecoration"] { display: none !important; }
    .js-plotly-plot, .plot-container {
        width: 100% !important;
    }
    section[data-testid="stSidebar"] {
        /* Let Streamlit manage height naturally */
    }
    button[kind="header"] {
        display: block !important; 
    }
    header[data-testid="stHeader"] {
        z-index: 100000 !important;
    }
}
</style>
"""

# 2. LIGHT MODE STYLES (Strictly for Light Mode)
LIGHT_CSS = """
<style>

.stApp {
    background: linear-gradient(to bottom right, #ffffff 0%, #f4f6f9 100%) !important;
    color: #1a1a1a !important;
}

.stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {
    color: #0f172a !important; 
}
.stApp p, .stApp span, .stApp div, .stApp label, .stApp li, .stApp td, .stApp th {
    color: #334155 !important; 
}
.stApp a {
    color: #2563eb !important; 
}


section[data-testid="stSidebar"] {
    background-color: #f8f9fa !important;
    border-right: 1px solid rgba(0,0,0,0.05);
}
section[data-testid="stSidebar"] h1, section[data-testid="stSidebar"] h2, section[data-testid="stSidebar"] h3, 
section[data-testid="stSidebar"] span, section[data-testid="stSidebar"] div, section[data-testid="stSidebar"] label, section[data-testid="stSidebar"] p {
    color: #333333 !important;
}





.stTextInput > div > div > input, .stSelectbox > div > div > div, .stTextArea > div > div > textarea {
    background-color: #ffffff !important;
    color: #1a1a1a !important;
    border: 1px solid #e0e0e0 !important;
    border-radius: 4px !important;
}

.stTextInput > div > div > input:focus, 
.stSelectbox > div > div > div:focus {
    border-color: #1a73e8 !important;
    box-shadow: 0 0 0 2px rgba(26, 115, 232, 0.2) !important;
}



.metric-card {
    background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
    border: 1px solid #e2e8f0;
    border-radius: 16px;
    padding: 24px;
    text-align: center;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
    transition: transform 0.2s ease, box-shadow 0.2s ease;
    height: 100%;
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
}
.metric-card:hover {
    transform: translateY(-3px);
    box-shadow: 0 10px 20px -5px rgba(0, 0, 0, 0.1);
    border-color: #cbd5e1;
}
.metric-icon {
    font-size: 1.5rem;
    margin-bottom: 12px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 48px;
    height: 48px;
    border-radius: 12px;
    background: #eff6ff; 
    color: #3b82f6; 
}
.metric-label {
    font-size: 0.75rem;
    font-weight: 700;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 4px;
}
.metric-value {
    font-size: 2.25rem;
    font-weight: 800;
    color: #0f172a;
    line-height: 1.1;
}


.meeting-card {
    display: flex;
    flex-direction: row;
    align-items: center;
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-left: 4px solid #10b981; 
    border-radius: 12px;
    padding: 12px;
    margin-bottom: 12px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.03);
    transition: all 0.2s ease;
    cursor: pointer; 
}
.meeting-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 16px rgba(0,0,0,0.08);
    border-color: #cbd5e1;
}
.mc-date-box {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    min-width: 52px;
    height: 52px;
    background: #f1f5f9;
    border-radius: 10px;
    margin-right: 12px;
    flex-shrink: 0;
}
.mc-month {
    font-size: 0.65rem;
    font-weight: 700;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    line-height: 1;
    margin-bottom: 2px;
}
.mc-day {
    font-size: 1.3rem;
    font-weight: 700;
    color: #1e293b;
    line-height: 1;
}
.mc-details {
    flex: 1;
    display: flex;
    flex-direction: column;
    overflow: hidden; 
    min-width: 0; 
}
.mc-title { 
    color: #1a202c !important; 
    font-weight: 600;
    font-size: 0.95rem;
    margin-bottom: 4px;
    white-space: nowrap; 
    text-overflow: ellipsis;
    line-height: 1.2;
}
.mc-company { 
    color: #64748b !important;
    font-size: 0.85rem;
    margin-bottom: 0; 
    white-space: nowrap;
    text-overflow: ellipsis;
    line-height: 1.3;
}


div[data-testid="stDataFrame"] {
    background-color: #ffffff !important;
    border: 1px solid #eee !important;
}
div[data-testid="stDataFrame"] th {
    background-color: #f8f9fa !important;
    color: #444 !important;
}
div[data-testid="stDataFrame"] td {
    background-color: #ffffff !important;
    color: #333 !important;
    border-bottom: 1px solid #f0f0f0 !important;
}


div[data-testid="stDataEditor"], [data-testid="stDataEditor"] * {
    background-color: #ffffff !important;
    color: #000000 !important;
    border-color: #e2e3e3 !important; 
    font-family: 'Roboto', sans-serif !important;
    font-size: 13px !important; 
}
div[data-testid="stDataEditor"] canvas {
    filter: none !important;
}


div[data-testid="stDataEditor"] thead tr th, 
[role="columnheader"], 
div[data-testid="stDataFrame"] thead tr th {
    background-color: #f8f9fa !important; 
    color: #444 !important;
    font-weight: 700 !important;
    text-align: center !important;
    border-right: 1px solid #ccc !important;
    border-bottom: 2px solid #ccc !important;
    font-size: 13px !important;
}


div[data-testid="stDataEditor"] tbody th {
    background-color: #f8f9fa !important;
    color: #666 !important;
    text-align: center !important;
    border-right: 2px solid #ccc !important;
}



div[data-testid="stFileUploader"] {
    background-color: #ffffff !important;
}
div[data-testid="stFileUploader"] > div {
    background-color: #f8f9fa !important;
    border: 2px dashed #cbd5e0 !important;
    border-radius: 8px !important;
}
div[data-testid="stFileUploader"] section {
    background-color: #f8f9fa !important;
    color: #1a1a1a !important;
}
div[data-testid="stFileUploader"] small {
    color: #4a5568 !important;
}
div[data-testid="stFileUploader"] svg {
    fill: #4a5568 !important;
}

div[data-testid="stFileUploader"] span,
div[data-testid="stFileUploader"] p,
div[data-testid="stFileUploader"] label {
    color: #2d3748 !important;
}

div[data-testid="stFileUploader"] button {
    background-color: #ffffff !important;
    color: #1a1a1a !important;
    border: 1px solid #cbd5e0 !important;
}
div[data-testid="stFileUploader"] button:hover {
    background-color: #f1f3f5 !important;
}



div[data-testid="stExpander"] {
    background-color: #ffffff !important;
    border: 1px solid #e0e0e0 !important;
}
div[data-testid="stExpander"] summary {
    background-color: #f8f9fa !important;
    color: #1a1a1a !important;
}


code {
    background-color: #f1f3f5 !important;
    color: #1a1a1a !important;
}
pre {
    background-color: #f8f9fa !important;
    border: 1px solid #e0e0e0 !important;
}










.stApp section[data-testid="stMain"] div[data-testid="stRadio"] div[role="radiogroup"] {
    background-color: #222222 !important; 
    padding: 4px !important;
    border-radius: 50px !important;
    display: flex !important;
    flex-direction: row !important;
    gap: 0px !important;
    border: 1px solid #444444 !important; 
    box-shadow: 0 4px 12px rgba(0,0,0,0.15) !important;
    width: fit-content !important;
    margin-left: auto !important;
    margin-top: 12px !important; 
}

.stApp section[data-testid="stMain"] div[data-testid="stRadio"] div[role="radiogroup"] label {
    padding: 8px 20px !important; 
    border-radius: 50px !important;
    margin: 0 !important;
    border: none !important;
    background-color: transparent !important;
    color: #aaaaaa !important;
    transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1) !important;
    text-align: center !important;
    flex: 1 !important;
    justify-content: center !important;
    font-weight: 600 !important;
    white-space: nowrap !important;
    font-size: 0.9rem !important;
}

.stApp section[data-testid="stMain"] div[data-testid="stRadio"] div[role="radiogroup"] label[data-checked="true"] {
    background-color: #ffffff !important;
    color: #000000 !important;
    font-weight: 800 !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.1) !important;
    transform: scale(1.02);
}

.stApp section[data-testid="stMain"] div[data-testid="stRadio"] div[role="radiogroup"] label:hover {
    color: #ffffff !important;
}


div[role="radiogroup"] label span,
div[role="radiogroup"] label p,
div[role="radiogroup"] label div {
    color: inherit !important;
}

</style>
"""

# 3. DARK MODE STYLES (Strictly for Dark Mode)
DARK_CSS = """
<style>

.stApp {
    background: linear-gradient(to bottom right, #0f172a 0%, #020617 100%) !important;
    color: #cbd5e1 !important; 
}


.stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {
    color: #f8fafc !important; 
}
.stApp p, .stApp span, .stApp div, .stApp label, .stApp li, .stApp td, .stApp th {
    color: #cbd5e1 !important; 
}
.stApp a {
    color: #60a5fa !important; 
}
.stApp a:hover {
    color: #93c5fd !important;
}

.stApp .stCaption, .stApp small, .stApp [data-testid="stCaption"] {
    color: #94a3b8 !important;
}

.stApp code, .stApp pre {
    background-color: #1e293b !important;
    color: #e2e8f0 !important;
}


section[data-testid="stSidebar"] {
    background-color: rgba(15, 23, 42, 0.98) !important; 
    border-right: 1px solid rgba(255, 255, 255, 0.05) !important;
}

section[data-testid="stSidebar"] h1, section[data-testid="stSidebar"] h2, section[data-testid="stSidebar"] h3, 
section[data-testid="stSidebar"] span, section[data-testid="stSidebar"] div, section[data-testid="stSidebar"] label, section[data-testid="stSidebar"] p {
    color: #e2e8f0 !important; 
}





.stTextInput > div > div > input, .stSelectbox > div > div > div, .stTextArea > div > div > textarea {
    background-color: rgba(30, 41, 59, 0.6) !important; 
    color: #f1f5f9 !important; 
    border: 1px solid rgba(255, 255, 255, 0.08) !important;
    border-radius: 8px !important;
}

.stTextInput > div > div > input:focus, .stSelectbox > div > div > div:focus, .stTextArea > div > div > textarea:focus {
    border-color: #fbbf24 !important; 
    box-shadow: 0 0 0 2px rgba(251, 191, 36, 0.2) !important;
}



.metric-card {
    background: rgba(30, 41, 59, 0.4) !important;
    border: 1px solid rgba(255, 255, 255, 0.05) !important;
    backdrop-filter: blur(10px);
    border-radius: 16px;
    padding: 24px;
    text-align: center;

    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.2);
    transition: transform 0.2s ease;
    height: 100%;
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
}
.metric-card:hover {
    transform: translateY(-3px);
    background: rgba(30, 41, 59, 0.6) !important;
    border-color: rgba(251, 191, 36, 0.2) !important; 
}
.metric-icon {
    font-size: 1.5rem;
    margin-bottom: 12px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 48px;
    height: 48px;
    border-radius: 12px;
    background: rgba(255, 255, 255, 0.05);
    color: #e2e8f0;
}
.metric-label {
    font-size: 0.75rem;
    font-weight: 700;
    color: #94a3b8 !important;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 4px;
}
.metric-value {
    font-size: 2.25rem;
    font-weight: 800;
    color: #f8fafc !important;
    line-height: 1.1;
}
.meeting-card {
    background: rgba(30, 41, 59, 0.4) !important;
    border: 1px solid rgba(255, 255, 255, 0.05) !important;
    border-left: 4px solid #10b981 !important; 
    border-radius: 12px;
    padding: 12px;
    margin-bottom: 10px;
    display: flex;
    flex-direction: row;
    align-items: center;
}
.mc-date-box {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    min-width: 52px;
    height: 52px;
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.05);
    border-radius: 10px;
    margin-right: 12px;
    flex-shrink: 0;
}
.mc-month {
    font-size: 0.65rem;
    font-weight: 700;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    line-height: 1;
    margin-bottom: 2px;
}
.mc-day {
    font-size: 1.3rem;
    font-weight: 700;
    color: #e2e8f0;
    line-height: 1;
}
.mc-details {
    flex: 1;
    display: flex;
    flex-direction: column;
    overflow: hidden;
}
.mc-title { 
    color: #f8fafc !important; 
    font-weight: 600; 
    font-size: 0.95rem;
    margin-bottom: 4px;
    white-space: normal;
    line-height: 1.2;
    word-break: break-word;
}
.mc-company { 
    color: #94a3b8 !important; 
    font-size: 0.85rem;
    margin-bottom: 8px;
    white-space: normal;
    line-height: 1.3;
}
.mc-action {
    display: inline-flex;
    align-items: center;
    gap: 4px;
    background: rgba(16, 185, 129, 0.1);
    color: #34d399;
    border: 1px solid rgba(16, 185, 129, 0.2);
    font-size: 0.75rem;
    font-weight: 600;
    padding: 4px 10px;
    border-radius: 6px;
    text-decoration: none !important;
    align-self: flex-start;
    transition: all 0.2s;
}
.mc-action:hover {
    background: rgba(16, 185, 129, 0.2);
    color: #6ee7b7;
}


div[data-testid="stDataFrame"], div[data-testid="stTable"], div[data-testid="stDataEditor"] {
    background-color: transparent !important;
    border: 1px solid rgba(255, 255, 255, 0.05) !important;
}

div[data-testid="stDataFrame"] th, 
div[data-testid="stTable"] th,
div[data-testid="stDataFrame"] thead tr th,
div[data-testid="stDataFrame"] thead th,
table thead tr th,
table th,
.stDataFrame th,
[data-testid="stDataFrame"] table th,
[data-testid="stTable"] table th {
    background-color: #1e293b !important; 
    color: #f8fafc !important;
    font-weight: 600 !important;
    border-bottom: 1px solid #334155 !important;
}

div[data-testid="stDataFrame"] td, 
div[data-testid="stTable"] td,
table tbody tr td,
table td {
    background-color: #0f172a !important; 
    color: #cbd5e1 !important;
    border-bottom: 1px solid #1e293b !important;
}


[data-testid="stDataFrame"] thead th,
[data-testid="stTable"] thead th,
[data-testid="stDataEditor"] thead th,
[data-testid="stDataEditor"] th,
.stDataFrame thead th,
table thead th,
th {
    background-color: #1e293b !important; 
    color: #ffffff !important;
    border-bottom: 1px solid #334155 !important;
}

[data-testid="stDataEditor"] td,
[data-testid="stDataEditor"] table tbody tr td {
    background-color: #0f172a !important;
    color: #cbd5e1 !important;
    border-bottom: 1px solid #1e293b !important;
}


div[class*="stButton"] > button, 
div[class*="stDownloadButton"] > button,
button[kind="primary"],
button[kind="secondary"] {
    background-color: transparent !important;
    background: transparent !important;
    color: #475569 !important; 
    border: 1px solid rgba(0, 0, 0, 0.15) !important; 
    border-radius: 50px !important; 
    font-weight: 600 !important;
    padding: 0.4rem 1.2rem !important;
    backdrop-filter: none; 
    transition: all 0.2s ease;
    box-shadow: none !important;
}

div[class*="stButton"] > button:hover, 
div[class*="stDownloadButton"] > button:hover {
    background-color: rgba(0, 0, 0, 0.05) !important; 
    border-color: rgba(0, 0, 0, 0.3) !important; 
    color: #0f172a !important; 
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05) !important;
}


div[class*="stButton"] > button[kind="primary"] {
    border-width: 2px !important;
    border-color: rgba(0, 0, 0, 0.2) !important;
    color: #0f172a !important;
}



div[data-testid="stFileUploader"] > div {
    background-color: rgba(30, 41, 59, 0.3) !important;
    border: 2px dashed rgba(255, 255, 255, 0.1) !important;
    border-radius: 12px !important;
}
div[data-testid="stFileUploader"] section {
    background-color: transparent !important;
}
div[data-testid="stFileUploader"] svg { fill: #94a3b8 !important; }
div[data-testid="stFileUploader"] small { color: #64748b !important; }
div[data-testid="stFileUploader"] span, div[data-testid="stFileUploader"] p { color: #cbd5e1 !important; }


div[role="radiogroup"] label, 
div[role="radiogroup"] label p,
div[data-testid="stRadio"] label p {
    color: #f1f5f9 !important; 
    font-weight: 500 !important;
}

label[data-testid="stLabel"], 
div[data-testid="stWidgetLabel"] label,
div[data-testid="stWidgetLabel"] p {
    color: #e2e8f0 !important;
    font-weight: 600 !important;
    font-size: 14px !important;
}

div[role="radiogroup"] div[aria-checked="true"] label p {
    color: #f59e0b !important; 
    font-weight: 700 !important;
}


div[data-testid="stExpander"] {
    background-color: transparent !important;
    border: 1px solid rgba(255, 255, 255, 0.08) !important;
}
div[data-testid="stExpander"] summary {
    color: #e2e8f0 !important;
}
div[data-testid="stExpander"] summary:hover {
    color: #fff !important;
    background-color: rgba(255, 255, 255, 0.02) !important;
}


div[data-testid="stProgress"] > div > div {
    background-color: #f59e0b !important; 
}


div[data-testid="stAlert"] {
    background-color: rgba(30, 41, 59, 0.5) !important;
    border: 1px solid rgba(255, 255, 255, 0.1) !important;
    color: #e2e8f0 !important;
}


div[role="radiogroup"] {
    gap: 6px !important;
}
div[role="radiogroup"] label {
    background-color: transparent !important;
    border: 1px solid transparent !important;
    padding: 8px 12px !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
    color: #94a3b8 !important;
    transition: all 0.2s ease;
}



</style>
"""

# 4. SHARED RESPONSIVE CSS (Mobile Optimization - Enhanced)
RESPONSIVE_CSS = """
<style>

html {
    scroll-behavior: smooth !important;
    -webkit-overflow-scrolling: touch !important;
}

body {
    -webkit-font-smoothing: antialiased !important;
    -moz-osx-font-smoothing: grayscale !important;
}


@media only screen and (max-width: 768px) {
    
    
    .block-container {
        padding-top: 1rem !important;
        padding-left: 0.75rem !important;
        padding-right: 0.75rem !important;
        padding-bottom: 3rem !important;
        max-width: 100vw !important;
        overflow-x: hidden !important;
    }

    
    h1 { 
        font-size: 1.5rem !important; 
        line-height: 1.3 !important;
        margin-bottom: 0.75rem !important;
    }
    h2 { 
        font-size: 1.25rem !important;
        margin-bottom: 0.5rem !important;
    }
    h3 { 
        font-size: 1.1rem !important;
        margin-bottom: 0.5rem !important;
    }
    p, div, label { 
        font-size: 14px !important;
        line-height: 1.5 !important;
    }
    
    
    div[data-testid="column"] {
        width: 100% !important;
        flex: 1 1 auto !important;
        min-width: 100% !important;
        margin-bottom: 0.75rem !important;
    }
    
    
    div[data-testid="stVerticalBlock"]:has(.css-actions-row) div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {
        width: auto !important;
        min-width: unset !important;
        flex: 1 !important;
        margin-bottom: 0 !important;
    }

    
    div.stButton > button {
        width: 100% !important;
        min-height: 50px !important;
        margin-bottom: 0.5rem !important;
        font-size: 15px !important;
        padding: 12px 20px !important;
        border-radius: 8px !important;
        transition: all 0.2s ease !important;
    }
    
    div.stButton > button:active {
        transform: scale(0.98) !important;
    }
    
    
    div[data-testid="stVerticalBlock"]:has(.css-actions-row) div.stButton > button {
         width: auto !important;
         min-height: 48px !important;
         padding: 10px 16px !important;
    }
    
    
    div[data-testid="stTextInput"] input, 
    div[data-testid="stSelectbox"] div[data-baseweb="select"],
    div[data-testid="stTextArea"] textarea {
        min-height: 50px !important;
        font-size: 16px !important;
        padding: 12px !important;
        border-radius: 8px !important;
    }
    
    
    div[data-testid="stDataFrame"], 
    div[data-testid="stDataEditor"] {
        width: 100% !important;
        overflow-x: auto !important;
        display: block !important;
        -webkit-overflow-scrolling: touch !important;
        border-radius: 8px !important;
    }
    
    
    section[data-testid="stSidebar"] {
        z-index: 99999 !important;
    }
    
    section[data-testid="stSidebar"] > div {
        padding-top: 1rem !important;
    }
    
    

    
    
    div[data-testid="stVerticalBlock"] > div {
        gap: 0.75rem !important;
    }
    
    
    .metric-card {
        padding: 1rem !important;
        margin-bottom: 0.75rem !important;
        border-radius: 12px !important;
    }
    
    .metric-icon {
        width: 40px !important;
        height: 40px !important;
        font-size: 1.25rem !important;
    }
    
    .metric-label {
        font-size: 0.875rem !important;
    }
    
    .metric-value {
        font-size: 1.5rem !important;
    }
    
    
    .meeting-card {
        padding: 0.75rem !important;
        margin-bottom: 0.5rem !important;
    }
    
    .mc-date-box {
        min-width: 50px !important;
        padding: 0.5rem !important;
    }
    
    .mc-month {
        font-size: 0.7rem !important;
    }
    
    .mc-day {
        font-size: 1.25rem !important;
    }
    
    .mc-title, .mc-company {
        font-size: 0.875rem !important;
    }
    
    
    iframe {
        max-width: 100% !important;
        height: auto !important;
    }
    
    
    div[data-testid="stExpander"] {
        border-radius: 8px !important;
    }
    
    div[data-testid="stExpander"] summary {
        min-height: 48px !important;
        padding: 12px !important;
        font-size: 15px !important;
    }
    
    
    div[data-baseweb="select"] > div {
        min-height: 50px !important;
    }
    
    
    div[role="radiogroup"] label {
        min-height: 44px !important;
        padding: 10px 12px !important;
        margin-bottom: 6px !important;
    }
    
    
    div[data-testid="stFileUploader"] {
        padding: 1rem !important;
    }
    
    div[data-testid="stFileUploader"] button {
        min-height: 50px !important;
        width: 100% !important;
    }
    
    
    div[data-baseweb="tab-list"] button {
        min-height: 48px !important;
        padding: 12px 16px !important;
        font-size: 14px !important;
    }
    
    
    div[data-testid="stAlert"] {
        padding: 1rem !important;
        border-radius: 8px !important;
        font-size: 14px !important;
    }
    
    
    div[data-testid="stProgress"] {
        height: 8px !important;
    }
    
    
    div[data-testid="stSpinner"] {
        min-height: 60px !important;
    }
}


@media only screen and (min-width: 769px) and (max-width: 1024px) {
    .block-container {
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    
    h1 { font-size: 1.75rem !important; }
    h2 { font-size: 1.5rem !important; }
    h3 { font-size: 1.25rem !important; }
}
</style>
"""

# ================== HELPER FUNCTIONS ==================
def metric_card(label, value, icon="📊", color="blue"):
    """
    Renders a styled metric card.
    color options: blue, green, amber, purple, rose
    """
    
    # Define color maps for Light/Dark styling adjustments if needed
    # using inline css vars or just relying on helper classes could work, 
    # but let's do a simple inline override for the icon bg
    
    color_map = {
        "blue":   {"bg": "#eff6ff", "text": "#3b82f6", "dark_bg": "rgba(59, 130, 246, 0.1)", "dark_text": "#60a5fa"},
        "green":  {"bg": "#f0fdf4", "text": "#22c55e", "dark_bg": "rgba(34, 197, 94, 0.1)",  "dark_text": "#4ade80"},
        "amber":  {"bg": "#fffbeb", "text": "#f59e0b", "dark_bg": "rgba(245, 158, 11, 0.1)", "dark_text": "#fbbf24"},
        "purple": {"bg": "#f3e8ff", "text": "#a855f7", "dark_bg": "rgba(168, 85, 247, 0.1)", "dark_text": "#c084fc"},
        "rose":   {"bg": "#ffe4e6", "text": "#f43f5e", "dark_bg": "rgba(244, 63, 94, 0.1)",  "dark_text": "#fb7185"},
    }
    
    c = color_map.get(color, color_map["blue"])
    
    # We use a unique ID or class to inject specific colors for this card instance? 
    # Actually, simpler to just inject style attribute on the icon.
    
    # Check Theme (Python side check for rendering slightly different inline styles if strictly needed, 
    # but CSS classes are cleaner. Let's send both vars and let CSS vars handle it? 
    # Streamlit doesn't support CSS vars easily dynamically.
    # We'll use the session state theme.
    
    theme = st.session_state.get("theme", "light")
    icon_bg = c["dark_bg"] if theme == "dark" else c["bg"]
    icon_col = c["dark_text"] if theme == "dark" else c["text"]
    
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-icon" style="background-color: {icon_bg}; color: {icon_col};">
            {icon}
        </div>
        <div class="metric-label">{label}</div>
        <div class="metric-value">{value}</div>
    </div>
    """, unsafe_allow_html=True)

@st.cache_data(ttl=30, show_spinner=False)
def fetch_data(url):
    try:
        r = requests.get(url, timeout=3)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return []

def update_lead(lead_id, data):
    try:
        requests.put(f"{LEADS_API}/{lead_id}", json=data)
        return True
    except:
        return False

def create_lead(data):
    try:
        requests.post(f"{BACKEND_BASE}/leads", json=data)
        return True
    except:
        return False

# Normalization Helpers
def normalize_text(text):
    if pd.isna(text): return ""
    return str(text).strip().lower()

def normalize_phone(phone):
    if pd.isna(phone): return ""
    s = str(phone).strip()
    # Remove chars
    import re
    s = re.sub(r'[\s\-\+\(\)]', '', s)
    # Remove leading 91 (Indian code) if 12 digits (91 + 10 digits) or just starts with 91 and is long
    if s.startswith('91') and len(s) > 10:
        s = s[2:]
    return s

def delete_lead(lead_id):
    try:
        requests.delete(f"{LEADS_API}/{lead_id}")
        return True
    except:
        return False

import urllib.parse

# ================== STATUS THEME CONFIGURATION ==================
# COLORS UPDATED TO MATCH USER REFERENCE IMAGE (VIBRANT)
STATUS_PALETTE = {
    "Interested":          {"bg": "#DFF5E1", "txt": "#1B5E20"}, # Light Green
    "Not picking":         {"bg": "#F0F0F0", "txt": "#616161"}, # Light Grey
    "Asked to call later": {"bg": "#FFF8E1", "txt": "#8D6E00"}, # Light Yellow
    "Meeting set":         {"bg": "#E3F2FD", "txt": "#0D47A1"}, # Light Blue
    "Meeting Done":        {"bg": "#E0F2F1", "txt": "#004D40"}, # Teal
    "Proposal sent":       {"bg": "#F3E5F5", "txt": "#4A148C"}, # Lavender
    "Follow-up scheduled": {"bg": "#FFE0B2", "txt": "#E65100"}, # Soft Orange
    "Not interested":      {"bg": "#FDECEA", "txt": "#B71C1C"}, # Light Red
    "Closed - Won":        {"bg": "#C8E6C9", "txt": "#1B5E20"}, # Strong Green
    "Closed - Lost":       {"bg": "#ECEFF1", "txt": "#37474F"}, # Soft Dark Grey
    "Generated":           {"bg": "#F5F5F5", "txt": "#616161"}, # Neutral
}

# Add normalized keys to palette (En-dash variants)
STATUS_PALETTE["Closed – Won"] = STATUS_PALETTE["Closed - Won"]
STATUS_PALETTE["Closed – Lost"] = STATUS_PALETTE["Closed - Lost"]

def get_status_colors(theme):
    """
    Returns a dictionary mapping status strings to CSS style strings.
    """
    css_map = {}
    for status, colors in STATUS_PALETTE.items():
        # User defined fixed colors (ignores theme)
        style = f"background-color: {colors['bg']}; color: {colors['txt']}; font-weight: 600; padding: 4px 12px; border-radius: 12px; text-align: center;"
        css_map[status] = style
        
    return css_map

# ================== THEME INJECTION ==================
# Force "light" theme to ensure consistent UI rendering across all components
st.session_state.theme = "light"

# --- INIT ZOOM STATE ---
if "zoom_level" not in st.session_state:
    st.session_state.zoom_level = 90 # Default comfortable zoom

st.markdown(COMMON_CSS, unsafe_allow_html=True)

# Force Light CSS
st.markdown(LIGHT_CSS, unsafe_allow_html=True)
st.markdown(RESPONSIVE_CSS, unsafe_allow_html=True)

# 🚀 FORCE BUTTON STYLES (Injected Last to Override Everything)
st.markdown("""

        <style>
        /* Compress whitespace at the top of the app globally */
        .block-container {
            padding-top: 1rem !important;
            margin-top: -10px !important;
        }
        /* Compress the Streamlit header space */
        header[data-testid="stHeader"] {
            height: 2.5rem !important;
            min-height: 2.5rem !important;
        }

div[data-testid="stButton"] > button, 
div[data-testid="stDownloadButton"] > button,
div[class*="stButton"] > button,
div[class*="stDownloadButton"] > button,
div[data-testid="stTooltipHoverTarget"] > button,
div[data-testid="stPopover"] > button,
button[data-testid="stPopoverButton"] {
    background-color: transparent !important;
    background: transparent !important;
    color: #334155 !important; 
    border: 1px solid #cbd5e1 !important; 
    border-radius: 50px !important;
    font-weight: 600 !important;
    padding: 0.35rem 1.1rem !important;
    box-shadow: none !important;
    transition: all 0.2s ease;
}

div[data-testid="stButton"] > button:hover, 
div[data-testid="stDownloadButton"] > button:hover,
div[class*="stButton"] > button:hover,
div[class*="stDownloadButton"] > button:hover,
div[data-testid="stTooltipHoverTarget"] > button:hover,
div[data-testid="stPopover"] > button:hover,
button[data-testid="stPopoverButton"]:hover {
    background-color: rgba(0, 0, 0, 0.04) !important;
    border-color: #94a3b8 !important; 
    color: #0f172a !important; 
    box-shadow: 0 2px 4px rgba(0,0,0,0.05) !important;
    transform: translateY(-1px);
}


div[data-testid="stButton"] > button[kind="primary"],
div[class*="stButton"] > button[kind="primary"],
div[data-testid="stTooltipHoverTarget"] > button[kind="primary"],
div[data-testid="stPopover"] > button[kind="primary"] {
    background-color: transparent !important;
    border: 2px solid #f59e0b !important; 
    color: #d97706 !important; 
}
div[data-testid="stButton"] > button[kind="primary"]:hover,
div[class*="stButton"] > button[kind="primary"]:hover,
div[data-testid="stTooltipHoverTarget"] > button[kind="primary"]:hover,
div[data-testid="stPopover"] > button[kind="primary"]:hover {
    background-color: rgba(245, 158, 11, 0.1) !important;
    box-shadow: 0 4px 12px rgba(245, 158, 11, 0.15) !important;
}
</style>
""", unsafe_allow_html=True)

# Global Padding Fix (Move content higher)
st.markdown("""
<style>

.block-container {
    padding-top: 1.5rem !important;
    padding-bottom: 2rem !important;
}

h1 {
    padding-top: 0 !important;
}
</style>
""", unsafe_allow_html=True)

# ================== SIDEBAR NAVIGATION ==================
# ================== SIDEBAR NAVIGATION ==================
import base64
import os

def get_sidebar_img_b64(file_path):
    if os.path.exists(file_path):
        with open(file_path, "rb") as f:
            data = f.read()
        return base64.b64encode(data).decode()
    return None

sb_logo_dark = get_sidebar_img_b64("logo.png")
sb_logo_light = get_sidebar_img_b64("logo_light.png")

# Force Light Logo since we killed dark mode
target_logo = sb_logo_light if sb_logo_light else sb_logo_dark

if target_logo:
    st.sidebar.markdown(f"""
    <div style="margin-bottom: 20px;">
        <img src="data:image/png;base64,{target_logo}" style="width: 100%; max-width: 240px; height: auto;">
    </div>
    """, unsafe_allow_html=True)
elif sb_logo_dark:
    st.sidebar.image("logo.png", width=240)
else:
    st.sidebar.title("SHDPIXEL")

# Helper for Animated Icons
# Helper for Animated Icons
def get_icon_html(icon_name, width=60, class_name=""):
    # Check paths
    possible_paths = [f"assets/icons/{icon_name}", f"assets/{icon_name}"]
    icon_path = None
    for p in possible_paths:
        if os.path.exists(p):
            icon_path = p
            break
            
    if icon_path:
        b64 = get_sidebar_img_b64(icon_path)
        return f'<img src="data:image/png;base64,{b64}" class="{class_name}" style="width: {width}px; height: auto; vertical-align: middle; margin-right: 15px;">'
    return ""

# Navigation
# Import Component
from components.email_verifier import render_email_verifier

# Navigation
# --- MODERN SIDEBAR NAVIGATION ---
# Hide default Streamlit sidebar nav
st.markdown("""
<style>
    [data-testid="stSidebarNav"] {display: none;}
    button[data-testid="collapsedControl"],
    button[data-testid="stSidebarCollapseButton"],
    button[data-testid="stExpandSidebarButton"] {
        display: none !important;
    }
    section[data-testid="stSidebar"] {
        padding-top: 1rem; 
        background-color: #f8f9fa;
        border-right: 1px solid #e5e7eb;
    }
</style>
""", unsafe_allow_html=True)

with st.sidebar:
    # 1. Profile Section with Popover (Dropdown)
    profile_container = st.container()
    with profile_container:
        # --- NEW MODERN PROFILE UI ---
        # 1. Prepare Avatar Data (Base64 or Initials)
        avatar_path = getattr(current_user, "avatar_file", None)
        b64_avatar = None
        user_initial = current_user.name[0].upper() if current_user.name else "U"
        
        has_image = False
        if avatar_path and os.path.exists(avatar_path):
             import base64
             with open(avatar_path, "rb") as f:
                 b64_avatar = base64.b64encode(f.read()).decode()
             has_image = True

        # 2. Add Popover (The "Button")
        # Use Name as label so it renders in DOM, allowing us to style it as the primary text.
        with st.popover(current_user.name, use_container_width=True):
             st.markdown(f"**{current_user.name}**")
             st.caption(current_user.role)
             if st.button("View Profile", key="btn_view_profile", use_container_width=True):
                 st.session_state.nav_selection = "Profile Settings"
                 st.rerun()
             if st.button("Logout", key="btn_logout_pop", type="primary", use_container_width=True):
                 st.session_state['logout_clicked'] = True
                 st.rerun()

        # 3. Inject CSS to Transform the Button into a Profile Card
        
        avatar_css_rule = ""
        if has_image:
            avatar_css_rule = f"""
                background-image: url("data:image/png;base64,{b64_avatar}") !important;
                background-size: contain !important;
                background-position: center !important;
                background-repeat: no-repeat !important;
                border: none !important;
                content: "" !important; /* CRITICAL: Override default icon text */
                color: transparent !important;
            """
        else:
            # Initials Avatar styling
            avatar_css_rule = f"""
                background-color: #EEF2FF !important;
                color: #4F46E5 !important;
                content: "{user_initial}" !important;
                text-align: center !important;
                line-height: 38px !important;
                font-weight: 600 !important;
                font-size: 16px !important; 
                border: none !important;
            """

        css_profile = f"""
        <style>
            /* Reduce Sidebar Top Padding */
            section[data-testid="stSidebar"] .block-container {{
                padding-top: 1.5rem !important; 
                padding-bottom: 3rem !important;
            }}

            /* Target the profile popover button inside sidebar ONLY */
            section[data-testid="stSidebar"] [data-testid="stPopoverButton"] {{
                border: none !important;
                background: transparent !important;
                box-shadow: none !important;
                padding: 0 !important;
                padding-left: 56px !important; /* Space for Avatar */
                height: 52px !important;
                min-height: 52px !important;
                display: flex !important;
                align-items: flex-start !important; /* Align text top */
                justify-content: center !important; /* Vertical center via flex direction? No, we use column padding */
                flex-direction: column !important;
                text-align: left !important;
                transition: background 0.2s ease;
                border-radius: 8px !important;
                margin-bottom: 0 !important;
                overflow: visible !important;
                position: relative !important;
            }}
            
            section[data-testid="stSidebar"] button[data-testid="stPopoverButton"]:hover {{
                background-color: #F8F9FA !important;
            }}

            /* AVATAR (::before) - Absolute Positioned */
            section[data-testid="stSidebar"] [data-testid="stPopoverButton"]::before {{
                content: "" !important; /* Force override of Streamlit's default icon */
                position: absolute;
                left: 8px;
                top: 50%;
                transform: translateY(-50%);
                width: 38px;
                height: 38px;
                border-radius: 50%;
                {avatar_css_rule}
                display: block !important;
                z-index: 2;
                box-shadow: 0 0 0 1px #E5E7EB; /* Add subtle border */
            }}

            /* NAME (Default Streamlit Text) */
            section[data-testid="stSidebar"] button[data-testid="stPopoverButton"] div[data-testid="stMarkdownContainer"] p {{
                font-weight: 600 !important;
                font-size: 14px !important;
                color: #1F2937 !important; /* Darker grey */
                margin: 0 !important;
                padding: 0 !important;
                line-height: 1.2 !important;
                position: absolute;
                top: 8px; /* Adjusted top */
                left: 58px;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
                max-width: 150px;
                pointer-events: none;
            }}
            
            /* ROLE (::after) - Absolute Positioned below Name */
            section[data-testid="stSidebar"] button[data-testid="stPopoverButton"]::after {{
                content: "{current_user.role}";
                position: absolute;
                left: 58px;
                top: 26px; /* Adjusted to sit nicely below name */
                color: #6B7280;
                font-size: 12px;
                font-weight: 400;
                line-height: 1.2;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
                max-width: 150px;
                display: block !important;
                pointer-events: none;
            }}
            
            /* HIDE ALL UNWANTED ICONS/IMAGES (The "Person Emoji" or Chevron) */
            section[data-testid="stSidebar"] button[data-testid="stPopoverButton"] svg,
            section[data-testid="stSidebar"] button[data-testid="stPopoverButton"] img,
            section[data-testid="stSidebar"] button[data-testid="stPopoverButton"] span[data-testid="stIcon"],
            section[data-testid="stSidebar"] button[data-testid="stPopoverButton"] span[data-testid="stPopoverIcon"] {{
                display: none !important;
                opacity: 0 !important;
                width: 0 !important;
                height: 0 !important;
                margin: 0 !important;
                pointer-events: none !important;
            }}


             /* --- COLLAPSED STATE OVERRIDES --- */
             /* Hide Name */
             body.sidebar-mj-collapsed section[data-testid="stSidebar"] button[data-testid="stPopoverButton"] div[data-testid="stMarkdownContainer"] p {{
                 display: none !important;
             }}
             
             /* Hide Role */
             body.sidebar-mj-collapsed section[data-testid="stSidebar"] button[data-testid="stPopoverButton"]::after {{
                 display: none !important;
             }}

             /* Center Avatar */
             body.sidebar-mj-collapsed section[data-testid="stSidebar"] button[data-testid="stPopoverButton"]::before {{
                 left: 50% !important;
                 transform: translate(-50%, -50%) !important;
                 margin: 0 !important;
             }}
             
             /* Reset container padding */
             body.sidebar-mj-collapsed section[data-testid="stSidebar"] button[data-testid="stPopoverButton"] {{
                 padding-left: 0 !important;
             }}
        </style>
        """
        st.markdown(css_profile, unsafe_allow_html=True)

    # Custom Divider with reduced spacing
    st.markdown('<div style="margin-top: 4px; margin-bottom: 12px; border-top: 1px solid #E5E7EB;"></div>', unsafe_allow_html=True)

    # 2. Define Menu Options
    opts_map = {}
    
    # Core
    opts_map["Dashboard"] = "Dashboard"
    opts_map["My CRM"] = "My CRM"
    opts_map["SHD PIXEL"] = "SHD PIXEL"
    opts_map["Search Leads"] = "Search Leads"
    opts_map["Tasks"] = "Tasks"

        
    # Management
    if current_user.role in [Role.HR, Role.CEO]:
        opts_map["Team CRMs"] = "Team CRMs"
        opts_map["User Mgmt"] = "User Management"
        
        if current_user.role == Role.CEO:
             opts_map["Analytics"] = "Analytics"

    # Tools
    opts_map["Power Dialer"] = "Power Dialer"
    opts_map["Email Verifier"] = "Email Verifier"
    opts_map["Google Maps Scraper"] = "Google Maps Scraper"
    opts_map["Spreadsheet Tool"] = "Spreadsheet Tool"
    opts_map["Websites"] = "Websites"
    
    # Determine Index
    current_val = st.session_state.get('nav_selection', 'Dashboard')
    nav_keys = list(opts_map.keys())
    nav_index = 0
    
    display_keys = list(opts_map.values())
    if current_val in display_keys:
        for k, v in opts_map.items():
            if v == current_val:
                nav_index = nav_keys.index(k)
                break
    elif current_val == "CRM Grid":
         current_path = st.session_state.get("current_crm_path", current_user.username)
         if current_path == "shared/shd_pixel" and "SHD PIXEL" in nav_keys:
             try: nav_index = nav_keys.index("SHD PIXEL")
             except: pass
         elif current_path == current_user.username and "My CRM" in nav_keys:
             try: nav_index = nav_keys.index("My CRM")
             except: pass
         elif "Team CRMs" in nav_keys:
             try: nav_index = nav_keys.index("Team CRMs")
             except: pass
    elif current_val == "Profile Settings":
        nav_index = None
    elif current_val == "Scraped Leads":
        if "Search Leads" in nav_keys:
            try: nav_index = nav_keys.index("Search Leads")
            except: pass
    elif current_val == "Admin Panel":
        if "User Mgmt" in nav_keys:
            try: nav_index = nav_keys.index("User Mgmt")
            except: pass
    
    # --- DYNAMIC ICON INJECTION ---
    # Map menu items to FontAwesome codes
    icon_map = {
        "Dashboard": "\\f015",          # fa-home
        "My CRM": "\\f0c0",             # fa-users (or fa-user-group in v6)
        "SHD PIXEL": "\\f1ad",          # fa-building
        "Search Leads": "\\f002",       # fa-search
        "Tasks": "\\f0ae",              # fa-tasks
        "Team CRMs": "\\f500",          # fa-user-tie
        "User Mgmt": "\\f508",          # fa-user-shield
        "User Management": "\\f508",
        "Analytics": "\\f201",          # fa-chart-line
        "Power Dialer": "\\f095",       # fa-phone
        "Email Verifier": "\\f003",     # fa-envelope
        "Google Maps Scraper": "\\f3c5",# fa-map-marker-alt
        "Spreadsheet Tool": "\\f1c3",   # fa-file-excel
        "Websites": "\\f0ac"            # fa-globe
    }
    
    # Generate CSS for icons
    icon_css = """<style>
        /* Compact Menu Spacing */
        section[data-testid="stSidebar"] div[role="radiogroup"] {
            margin-top: -10px !important;
            padding-top: 0 !important;
            gap: 0px !important; /* Remove gap entirely, control via padding */
        }
        
        /* Reduce individual item height and set up label bounds */
        section[data-testid="stSidebar"] div[role="radiogroup"] label {
            padding: 8px 12px !important;
            margin-bottom: 2px !important;
            border-radius: 8px !important;
            transition: all 0.2s ease !important;
            cursor: pointer !important;
            border: 1px solid transparent !important;
        }
        
        /* HOVER STATE */
        section[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
            background-color: #F8FAFC !important;
        }
        
        /* SUPPRESS ALL NATIVE STREAMLIT RADIO BG/BORDER ARTIFACTS */
        section[data-testid="stSidebar"] div[role="radiogroup"] label > div:first-child {
            display: none !important; /* Hide native circle/pill background wrapper entirely */
        }
        section[data-testid="stSidebar"] div[role="radiogroup"] label[data-checked="true"] p {
            color: inherit !important;
        }
    """
    
    # Append Base Icon Classes
    for i, key in enumerate(nav_keys):
        # Default icon if not found
        code = icon_map.get(key, "\\f111") # fa-circle 
        icon_css += f"""
        section[data-testid="stSidebar"] div[role="radiogroup"] label:nth-child({i+1})::before {{
            content: "{code}" !important;
            transition: color 0.2s ease !important;
        }}
        """
        
    # Append the REAL Active Override based on nav_index
    if nav_index is not None:
        icon_css += f"""
        /* Explicity Highlight the Logically Active Nav Index */
        section[data-testid="stSidebar"] div[role="radiogroup"] label:nth-child({nav_index+1}) {{
            background-color: #EFF6FF !important;
            background-image: linear-gradient(90deg, #3B82F6 0%, #3B82F6 4px, transparent 4px) !important;
            box-shadow: 0 1px 2px rgba(0,0,0,0.02) !important;
            border: 1px solid #DBEAFE !important;
        }}
        /* Keep hover nice on active */
        section[data-testid="stSidebar"] div[role="radiogroup"] label:nth-child({nav_index+1}):hover {{
            background-color: #E0EFFF !important;
        }}
        section[data-testid="stSidebar"] div[role="radiogroup"] label:nth-child({nav_index+1}) div[data-testid="stMarkdownContainer"] p {{
            color: #1D4ED8 !important;
            font-weight: 700 !important;
        }}
        section[data-testid="stSidebar"] div[role="radiogroup"] label:nth-child({nav_index+1})::before {{
            color: #1D4ED8 !important;
        }}
        """
        
    icon_css += "</style>"
    st.markdown(icon_css, unsafe_allow_html=True)

    # 3. Render Navigation
    selected_label = st.radio("Navigation", nav_keys, index=nav_index, label_visibility="collapsed")
    
    if selected_label:
        selected_val = opts_map[selected_label]
        
        # Logic Mapping
        if selected_val == 'Dashboard':
            if st.session_state.get('nav_selection') != "Dashboard":
                st.session_state.nav_selection = "Dashboard"
                st.rerun()

        elif selected_val == 'My CRM':
            st.session_state.nav_selection = "CRM Grid"
            new_path = current_user.username
            if st.session_state.get("current_crm_path") != new_path:
                st.session_state.current_crm_path = new_path
                st.session_state.crm_sheets = {}
                st.rerun()

        elif selected_val == 'SHD PIXEL':
            st.session_state.nav_selection = "CRM Grid"
            new_path = "shared/shd_pixel"
            if st.session_state.get("current_crm_path") != new_path:
                st.session_state.current_crm_path = new_path
                st.session_state.crm_sheets = {}
                st.rerun()

        elif selected_val == 'Team CRMs':
            rerun_needed = False
            if st.session_state.get("nav_selection") != "CRM Grid":
                st.session_state.nav_selection = "CRM Grid"
                rerun_needed = True

            # Auto-switch to the first team CRM if the user is currently looking at their own or SHD
            acc_crms = auth_manager.get_accessible_crms(current_user)
            team_crms = [c for c in acc_crms if c not in ("My CRM", "SHD PIXEL")]
            
            if team_crms:
                new_path = auth_manager.resolve_crm_path(team_crms[0], current_user)
                resolved_all_teams = [auth_manager.resolve_crm_path(tc, current_user) for tc in team_crms]
                
                if st.session_state.get("current_crm_path") not in resolved_all_teams:
                    st.session_state.current_crm_path = new_path
                    st.session_state.crm_sheets = {}
                    rerun_needed = True
                    
            if rerun_needed:
                st.rerun()
        elif selected_val == 'Search Leads':
            if st.session_state.get('nav_selection') != "Scraped Leads":
                st.session_state.nav_selection = "Scraped Leads"
                st.rerun()
        
        elif selected_val == 'User Management':
            if st.session_state.get('nav_selection') != "Admin Panel":
                st.session_state.nav_selection = "Admin Panel"
                st.rerun()
        
        elif selected_val == 'Tasks':
            if st.session_state.get('nav_selection') != "Tasks":
                st.session_state.nav_selection = "Tasks"
                st.rerun()

        elif selected_val in ['Power Dialer', 'Email Verifier', 'Google Maps Scraper', 'Spreadsheet Tool', 'Analytics', 'Websites']:
             if st.session_state.get('nav_selection') != selected_val:
                 st.session_state.nav_selection = selected_val
                 st.rerun()

    ICON_MAP = {
        "Dashboard": "\\f201", "My CRM": "\\f1c0", "SHD PIXEL": "\\f135",
        "Search Leads": "\\f002", "Tasks": "\\f0ae", "User Mgmt": "\\f509",
        "Analytics": "\\f200", "Power Dialer": "\\f095", "Email Verifier": "\\f658",
        "Google Maps Scraper": "\\f3c5", "Spreadsheet Tool": "\\f0ce",
        "Team CRMs": "\\f500", "Websites": "\\f0ac"
    }
    css_icons = ""
    for idx, key in enumerate(nav_keys):
        icon_code = ICON_MAP.get(key, "\\f111")
        css_icons += f"""
        section[data-testid="stSidebar"] div[role="radiogroup"] label:nth-child({idx+1})::before {{ content: "{icon_code}" !important; }}
        body.sidebar-mj-collapsed section[data-testid="stSidebar"] div[role="radiogroup"] label:nth-child({idx+1}):hover::after {{ content: "{key}" !important; }}
        """
    st.markdown(f"<style>{css_icons}</style>", unsafe_allow_html=True)

# 6. Global Page Handler
if st.session_state.get("nav_selection") == "Profile Settings":
    render_profile_page(auth_manager, current_user)

page = st.session_state.nav_selection




# Hide Decoration Bar Only (Keep Sidebar Toggle)
st.markdown("""
<style>
div[data-testid="stDecoration"] {
    visibility: hidden !important;
}
</style>
""", unsafe_allow_html=True)



# --- UPCOMING MEETINGS WIDGET ---
st.sidebar.markdown("""
    <style>
    @keyframes marquee {
        0% { transform: translateX(0); }
        100% { transform: translateX(-25%); }
    }
    .upcoming-header-base {
        margin-top: 16px; 
        margin-bottom: 8px; 
        color: #9CA3AF; 
        font-size: 0.75rem; 
        font-weight: 600; 
        letter-spacing: 0.05em; 
        text-transform: uppercase;
    }
    .header-static {
        display: flex; 
        align-items: center; 
        gap: 6px;
    }
    .header-scroll {
        display: none;
        overflow: hidden;
        white-space: nowrap;
        mask-image: linear-gradient(to right, transparent, black 10%, black 90%, transparent);
        -webkit-mask-image: linear-gradient(to right, transparent, black 10%, black 90%, transparent);
    }
    .header-scroll-inner {
        display: inline-block;
        animation: marquee 10s linear infinite;
    }
    body.sidebar-mj-collapsed .header-static {
        display: none !important;
    }
    body.sidebar-mj-collapsed .header-scroll {
        display: block !important;
    }
    </style>

    <div class="upcoming-header-base header-static">
        <span style="font-size: 1rem;">📅</span> UPCOMING MEETINGS
    </div>
    
    <div class="upcoming-header-base header-scroll">
        <div class="header-scroll-inner">
            <span>📅</span> UPCOMING MEETINGS &nbsp;&nbsp;&nbsp; <span>📅</span> UPCOMING MEETINGS &nbsp;&nbsp;&nbsp; <span>📅</span> UPCOMING MEETINGS &nbsp;&nbsp;&nbsp; <span>📅</span> UPCOMING MEETINGS &nbsp;&nbsp;&nbsp;
        </div>
    </div>
""", unsafe_allow_html=True)
try:
    # Quick lightweight check
    m_leads_raw = fetch_data(LEADS_API)
    if m_leads_raw:
        m_df = pd.DataFrame(m_leads_raw)
        if "meetingDate" in m_df.columns:
            m_df["meetingDate"] = pd.to_datetime(m_df["meetingDate"], errors='coerce').dt.date
            today = datetime.now().date()
            
            # Filter: Date exists AND is >= Today
            # We treat NaT as null
            future_meetings = m_df[
                (m_df["meetingDate"].notna()) & 
                (m_df["meetingDate"] >= today)
            ].sort_values("meetingDate")
            
            if future_meetings.empty:
                st.sidebar.caption("No upcoming meetings.")
            else:
                meetings_html = '<div style="max-height: 350px; overflow-y: auto; padding-right: 4px;">'
                
                for idx, row in future_meetings.iterrows():
                    # Create Link
                    m_date = row["meetingDate"]
                    name = str(row.get("contactName", "Lead")).strip() or "Lead"
                    comp = str(row.get("businessName", "Company")).strip() or "Company"
                    
                    # Generate Link
                    start_str = m_date.strftime("%Y%m%d")
                    end_str = (m_date + pd.Timedelta(days=1)).strftime("%Y%m%d")
                    title = urllib.parse.quote(f"Meeting with {name} ({comp})")
                    details = urllib.parse.quote(f"Phone: {row.get('phone', '')}\nAddress: {row.get('address', '')}")
                    cal_url = f"https://calendar.google.com/calendar/render?action=TEMPLATE&text={title}&details={details}&dates={start_str}/{end_str}"
                    
                    # Append Card HTML (No indentation to prevent Markdown code block rendering)
                    meetings_html += f"""
<a href="{cal_url}" target="_blank" style="text-decoration: none; color: inherit;">
<div class="meeting-card">
<div class="mc-date-box">
<span class="mc-month">{m_date.strftime('%b').upper()}</span>
<span class="mc-day">{m_date.strftime('%d')}</span>
</div>
<div class="mc-details" style="min-width: 0;">
<div class="mc-title" title="{comp}" style="white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">{comp}</div>
<div class="mc-company" title="{name}" style="white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">👤 {name}</div>
</div>
</div>
</a>"""
                
                meetings_html += "</div>"
                st.sidebar.markdown(meetings_html, unsafe_allow_html=True)
except Exception as e:
    # Fail silently to not break main app
    st.sidebar.caption(f"Syncing calendar... ({e})")

# ================== ADMIN PANEL ==================
if "Admin Panel" in page:
    render_admin_panel(auth_manager, current_user)
    st.stop()

# ================== DASHBOARD ==================
if "Dashboard" in page:
    st.markdown(f"""
    <div style="margin-bottom: 20px; margin-top: 0px;">
        <h1 style="margin: 0; padding: 0; font-size: 2.2rem; font-weight: 800; display: flex; align-items: center;">
            {get_icon_html('rocket.png', 70)}
            <span>Sales & Growth Engine</span>
        </h1>
        <p style="margin-left: 85px; margin-top: -5px; opacity: 0.7; font-size: 1rem;">
            End-to-end command center: from data mining to closed deals
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Fetch all leads directly to calculate custom splits
    leads = fetch_data(LEADS_API)
    df_all = pd.DataFrame(leads)
    
    # ALWAYS RENDER CARDS (Empty or Not)
    # Split Data
    if df_all.empty:
        df_generated = pd.DataFrame(columns=["id", "status"])
        df_crm = pd.DataFrame(columns=["id", "status"])
    else:
        # 1. Generated (Fresh leads, unworked)
        df_generated = df_all[df_all['status'] == 'Generated']
        # 2. CRM (Everything else)
        df_crm = df_all[df_all['status'] != 'Generated']
    
    # --- SECTION 1: LEAD GENERATION ---
    st.subheader("⚡ Lead Generation (Data Mining)")
    
    # Calculate Recent Imports (CRM)
    recent_crm_count = 0
    if not df_all.empty and 'createdAt' in df_all.columns:
        # Convert to datetime and strip timezone for comparison
        df_all["createdAt"] = pd.to_datetime(df_all["createdAt"], errors='coerce').dt.tz_localize(None)
        last_24h = datetime.now() - pd.Timedelta(hours=24)
        recent_crm_count = len(df_all[df_all["createdAt"] >= last_24h])

    # Calculate Fresh Scraped Leads (From History, not yet in CRM)
    executions = fetch_data(EXECUTIONS_API)
    scraped_today_count = 0
    if executions:
        df_exec = pd.DataFrame(executions)
        if 'date' in df_exec.columns and 'leadsGenerated' in df_exec.columns:
            df_exec["date"] = pd.to_datetime(df_exec["date"], errors='coerce').dt.tz_localize(None)
            last_24h = datetime.now() - pd.Timedelta(hours=24)
            # Sum up leads generated since yesterday
            scraped_today_count = df_exec[df_exec["date"] >= last_24h]["leadsGenerated"].fillna(0).astype(int).sum()

    c1, c2, c3 = st.columns(3)
    with c1: metric_card("Fresh Scraped (Today)", int(scraped_today_count), icon="✨", color="blue")
    with c2: metric_card("Imported to CRM (Today)", recent_crm_count, icon="📥", color="purple")
    
    # --- SECTION 2: CRM PIPELINE ---
    st.markdown("---")
    st.subheader("💼 Active Pipeline (CRM)")
    
    # Calculate CRM specific metrics
    crm_total = len(df_crm)
    
    hot_leads = 0
    meetings = 0
    closed_won = 0
    
    if not df_crm.empty:
        # Hot Leads (Robust Check)
        if 'priority' in df_crm.columns:
            hot_leads = len(df_crm[df_crm['priority'].astype(str).str.upper() == 'HOT'])
            
        # Meetings (Robust Check)
        if 'status' in df_crm.columns:
            meetings = len(df_crm[df_crm['status'].astype(str).str.contains("Meeting", case=False, na=False)])
            # Closed Won (Robust Check for 'Won' or 'Closed - Won')
            closed_won = len(df_crm[df_crm['status'].astype(str).str.contains("Won", case=False, na=False)])
    
    c_a, c_b, c_c, c_d = st.columns(4)
    with c_a: metric_card("In Pipeline", crm_total, icon="📊", color="blue")
    with c_b: metric_card("Hot Opportunities", hot_leads, icon="🔥", color="rose")
    with c_c: metric_card("Meetings Set", meetings, icon="📅", color="amber")
    with c_d: metric_card("Closed Won", closed_won, icon="🏆", color="green")

    # st.markdown("---")
    # if breakdown:
    #     ...

# ================== CRM GRID (PIPELINE) ==================
# ================== CRM GRID (NEW GOOGLE SHEETS STYLE) ==================
if "CRM Grid" in page:
    # --- 1. SESSION STATE INITIALIZATION ---
    if "crm_data" not in st.session_state:
        st.session_state.crm_data = pd.DataFrame()
    if "show_add_lead" not in st.session_state:
        st.session_state.show_add_lead = False
        
    # --- 2. GOOGLE SHEETS STYLING (CSS) ---
    st.markdown("""
    <style>
    
    .stApp {
        background-color: #f1f3f4 !important;
    }
    
    
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 0rem !important;
        padding-left: 1rem !important;
        padding-right: 1rem !important;
        max-width: 100% !important;
    }

    
    .crm-menu-bar {
        font-family: 'Roboto', sans-serif;
        font-size: 13px !important;
        color: #5f6368 !important;
        display: flex;
        gap: 15px;
        margin-top: -5px;
        cursor: pointer;
    }
    
    
    div[data-testid="stHorizontalBlock"] > div {
        align-items: center;
    }
    
    
    .stTextInput input, .stSelectbox div[data-baseweb="select"] > div, .stDateInput input {
        background-color: white !important;
        border-radius: 4px !important;
        border: 1px solid #dadce0 !important;
        color: #3c4043 !important;
        font-size: 14px !important;
        min-height: 36px;
    }
    
    .stTextInput input:focus, .stSelectbox div[data-baseweb="select"] > div:focus-within {
        border-color: #1a73e8 !important;
        box-shadow: 0 0 0 1px #1a73e8 !important;
    }

    
    div[data-testid="stButton"] button {
        border-radius: 4px !important;
        font-size: 14px !important;
        font-weight: 500 !important;
        padding: 6px 16px !important;
        border: 1px solid #dadce0 !important;
        background-color: white !important;
        color: #3c4043 !important;
        transition: all 0.2s;
        height: 36px;
    }
    div[data-testid="stButton"] button:hover {
        background-color: #f8f9fa !important;
        border-color: #dadce0 !important;
        color: #202124 !important;
        box-shadow: 0 1px 2px rgba(60,64,67,0.3) !important;
    }
    
    
    button[kind="primary"] {
        background-color: #1a73e8 !important;
        border-color: #1a73e8 !important;
        color: white !important;
    }
    button[kind="primary"]:hover {
        background-color: #185abc !important;
        border-color: #185abc !important;
        box-shadow: 0 1px 3px rgba(66,133,244,0.3) !important;
        color: white !important;
    }

    
    div[data-testid="stDataEditor"] {
        border: 1px solid #e0e0e0 !important;
        background: white !important;
        border-radius: 0px !important;
    }
    
    
    div[data-testid="stDataEditor"] thead tr th {
        background-color: #f8f9fa !important;
        color: #3c4043 !important;
        font-weight: 700 !important;
        font-size: 13px !important;
        border-bottom: 2px solid #e0e0e0 !important;
        border-right: 1px solid #e0e0e0 !important;
        text-align: center !important;
    }
    
    
    div[data-testid="stDataEditor"] tbody td {
        font-family: 'Roboto', sans-serif !important;
        font-size: 13px !important;
        color: #202124 !important;
        border-right: 1px solid #e0e0e0 !important;
        border-bottom: 1px solid #e0e0e0 !important;
        padding: 6px 8px !important;
    }
    
    
    div[data-testid="stDataEditor"] tbody tr:hover td {
        background-color: #e8f0fe !important;
    }
    
    
    div[data-testid="stDataEditor"] tbody th {
        background-color: #f8f9fa !important;
        color: #5f6368 !important;
        border-right: 2px solid #e0e0e0 !important;
        font-size: 12px;
    }
    
    
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    </style>
    """, unsafe_allow_html=True)
    
    # --- 2b. ADD LEAD DIALOG ---
    # --- 2b. ADD LEAD DIALOG (Redesigned) ---
    # --- 2b. ADD LEAD DIALOG (Modern SaaS Style) ---
    @st.dialog("Add New Lead")
    def add_lead_modal():
        # --- 1. MODERN CSS ---
        st.markdown("""
            <style>
            
            div[data-testid="stDialog"] div[role="dialog"] {
                width: 800px !important;
                max-width: 90vw !important;
                border-radius: 16px !important;
                padding: 0 !important;
                background-color: white !important;
                box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04) !important;
            }
            
            
            div[data-testid="stForm"] {
                border: none !important;
                padding: 0 16px !important;
            }
            
            
            .crm-modal-header {
                padding-bottom: 12px;
                border-bottom: 1px solid #e5e7eb;
                margin-bottom: 20px;
            }
            .crm-modal-title {
                font-size: 24px;
                font-weight: 600;
                color: #111827;
                margin: 0;
            }
            .crm-modal-subtitle {
                font-size: 14px;
                color: #6b7280;
                margin-top: 2px;
            }
            
            
            .crm-section-title {
                font-size: 12px;
                text-transform: uppercase;
                letter-spacing: 0.05em;
                color: #6b7280;
                font-weight: 600;
                margin-top: 16px;
                margin-bottom: 8px;
                border-bottom: 1px dashed #e5e7eb;
                padding-bottom: 4px;
            }

            
            .stTextInput input, .stSelectbox div[data-baseweb="select"] > div, .stDateInput input {
                height: 42px !important;
                border-radius: 8px !important;
                border: 1px solid #d1d5db !important;
                background-color: #fafafa !important;
                color: #111827 !important;
                font-size: 14px !important;
                transition: all 0.2s;
            }
            .stTextInput input:focus, .stSelectbox div[data-baseweb="select"] > div:focus-within, .stDateInput input:focus {
                border-color: #3b82f6 !important;
                box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1) !important;
                background-color: white !important;
            }
            
            
            .stTextArea textarea {
                border-radius: 8px !important;
                border: 1px solid #d1d5db !important;
                background-color: #fafafa !important;
                font-size: 14px !important;
            }
            .stTextArea textarea:focus {
                border-color: #3b82f6 !important;
                background-color: white !important;
            }

            
            label p {
                font-size: 13px !important;
                font-weight: 500 !important;
                color: #374151 !important;
            }
            
            
            div[data-testid="stFormSubmitButton"] {
                border: none !important;
                background: transparent !important;
            }
            
            
            div[data-testid="stFormSubmitButton"] button {
                border-radius: 8px !important;
                height: 44px !important;
                font-weight: 500 !important;
                border: 1px solid transparent !important;
                transition: all 0.2s ease-in-out !important;
                padding: 0 16px !important;
            }
            
            
            
            div[data-testid="stFormSubmitButton"] button[kind="primary"] {
                background-color: #1a73e8 !important;
                color: white !important;
                border: none !important;
                box-shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.1), 0 1px 2px -1px rgba(0, 0, 0, 0.1) !important;
            }
            div[data-testid="stFormSubmitButton"] button[kind="primary"]:hover {
                background-color: #1557b0 !important;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06) !important;
                transform: translateY(-1px);
            }
            div[data-testid="stFormSubmitButton"] button[kind="primary"]:active {
                background-color: #1a73e8 !important;
                transform: translateY(0px);
            }

            
            
            div[data-testid="stFormSubmitButton"] button[kind="secondary"] {
                background-color: white !important;
                color: #374151 !important;
                border: 1px solid #d1d5db !important;
            }
            div[data-testid="stFormSubmitButton"] button[kind="secondary"]:hover {
                background-color: #f9fafb !important;
                border-color: #9ca3af !important;
                color: #111827 !important;
            }
            div[data-testid="stFormSubmitButton"] button[kind="secondary"]:active {
                background-color: #f3f4f6 !important;
            }
            
            
            div[data-testid="stFormSubmitButton"] button:not([kind="primary"]) {
                background-color: white !important;
                color: #374151 !important;
                border: 1px solid #d1d5db !important;
            }
            </style>
        """, unsafe_allow_html=True)

        # --- 2. HEADER ---
        st.markdown("""
        <div class="crm-modal-header">
            <!-- <div class="crm-modal-title">Add New Lead</div> -->
            <div class="crm-modal-subtitle">Fields based on current sheet structure</div>
        </div>
        """, unsafe_allow_html=True)
        
        # --- 3. DYNAMIC LOGIC ---
        current_sheet = st.session_state.active_sheet
        df = st.session_state.crm_sheets[current_sheet]
        columns = [c for c in df.columns if c not in ["id", "created_at", "updated_at", "➕"] and not c.startswith("Unnamed")]
        
        # Grouping Logic
        groups = {
            "Basic Info": ["businessName", "contactName", "phone", "email", "website"],
            "Location": ["address", "location", "city", "country", "zip"],
            "Lead Details": ["status", "priority", "source", "query", "dealValue"],
            "Assignment": ["calledBy", "meetingBy", "closedBy", "assignedTo"],
            "Follow-up": ["lastFollowUpDate", "nextFollowUpDate", "meetingDate"],
            "Other": []
        }
        
        # Sort columns into groups
        sorted_cols = {k: [] for k in groups}
        assigned_cols = set()
        
        for col in columns:
            found = False
            for group, keywords in groups.items():
                # Check explicit match first, then keyword containment
                norm_col = col.lower().replace(" ", "")
                if any(k.lower() == norm_col for k in keywords) or \
                   (group == "Basic Info" and any(x in norm_col for x in ["name", "phone", "email"])) or \
                   (group == "Location" and any(x in norm_col for x in ["address", "loc", "city"])) or \
                   (group == "Lead Details" and any(x in norm_col for x in ["status", "source", "deal"])) or \
                   (group == "Follow-up" and "date" in norm_col):
                        if col not in assigned_cols:
                            sorted_cols[group].append(col)
                            assigned_cols.add(col)
                            found = True
                        break
            if not found:
                sorted_cols["Other"].append(col)

        form_values = {}
        
        with st.form("dynamic_add_lead_form_saas", border=False):
            # Iterate through groups
            for group_name, group_cols in sorted_cols.items():
                if not group_cols: continue
                
                # Render Section Title
                st.markdown(f'<div class="crm-section-title">{group_name}</div>', unsafe_allow_html=True)
                
                # Render Fields in 2-Col Grid
                # Logic: collect normal fields, flush full-width fields immediately
                
                # --- Widget Makers ---
                input_buffer = [] 
                def flush_buffer(buf):
                    if not buf: return
                    cols = st.columns(len(buf))
                    for idx, (c_name, output_dict, widget_maker) in enumerate(buf):
                        with cols[idx]:
                            output_dict[c_name] = widget_maker()
                
                for col in group_cols:
                    c_lower = col.lower()
                    
                    is_full_width = False
                    widget_maker = None
                    
                    # Define Widget Types
                    if any(x in c_lower for x in ["notes", "address", "description", "details", "query"]):
                        is_full_width = True
                        def widget_maker(c=col): return st.text_area(c, height=100)
                    elif "status" in c_lower:
                        def widget_maker(c=col): return st.selectbox(c, ["New", "Contacted", "Qualified", "Proposal", "Negotiation", "Closed-Won", "Closed-Lost"])
                    elif "priority" in c_lower:
                        def widget_maker(c=col): return st.selectbox(c, ["HOT", "WARM", "COLD"])
                    elif any(x in c_lower for x in ["date", "follow", "at"]):
                         def widget_maker(c=col): return st.date_input(c, value=None)
                    elif "email" in c_lower:
                        def widget_maker(c=col): return st.text_input(c, placeholder="name@company.com")
                    elif any(x in c_lower for x in ["phone", "mobile"]):
                        def widget_maker(c=col): return st.text_input(c, placeholder="+1 ...")
                    else:
                        def widget_maker(c=col): return st.text_input(c)

                    # Render
                    if is_full_width:
                        flush_buffer(input_buffer)
                        input_buffer = []
                        form_values[col] = widget_maker()
                    else:
                        input_buffer.append((col, form_values, widget_maker))
                        if len(input_buffer) == 2:
                            flush_buffer(input_buffer)
                            input_buffer = []

                flush_buffer(input_buffer) # Flush remaining in group

            st.markdown("<br>", unsafe_allow_html=True) # Spacer
            
            # --- 4. BUTTONS ---
            c_cancel, c_submit = st.columns([1, 2])
            with c_submit:
                submitted = st.form_submit_button("Save Lead", type="primary", use_container_width=True)
            with c_cancel:
                if st.form_submit_button("Cancel", use_container_width=True):
                    st.rerun()

            if submitted:
                # Same Sync Logic as before (kept clean)
                errors = []
                for k, v in form_values.items():
                    if "email" in k.lower() and v and "@" not in str(v):
                         errors.append(f"⚠️ Invalid email format in '{k}'")
                
                if errors:
                    for e in errors: st.error(e)
                else:
                    new_row = form_values.copy()
                    # Formatting
                    for k, v in new_row.items():
                        if hasattr(v, 'isoformat'): new_row[k] = v.isoformat()
                        elif v is None: new_row[k] = ""
                    
                    if "created_at" not in new_row: new_row["created_at"] = datetime.now().isoformat()
                    
                    # Update Local
                    st.session_state.crm_sheets[current_sheet] = pd.concat(
                        [st.session_state.crm_sheets[current_sheet], pd.DataFrame([new_row])], 
                        ignore_index=True
                    )
                    
                    # --- AUTO SAVE TO DISK (Corrected) ---
                    try:
                        ctx_path = st.session_state.get("current_crm_path", current_user.username)
                        full_ctx_dir = os.path.join("crm_data", ctx_path)
                        if not os.path.exists(full_ctx_dir): os.makedirs(full_ctx_dir, exist_ok=True)
                        
                        safe_name = current_sheet.replace("/", "_") 
                        f_path = os.path.join(full_ctx_dir, f"{safe_name}.json")
                        
                        st.session_state.crm_sheets[current_sheet].to_json(f_path)
                        st.toast(f"Lead Saved & Persisted to {ctx_path}", icon="✅")
                    except Exception as e:
                        st.error(f"Failed to auto-save: {e}")
                    except Exception as e:
                        print(f"ERROR: Add lead auto-save failed: {e}")
                    
                    # Update Backend (Best Effort)
                    try:
                        backend_payload = {}
                        key_map = {
                            "businessname": "businessName", "company": "businessName",
                            "contactname": "contactName", "name": "contactName",
                            "phone": "phone", "email": "email",
                            "status": "status", "priority": "priority",
                            "notes": "callNotes", "address": "address",
                            "source": "source", "dealvalue": "dealValue"
                        }
                        for k, v in new_row.items():
                            k_norm = k.lower().replace(" ", "").replace("_", "")
                            if k_norm in key_map: backend_payload[key_map[k_norm]] = v
                        if backend_payload:
                             if "source" not in backend_payload: backend_payload["source"] = "Manual"
                             create_lead(backend_payload)
                    except: pass
                    
                    st.toast("Lead Saved Successfully! 🎉", icon="✅")
                    time.sleep(1)
                    st.rerun()

    # --- 3. SHEET MANAGEMENT & DATA FETCH ---
    # --- 3. SHEET MANAGEMENT & DATA FETCH (RBAC Aware) ---
    if "crm_sheets" not in st.session_state or not st.session_state.crm_sheets:
        # Determine context path
        ctx_path = st.session_state.get("current_crm_path", current_user.username)
        full_ctx_dir = os.path.join("crm_data", ctx_path)
        
        if not os.path.exists(full_ctx_dir):
            os.makedirs(full_ctx_dir, exist_ok=True)
            
        # Load files
        loaded_sheets = {}
        try:
            files = [f for f in os.listdir(full_ctx_dir) if f.endswith(".json")]
            for f in files:
                s_name = f.replace(".json", "")
                f_path = os.path.join(full_ctx_dir, f)
                try:
                    df_loaded = pd.read_json(f_path)
                    col_map = {"contactName": "Name", "businessName": "Company Name", "phone": "Phone Number", "email": "Email", "address": "Address", "status": "Status", "priority": "Priority", "nextFollowUpDate": "Next Follow-up Date", "callNotes": "Notes", "lastFollowUpDate": "Last Follow up Date"}
                    df_loaded.rename(columns=col_map, inplace=True)
                    loaded_sheets[s_name] = df_loaded
                except Exception as e:
                    st.error(f"Error loading {s_name}: {e}")
        except Exception as e:
            st.error(f"Error accessing directory: {e}")
            
        if not loaded_sheets:
            # Seed with API data only if it represents 'My CRM' and empty? 
            # Or just create default empty
            # If user is Intern/HR/CEO accessing their OWN CRM, maybe we fetch form API?
            # For simplistic implementation, just fallback to empty Sheet1
            # Try to fetch from backend
            required_cols = ["Name", "Company Name", "Phone Number", "Address", "Last Follow up Date", "Next Follow-up Date", "Status", "Notes", "Priority", "Called By", "Meeting By", "Closed By"]
            try:
                print("DEBUG: Syncing CRM from Backend...")
                api_leads = fetch_data(LEADS_API) # Expects list of dicts
                if api_leads:
                    load_df = pd.DataFrame(api_leads)
                    # Helper to safeguard against missing columns
                    for col in required_cols:
                        if col not in load_df.columns:
                            load_df[col] = None
                    loaded_sheets = {"Sheet1": load_df}
                else:
                    loaded_sheets = {"Sheet1": pd.DataFrame(columns=required_cols)}
            except Exception as e:
                print(f"ERROR: Backend sync failed: {e}")
                loaded_sheets = {"Sheet1": pd.DataFrame(columns=required_cols)}

        st.session_state.crm_sheets = loaded_sheets
        st.session_state.active_sheet = list(loaded_sheets.keys())[0]
    
    # Column Metadata for types (Dropdown options, etc.)
    if "crm_col_meta" not in st.session_state:
        st.session_state.crm_col_meta = {}

    # Ensure active sheet exists

    if st.session_state.active_sheet not in st.session_state.crm_sheets:
        st.session_state.active_sheet = list(st.session_state.crm_sheets.keys())[0]

    # Get Active Dataframe
    df = st.session_state.crm_sheets[st.session_state.active_sheet].copy()
    df = df.loc[:, ~df.columns.duplicated()]
    
    # Ensure columns exist
    required_cols = ["Name", "Company Name", "Phone Number", "Address", "Last Follow up Date", "Next Follow-up Date", "Status", "Notes", "Priority", "Called By", "Meeting By", "Closed By"]
    for col in required_cols:
        if col not in df.columns:
            df[col] = None

    # --- 4. HEADER SECTION (Functional) ---
    # --- 4. HEADER SECTION (Functional) ---
    # Google Sheets Styling for Toolbar
    st.markdown("""
        <style>
        .stApp { background-color: #f8f9fa; }
        
        /* Hide built-in Streamlit header safely */
        [data-testid="stHeader"] { display: none; }
        div[data-testid="stDecoration"] { display: none; }
        
        /* Compact Block Container */
        div[data-testid="stMainBlockContainer"],
        .stAppViewContainer .main .block-container {
            padding-top: 1rem !important; /* Minimal padding */
            padding-bottom: 0rem !important;
            padding-left: 1rem !important;
            padding-right: 1rem !important;
            max-width: 100vw !important;
            margin-top: 0rem !important; 
        }
        
        /* Remove internal gaps */
        div[data-testid="stVerticalBlock"] {
            gap: 0rem !important;
        }
        
        div[data-testid="stHorizontalBlock"] {
            gap: 0.5rem !important;
        }
        
        .main > div {
            padding-top: 0rem !important;
        }
        div[data-testid="stDataEditor"] { 
            border: 1px solid #e0e0e0; 
            border-radius: 0px; 
            background-color: white; 
            box-shadow: none;
        }
        .sheet-tab-active { border-bottom: 3px solid #1a73e8; color: #1a73e8; font-weight: bold; }
        .sheet-tab { color: #5f6368; cursor: pointer; }
        
        
        div[data-testid="stHorizontalBlock"] button {
            border-radius: 20px !important;
            border: 1px solid transparent !important;
            font-size: 14px !important;
            padding: 4px 12px !important;
            font-weight: 500 !important;
            background-color: transparent !important;
            color: #3c4043 !important;
            white-space: nowrap !important;
            min-width: auto !important;
            margin: 0px !important;
        }
        div[data-testid="stHorizontalBlock"] button:hover {
            background-color: #f1f3f4 !important;
            color: #202124 !important;
            border-color: transparent !important;
        }
        
        
        div[data-testid="column"]:has(button:contains("Add Lead")) button {
            background-color: #1a73e8 !important;
            color: white !important;
            border-radius: 24px !important;
            box-shadow: none !important; 
            border: none !important;
            padding: 4px 20px !important;
        }
        div[data-testid="column"]:has(button:contains("Add Lead")) button:hover {
             background-color: #185abc !important;
             color: white !important;
             box-shadow: 0 1px 2px 0 rgba(60,64,67,0.3) !important;
        }
        div[data-testid="column"]:has(button:contains("Add Lead")) button:focus {
             border: none !important;
             outline: none !important;
             box-shadow: none !important;
        }

        

    
    div[data-testid="stButton"] button[kind="primary"] {
        background-color: #1a73e8 !important;
        color: white !important;
        border: none !important;
        border-radius: 24px !important; 
        padding: 4px 24px !important;
        font-weight: 500 !important;
        font-size: 14px !important;
        box-shadow: 0 1px 2px rgba(60,64,67,0.3), 0 1px 3px 1px rgba(60,64,67,0.15) !important;
        transition: background-color 0.2s, box-shadow 0.2s !important;
        height: 36px !important;
        margin-top: 0px !important; 
    }
    div[data-testid="stButton"] button[kind="primary"]:hover {
        background-color: #174ea6 !important;
        box-shadow: 0 1px 3px rgba(60,64,67,0.3), 0 4px 8px 3px rgba(60,64,67,0.15) !important;
    }
    
    
    
    div[data-testid="column"] button:not([kind="primary"]) {
         background-color: white !important;
         border: 1px solid #dadce0 !important;
         border-radius: 8px !important;
         color: #3c4043 !important;
         height: 32px !important;
         font-size: 13px !important;
         font-weight: 500 !important;
         padding: 0 8px !important;
         box-shadow: none !important;
    }
    div[data-testid="column"] button:not([kind="primary"]):hover {
         background-color: #f1f3f4 !important;
         border-color: #dadce0 !important;
         color: #202124 !important;
    }

    
    
    div[data-testid="stPopover"] button {
        font-weight: 500 !important;
        color: #5f6368 !important;
        border: none !important;
        background: transparent !important;
    }
    div[data-testid="stPopover"] button:hover {
        background-color: #f1f3f4 !important;
        color: #202124 !important;
    }

    
    .stTextInput input {
        font-weight: 700 !important;
        font-size: 18px !important;
        border: none !important;
        padding-left: 0 !important;
        background: transparent !important;
        color: #202124 !important;
    }
    .stTextInput input:focus {
        box-shadow: none !important;
        background: white !important;
        border-bottom: 2px solid #1a73e8 !important;
    }
    
    
    hr.crm-header-divider {
        margin: 4px 0 12px 0 !important;
        border: 0 !important;
        border-bottom: 1px solid #e0e0e0 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # --- GOOGLE SHEETS STYLE HEADER & TOOLBAR ---
    
    # Custom CSS for the Spreadsheet Look
    st.markdown("""
        <style>
        /* Main Header Container */
        .gs-header-container {
            padding: 16px 20px;
            border-bottom: 1px solid #E5E7EB;
            background-color: #FFFFFF;
            box-shadow: 0 1px 2px rgba(0,0,0,0.04);
        }

        /* Title Input Styling */
        div[data-testid="stElementContainer"]:has(#crm-header-top-marker) ~ div[data-testid="stHorizontalBlock"] div[data-testid="stTextInput"] div[data-baseweb="input"] {
            background-color: #F8FAFC !important;
            border: 1px solid #E5E7EB !important;
            border-radius: 10px !important;
            padding: 10px 14px !important;
            font-size: 18px !important;
            font-weight: 600 !important;
            color: #1A1C21 !important;
            transition: all 0.2s;
            height: 48px !important;
            min-height: 48px !important;
        }

        /* Select CRM Dropdown */
        div[data-testid="stElementContainer"]:has(#crm-header-top-marker) ~ div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] {
            background-color: #FFFFFF !important;
            border: 1px solid #E5E7EB !important;
            border-radius: 10px !important;
            height: 40px !important;
            min-height: 40px !important;
            font-weight: 500 !important;
            color: #374151 !important;
            transition: all 0.2s ease;
        }
        div[data-testid="stElementContainer"]:has(#crm-header-top-marker) ~ div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"]:hover div[data-baseweb="select"] {
            background-color: #F8FAFC !important;
        }

        /* Add Lead Button */
        div[data-testid="stElementContainer"]:has(#crm-header-top-marker) ~ div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button[kind="primary"] {
            background: linear-gradient(135deg, #2F80ED 0%, #1C64F2 100%) !important;
            color: #FFFFFF !important;
            border-radius: 999px !important;
            padding: 10px 26px !important;
            height: 42px !important;
            font-weight: 600 !important;
            font-size: 14px !important;
            border: none !important;
            box-shadow: 0 4px 12px rgba(47,128,237,0.25) !important;
            transition: all 0.2s ease !important;
        }
        div[data-testid="stElementContainer"]:has(#crm-header-top-marker) ~ div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button[kind="primary"]:hover {
            transform: translateY(-2px) !important;
        }

        /* Menu Bar Row Background (Optional minimal separator) */
        div[data-testid="stElementContainer"]:has(#crm-menu-marker) + div[data-testid="stHorizontalBlock"] {
            padding-bottom: 12px !important;
            gap: 24px !important;
        }

        /* Menu Bar Buttons */
        div[data-testid="stElementContainer"]:has(#crm-menu-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stPopover"] > button {
            border: none !important;
            background: transparent !important;
            padding: 6px 10px !important;
            font-size: 14px !important;
            color: #374151 !important;
            font-weight: 500 !important;
            border-radius: 6px !important;
            transition: all 0.15s ease !important;
        }
        div[data-testid="stElementContainer"]:has(#crm-menu-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stPopover"] > button:hover {
            background-color: #F9FAFB !important;
            text-decoration: underline !important;
            text-decoration-thickness: 2px !important;
            text-underline-offset: 4px !important;
            color: #111827 !important;
        }

        /* Toolbar Row - the whole container */
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] {
            background-color: #FFFFFF !important;
            border: 1px solid #F1F5F9 !important;
            border-radius: 16px !important;
            padding: 14px 16px 14px 20px !important;
            margin: 12px 0 !important;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -1px rgba(0, 0, 0, 0.03) !important;
            align-items: flex-end !important;
            overflow: visible !important;
        }

        /* Filter group label styling */
        .toolbar-filter-label {
            font-family: 'Inter', sans-serif !important;
            font-size: 11px !important;
            font-weight: 700 !important;
            color: #64748b !important;
            text-transform: uppercase !important;
            letter-spacing: 0.8px !important;
            margin-bottom: 6px !important;
            display: block !important;
        }

        /* Toolbar Inputs and Selects */
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stTextInput"] div[data-baseweb="input"],
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"],
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stDateInput"] div[data-baseweb="input"] {
            background-color: #F8FAFC !important;
            border: 1px solid #E2E8F0 !important;
            border-radius: 9999px !important; /* Pill shape */
            font-size: 14px !important;
            height: 40px !important;
            min-height: 40px !important;
            padding: 0 4px !important;
            box-shadow: inset 0 1px 2px rgba(0, 0, 0, 0.02) !important;
            transition: all 0.2s ease !important;
        }
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stTextInput"] input::placeholder {
            color: #94A3B8 !important;
            font-weight: 500 !important;
        }
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stTextInput"]:focus-within div[data-baseweb="input"],
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"]:focus-within div[data-baseweb="select"],
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stDateInput"]:focus-within div[data-baseweb="input"] {
            border-color: #3B82F6 !important;
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.15) !important;
            background-color: #FFFFFF !important;
        }

        /* Toolbar action buttons - icon + text style */
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button {
            color: #475569 !important;
            background: #F8FAFC !important;
            border: 1px solid #E2E8F0 !important;
            border-radius: 9999px !important; /* Pill shape */
            height: 40px !important;
            width: auto !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            padding: 6px 16px !important;
            gap: 6px !important;
            font-size: 13px !important;
            font-weight: 600 !important;
            transition: all 0.2s ease !important;
            white-space: nowrap !important;
            box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05) !important;
        }
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button:hover {
            background-color: #FFFFFF !important;
            color: #0F172A !important;
            border-color: #CBD5E1 !important;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05) !important;
            transform: translateY(-1px) !important;
        }

        /* Toolbar popover button (Quick Settings cog) */
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stPopover"] > button {
            color: #4B5563 !important;
            background: #F3F4F6 !important;
            border: 1px solid #E5E7EB !important;
            border-radius: 8px !important;
            height: 40px !important;
            width: auto !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            padding: 6px 12px !important;
            gap: 4px !important;
            font-size: 12px !important;
            font-weight: 500 !important;
            transition: all 0.2s !important;
        }
        div[data-testid="stElementContainer"]:has(#crm-toolbar-marker) + div[data-testid="stHorizontalBlock"] div[data-testid="stPopover"] > button:hover {
            background-color: #E5E7EB !important;
            color: #111827 !important;
            border-color: #D1D5DB !important;
        }

        /* Toolbar vertical divider */
        .toolbar-divider {
            width: 1px !important;
            height: 32px !important;
            background-color: #E5E7EB !important;
            margin: 0 4px !important;
            display: inline-block !important;
            vertical-align: middle !important;
        }
        
        /* 1) Document Title Input — hide the input box, show plain text */
        div[data-testid="stTextInput"]:has(input[aria-label="Document Title"]) div[data-baseweb="input"] {
            background-color: transparent !important;
            border: 1px solid transparent !important;
            border-radius: 4px !important;
            box-shadow: none !important;
            transition: all 0.2s ease !important;
        }
        div[data-testid="stTextInput"]:has(input[aria-label="Document Title"]) div[data-baseweb="input"]:hover {
            border: 1px solid #dadce0 !important;
        }
        div[data-testid="stTextInput"]:has(input[aria-label="Document Title"]) div[data-baseweb="input"]:focus-within {
            border: 2px solid #4F46E5 !important;
            background-color: #ffffff !important;
        }
        div[data-testid="stTextInput"]:has(input[aria-label="Document Title"]) input {
            font-size: 15px !important;
            font-weight: 700 !important;
            color: #111827 !important;
            font-family: 'DM Sans', sans-serif !important;
            padding: 2px 4px !important;
            height: auto !important;
            line-height: 1.2 !important;
            background: transparent !important;
            outline: none !important;
        }

        /* 2) Global Primary Button — indigo pill */
        div[data-testid="stButton"] button[kind="primary"],
        button[kind="primary"] {
            background: #4F46E5 !important;
            border: none !important;
            border-radius: 8px !important;
            box-shadow: 0 2px 10px rgba(79,70,229,0.25) !important;
            color: #ffffff !important;
            font-weight: 600 !important;
            font-size: 13px !important;
            font-family: 'DM Sans', sans-serif !important;
            transition: all 0.15s ease !important;
            padding: 0 14px !important;
            height: 36px !important;
            white-space: nowrap !important;
        }
        div[data-testid="stButton"] button[kind="primary"]:hover,
        button[kind="primary"]:hover {
            background: #4338CA !important;
            box-shadow: 0 4px 16px rgba(79,70,229,0.35) !important;
        }

        /* 3) Header menu popovers — borderless text style */
        div[data-testid="stPopover"] > div > button {
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            color: #374151 !important;
            font-size: 13.5px !important;
            font-weight: 500 !important;
            padding: 2px 6px !important;
            height: auto !important;
            min-height: unset !important;
            border-radius: 4px !important;
            white-space: nowrap !important;
            font-family: 'DM Sans', sans-serif !important;
        }
        div[data-testid="stPopover"] > div > button:hover {
            background: #F3F4F6 !important;
            color: #111827 !important;
        }

        /* 4) Header icon buttons — tiny borderless */
        div[class*="crm-icon-btn"] button,
        div[data-testid="stButton"].crm-hdr-icon button {
            background: transparent !important;
            border: 1.5px solid #E5E7EB !important;
            border-radius: 7px !important;
            box-shadow: none !important;
            color: #6B7280 !important;
            font-size: 14px !important;
            padding: 0 !important;
            width: 32px !important;
            height: 32px !important;
            min-height: unset !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
        }
        div[data-testid="stButton"].crm-hdr-icon button:hover {
            background: #F9FAFB !important;
            border-color: #D1D5DB !important;
            color: #374151 !important;
        }

        /* 5) CRM Pill Selector via popover */
        .crm-pill-popover-btn > div > button {
            background: #EEF2FF !important;
            border: 1.5px solid #C7D2FE !important;
            border-radius: 20px !important;
            color: #4338CA !important;
            font-size: 12.5px !important;
            font-weight: 600 !important;
            padding: 3px 12px !important;
            height: 30px !important;
            min-height: unset !important;
            white-space: nowrap !important;
            font-family: 'DM Sans', sans-serif !important;
        }
        .crm-pill-popover-btn > div > button:hover {
            background: #E0E7FF !important;
            border-color: #A5B4FC !important;
        }

        /* 6) Header row global — white bg, fixed height, border-bottom */
        div.stVerticalBlock:has(span#crm-header-top-marker) {
            background: #FFFFFF !important;
            border-bottom: 1px solid #E5E7EB !important;
            padding: 6px 12px 6px 12px !important;
        }
        div.stVerticalBlock:has(span#crm-header-top-marker) div[data-testid="stHorizontalBlock"] {
            align-items: center !important;
            flex-wrap: nowrap !important;
            gap: 0 !important;
            min-height: 44px !important;
        }

        /* 7) Top Gradient Bar */
        .top-gradient-bar {
            position: fixed;
            top: 0; left: 0; right: 0;
            height: 3px;
            background: linear-gradient(90deg, #4F46E5, #818CF8, #C084FC);
            z-index: 999999;
        }

        /* 8) Font */
        @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap');
        .stApp, .stApp > header, .stApp .main, button, input, select {
            font-family: 'DM Sans', sans-serif !important;
        }
        </style>
    """, unsafe_allow_html=True)

    # --- FILE, EDIT, VIEW MENU ---
    # We bring the entire layout into one unified line as requested.
    st.markdown('<div class="top-gradient-bar"></div>', unsafe_allow_html=True)
    with st.container():
        st.markdown('<span id="crm-header-top-marker"></span>', unsafe_allow_html=True)
        # Determine Active CRM Path early for Title and Selector
        current_path_val = st.session_state.get("current_crm_path", current_user.username)
        # Calculate Dynamic Title & Avatar
        tgt_u = None
        user_initial = "U"
        doc_avatar_b64 = None
        
        if "shd_pixel" in current_path_val:
            doc_title_val = "SHDPIXEL CRM"
            user_initial = "S"
            logo_path = "logo.png" if os.path.exists("logo.png") else "logo_light.png"
            if os.path.exists(logo_path):
                import base64
                with open(logo_path, "rb") as f:
                    doc_avatar_b64 = base64.b64encode(f.read()).decode()
        elif current_path_val == current_user.username:
            doc_title_val = f"{current_user.name} CRM"
            tgt_u = current_user
        else:
            tgt_u = auth_manager.users.get(current_path_val)
            doc_title_val = f"{tgt_u.name} CRM" if tgt_u else "CRM Database"

        if tgt_u:
            user_initial = tgt_u.name[0].upper() if tgt_u.name else "U"
            avatar_path = getattr(tgt_u, "avatar_file", None)
            if avatar_path and os.path.exists(avatar_path):
                import base64
                with open(avatar_path, "rb") as f:
                    doc_avatar_b64 = base64.b64encode(f.read()).decode()

        # --- STATE INITIALIZATION FOR FILE MENUS ---
        if "undo_stack" not in st.session_state: st.session_state.undo_stack = []
        
        # Helper: Save State for Undo Logic
        def save_state_for_undo():
            c_sheet = st.session_state.active_sheet
            current_data = st.session_state.crm_sheets[c_sheet].copy()
            st.session_state.undo_stack.append({'sheet': c_sheet, 'data': current_data})
            if len(st.session_state.undo_stack) > 20: st.session_state.undo_stack.pop(0)

        # Helper: Save Active Sheet to Disk immediately after a menu action
        def save_active_to_disk(sheet_name=None):
            try:
                target_sheet = sheet_name or st.session_state.active_sheet
                ctx_path = st.session_state.get("current_crm_path", current_user.username)
                full_ctx_dir = os.path.join("crm_data", ctx_path)
                if not os.path.exists(full_ctx_dir): os.makedirs(full_ctx_dir, exist_ok=True)
                safe_name = target_sheet.replace("/", "_")
                f_path = os.path.join(full_ctx_dir, f"{safe_name}.json")
                st.session_state.crm_sheets[target_sheet].to_json(f_path)
            except Exception as e:
                print(f"Error saving to disk manually: {e}")

        # Helper: Delete Sheet from Disk
        def delete_sheet_from_disk(sheet_name):
            try:
                ctx_path = st.session_state.get("current_crm_path", current_user.username)
                full_ctx_dir = os.path.join("crm_data", ctx_path)
                safe_name = sheet_name.replace("/", "_")
                f_path = os.path.join(full_ctx_dir, f"{safe_name}.json")
                if os.path.exists(f_path): os.remove(f_path)
            except Exception:
                pass

        current_sheet_name = st.session_state.active_sheet
        df_current = st.session_state.crm_sheets[current_sheet_name]

        # ---------------------------------------------------------------
        # SINGLE UNIFIED HEADER ROW
        # Strategy: col[0] = HTML (avatar + title + pill + divider)
        #           col[1..6] = menu popovers (File/Edit/View/Insert/AddCol/More)
        #           col[7] = Add Lead button
        # CSS below makes popovers look like plain-text menu items
        # ---------------------------------------------------------------

        accessible_crms = auth_manager.get_accessible_crms(current_user)
        sel_index_crm = 0
        for idx, opt in enumerate(accessible_crms):
            if auth_manager.resolve_crm_path(opt, current_user) == current_path_val:
                sel_index_crm = idx
                break
        _raw_label = accessible_crms[sel_index_crm] if accessible_crms else "My CRM"
        # Clean up the CRM label for display (remove 'User: ' prefix, truncate long names)
        _clean = _raw_label.replace("User: ", "").replace("user: ", "")
        # Strip email domain for readability
        if "@" in _clean: _clean = _clean.split("@")[0]
        # Capitalize nicely
        current_crm_label = _clean[:14].strip() or "My CRM"

        if doc_avatar_b64:
            avatar_html = (
                f'<img src="data:image/png;base64,{doc_avatar_b64}" '
                f'style="width:32px;height:32px;border-radius:8px;object-fit:cover;'
                f'border:1px solid #C7D2FE;flex-shrink:0;">'
            )
        else:
            avatar_html = (
                f'<span style="display:inline-flex;width:32px;height:32px;border-radius:10px;'
                f'background:#4F46E5;color:white;align-items:center;justify-content:center;'
                f'font-weight:800;font-size:15px;flex-shrink:0;letter-spacing:-0.5px;">{user_initial}</span>'
            )

        # Col 0 HTML: avatar + title + pill + divider — matches reference exactly
        left_cluster = (
            '<div style="display:flex;align-items:center;gap:10px;height:48px;padding:0 4px;">'
            + avatar_html
            + f'<strong style="font-size:14px;color:#111827;font-weight:700;'
            f'letter-spacing:-0.2px;white-space:nowrap;">{doc_title_val}</strong>'
            + '<span style="display:inline-block;width:1px;height:20px;background:#D1D5DB;"></span>'
            + f'<span style="display:inline-flex;align-items:center;gap:5px;'
            f'background:#F4F3FF;border:1px solid #C7D2FE;border-radius:20px;'
            f'padding:2px 10px;font-size:12.5px;font-weight:600;color:#4338CA;'
            f'white-space:nowrap;cursor:pointer;">&#9679; {current_crm_label} &#9660;</span>'
            + '<span style="display:inline-block;width:1px;height:20px;background:#D1D5DB;"></span>'
            + '</div>'
        )

        # CSS: scope everything to this marker, make popovers look like plain text
        st.markdown("""
        <style>
        /* ── Header row container ── */
        div.stVerticalBlock:has(span#crm-unified-row) {
            background: #FFFFFF !important;
            border-bottom: 1.5px solid #E5E7EB !important;
            padding: 0 10px !important;
            margin-bottom: 0 !important;
        }
        div.stVerticalBlock:has(span#crm-unified-row) > div[data-testid="stHorizontalBlock"] {
            align-items: center !important;
            flex-wrap: nowrap !important;
            min-height: 52px !important;
            gap: 0 !important;
        }
        /* Columns: center all content vertically */
        div.stVerticalBlock:has(span#crm-unified-row) div[data-testid="column"] {
            display: flex !important;
            align-items: center !important;
            padding: 0 1px !important;
        }
        /* Popover trigger buttons: plain borderless text */
        div.stVerticalBlock:has(span#crm-unified-row) div[data-testid="stPopover"] > div > button {
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            color: #4B5563 !important;
            font-size: 13px !important;
            font-weight: 500 !important;
            padding: 0 8px !important;
            height: 52px !important;
            min-height: unset !important;
            border-radius: 0px !important;
            white-space: nowrap !important;
            border-bottom: 2px solid transparent !important;
            transition: color 0.15s, border-color 0.15s !important;
        }
        div.stVerticalBlock:has(span#crm-unified-row) div[data-testid="stPopover"] > div > button:hover {
            color: #4F46E5 !important;
            border-bottom-color: #4F46E5 !important;
            background: transparent !important;
        }
        div.stVerticalBlock:has(span#crm-unified-row) div[data-testid="stPopover"] > div > button svg {
            display: none !important;
        }
        /* ── Add Lead primary button — MUST stay purple ── */
        div.stVerticalBlock:has(span#crm-unified-row) div[data-testid="stButton"] > button {
            background: #4F46E5 !important;
            color: #FFFFFF !important;
            border: none !important;
            height: 36px !important;
            padding: 0 18px !important;
            font-size: 13px !important;
            font-weight: 600 !important;
            border-radius: 8px !important;
            white-space: nowrap !important;
            box-shadow: 0 1px 4px rgba(79,70,229,0.3) !important;
            cursor: pointer !important;
        }
        div.stVerticalBlock:has(span#crm-unified-row) div[data-testid="stButton"] > button:hover {
            background: #4338CA !important;
        }
        </style>
        """, unsafe_allow_html=True)


        st.markdown('<span id="crm-unified-row"></span>', unsafe_allow_html=True)

        # 9 columns: [branding] [File] [Edit] [View] [Insert] [AddCol] [More] [icons] [AddLead]
        hc = st.columns([2.8, 0.5, 0.5, 0.5, 0.6, 0.7, 0.72, 0.9, 1.5], gap="small", vertical_alignment="center")


        with hc[0]:
            st.markdown(left_cluster, unsafe_allow_html=True)

        with hc[1]:
            with st.popover("File", use_container_width=True):
                st.caption("Sheet Actions")
                if st.button("New Sheet 📄", use_container_width=True, key="file_new_sheet"):
                    new_n = f"Sheet{len(st.session_state.crm_sheets)+1}"
                    if new_n not in st.session_state.crm_sheets:
                        st.session_state.crm_sheets[new_n] = pd.DataFrame(columns=required_cols)
                        st.session_state.active_sheet = new_n
                        save_active_to_disk()
                        st.rerun()
                if st.button("Duplicate Sheet 📑", use_container_width=True, key="file_dup_sheet"):
                    save_state_for_undo()
                    dup_name = f"Copy of {current_sheet_name}"
                    st.session_state.crm_sheets[dup_name] = df_current.copy()
                    st.session_state.active_sheet = dup_name
                    save_active_to_disk()
                    st.rerun()
                st.divider()
                ren_name = st.text_input("Rename Sheet", value=current_sheet_name, key="menu_ren_input")
                if st.button("Apply Rename", key="menu_ren_btn"):
                    if ren_name and ren_name != current_sheet_name:
                        if ren_name not in st.session_state.crm_sheets:
                            st.session_state.crm_sheets[ren_name] = st.session_state.crm_sheets.pop(current_sheet_name)
                            st.session_state.active_sheet = ren_name
                            delete_sheet_from_disk(current_sheet_name)
                            save_active_to_disk(ren_name)
                            st.rerun()
                        else: st.error("Name exists")

        with hc[2]:
            with st.popover("Edit", use_container_width=True):
                can_undo = len(st.session_state.undo_stack) > 0
                if st.button(f"Undo ↩️ ({len(st.session_state.undo_stack)})", disabled=not can_undo, use_container_width=True, key="edit_undo_btn"):
                    last_state = st.session_state.undo_stack.pop()
                    st.session_state.crm_sheets[last_state['sheet']] = last_state['data']
                    st.session_state.active_sheet = last_state['sheet']
                    save_active_to_disk()
                    st.rerun()
                st.divider()
                find_s = st.text_input("Find", key="edit_find")
                repl_s = st.text_input("Replace", key="edit_repl")
                if st.button("Replace All", key="edit_repl_btn"):
                    if find_s:
                        save_state_for_undo()
                        st.session_state.crm_sheets[current_sheet_name] = df_current.replace(find_s, repl_s, regex=False)
                        save_active_to_disk()
                        st.rerun()

        with hc[3]:
            with st.popover("View", use_container_width=True):
                compact_val = st.toggle("Compact Mode", value=st.session_state.get("view_compact", False), key="view_compact_tog")
                if compact_val != st.session_state.get("view_compact", False):
                    st.session_state.view_compact = compact_val
                    st.rerun()

        with hc[4]:
            with st.popover("Insert", use_container_width=True):
                if st.button("Row Above ⬆️", use_container_width=True, key="ins_row_above"):
                    save_state_for_undo()
                    empty = pd.DataFrame([{c: None for c in df_current.columns}])
                    st.session_state.crm_sheets[current_sheet_name] = pd.concat([empty, df_current], ignore_index=True)
                    save_active_to_disk()
                    st.rerun()
                if st.button("Row Below ⬇️", use_container_width=True, key="ins_row_below"):
                    save_state_for_undo()
                    empty = pd.DataFrame([{c: None for c in df_current.columns}])
                    st.session_state.crm_sheets[current_sheet_name] = pd.concat([df_current, empty], ignore_index=True)
                    save_active_to_disk()
                    st.rerun()

        with hc[5]:
            with st.popover("Add Col", use_container_width=True):
                ac_name = st.text_input("Column Name", key="ac_name_in")
                ac_type = st.selectbox("Type", ["Text", "Number", "Date", "Dropdown"], key="ac_type_in")
                if st.button("Create Column", type="primary", use_container_width=True, key="ac_create_btn"):
                    if ac_name and ac_name not in df_current.columns:
                        save_state_for_undo()
                        st.session_state.crm_sheets[current_sheet_name][ac_name] = None
                        save_active_to_disk()
                        st.rerun()

        with hc[6]:
            with st.popover("⋯ More", use_container_width=True):
                st.caption("Switch CRM")
                for opt in accessible_crms:
                    resolved = auth_manager.resolve_crm_path(opt, current_user)
                    is_active = (resolved == current_path_val)
                    lbl = f"✓  {opt}" if is_active else f"    {opt}"
                    if st.button(lbl, key=f"crm_switch_{opt}", use_container_width=True):
                        if resolved != current_path_val:
                            st.session_state.current_crm_path = resolved
                            st.session_state.crm_sheets = {}
                            st.rerun()
                st.divider()
                st.caption("Tools")
                if st.button("Remove Duplicates 🧹", use_container_width=True, key="more_dedup"):
                    save_state_for_undo()
                    st.session_state.crm_sheets[current_sheet_name].drop_duplicates(inplace=True)
                    save_active_to_disk()
                    st.rerun()
                if st.button("Delete Empty Rows", use_container_width=True, key="more_del_empty"):
                    save_state_for_undo()
                    st.session_state.crm_sheets[current_sheet_name].dropna(how='all', inplace=True)
                    save_active_to_disk()
                    st.rerun()
                del_target = st.selectbox("Delete Column:", df_current.columns, key="del_col_sel")
                if st.button(f"Delete '{del_target}'", type="primary", use_container_width=True, key="more_del_col"):
                    if len(df_current.columns) > 1:
                        st.session_state.crm_sheets[current_sheet_name].drop(columns=[del_target], inplace=True)
                        save_active_to_disk()
                        st.rerun()
                st.divider()
                csv_data = df_current.to_csv(index=False).encode('utf-8')
                st.download_button("Export CSV ⬇️", csv_data, f"{current_sheet_name}.csv", "text/csv", use_container_width=True, key="more_export")
                if st.button("⚠️ CLEAR ALL DATA", use_container_width=True, key="more_clear_all"):
                    save_state_for_undo()
                    st.session_state.crm_sheets[current_sheet_name] = pd.DataFrame(columns=required_cols)
                    save_active_to_disk()
                    st.rerun()

        with hc[7]:
            # Icon buttons as HTML (decorative — non-interactive)
            icons_html = (
                '<div style="display:flex;align-items:center;gap:5px;justify-content:flex-end;">'
                '<button style="width:28px;height:28px;border:1.5px solid #E2E8F0;border-radius:6px;'
                'background:#fff;cursor:pointer;font-size:13px;display:inline-flex;'
                'align-items:center;justify-content:center;color:#6B7280;">&#128065;</button>'
                '<button style="width:28px;height:28px;border:1.5px solid #E2E8F0;border-radius:6px;'
                'background:#fff;cursor:pointer;font-size:13px;display:inline-flex;'
                'align-items:center;justify-content:center;color:#6B7280;">&#8595;</button>'
                '<button style="width:28px;height:28px;border:1.5px solid #E2E8F0;border-radius:6px;'
                'background:#fff;cursor:pointer;font-size:13px;display:inline-flex;'
                'align-items:center;justify-content:center;color:#6B7280;">&#8599;</button>'
                '</div>'
            )
            st.markdown(icons_html, unsafe_allow_html=True)

        with hc[8]:
            if st.button("+ Add Lead", type="primary", use_container_width=True, key="header_add_lead_right"):
                add_lead_modal()

    # --- END HEADER BLOCK ---

    # --- TOOLBAR SECTION ---
    st.markdown("""
    <style>
    div.stVerticalBlock:has(span#crm-toolbar-marker) {
        background: #FAFAFA !important;
        border-bottom: 1px solid #E5E7EB !important;
        padding: 6px 12px !important;
    }
    div.stVerticalBlock:has(span#crm-toolbar-marker) div[data-testid="stHorizontalBlock"] {
        align-items: center !important;
        flex-wrap: nowrap !important;
        gap: 8px !important;
    }
    /* Toolbar search input */
    div.stVerticalBlock:has(span#crm-toolbar-marker) div[data-testid="stTextInput"] input {
        border-radius: 20px !important;
        border: 1.5px solid #E2E8F0 !important;
        background: #fff !important;
        font-size: 13px !important;
        padding: 4px 14px !important;
        height: 32px !important;
    }
    /* Toolbar selectboxes */
    div.stVerticalBlock:has(span#crm-toolbar-marker) div[data-testid="stSelectbox"] > div {
        border-radius: 6px !important;
        border: 1.5px solid #E2E8F0 !important;
        background: #fff !important;
        min-height: 32px !important;
        font-size: 13px !important;
    }
    /* Toolbar filter labels */
    .tb-label {
        font-size: 11px;
        font-weight: 700;
        color: #6B7280;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        white-space: nowrap;
        display: block;
        margin-bottom: 2px;
    }
    /* Refresh & action buttons in toolbar */
    div.stVerticalBlock:has(span#crm-toolbar-marker) div[data-testid="stButton"] button {
        border: 1.5px solid #E2E8F0 !important;
        border-radius: 6px !important;
        background: #fff !important;
        font-size: 13px !important;
        color: #374151 !important;
        height: 32px !important;
        padding: 0 10px !important;
    }
    </style>
    """, unsafe_allow_html=True)
    st.markdown('<span id="crm-toolbar-marker"></span>', unsafe_allow_html=True)
    inst_id = f"{current_path_val}_{current_sheet_name}"
    t_c1, t_c2, t_c3, t_c4, t_c5, t_c6 = st.columns([2.5, 1.3, 1.3, 1.5, 0.1, 2.8], gap="small", vertical_alignment="bottom")

    with t_c1:
        st.markdown('<span class="tb-label">&#128269; Search lead</span>', unsafe_allow_html=True)
        q = st.text_input("Search", placeholder="Search lead", label_visibility="collapsed", key=f"gs_search_{inst_id}")

    with t_c2:
        st.markdown('<span class="tb-label">&#128202; STATUS</span>', unsafe_allow_html=True)
        all_stats = ["All"] + list(df['Status'].unique()) if 'Status' in df.columns else ["All"]
        s_fil = st.selectbox("Status", options=all_stats, label_visibility="collapsed", key=f"gs_status_{inst_id}", index=0)

    with t_c3:
        st.markdown('<span class="tb-label">&#128293; PRIORITY</span>', unsafe_allow_html=True)
        p_fil = st.selectbox("Priority", options=["All", "HOT", "WARM", "COLD"], label_visibility="collapsed", key=f"gs_prio_{inst_id}", index=0)

    with t_c4:
        # Date — with visible label
        st.markdown('<span class="toolbar-filter-label">📅 Follow-up Date</span>', unsafe_allow_html=True)
        d_fil = st.date_input("Filter Date", value=None, label_visibility="collapsed", key=f"gs_date_{inst_id}")

    with t_c5:
        # Visual divider between filters and actions
        st.markdown('<div class="toolbar-divider"></div>', unsafe_allow_html=True)

    with t_c6:
        # Action buttons with text labels for clarity
        ic1, ic2, ic3 = st.columns(3)
        with ic1:
             st.button("🔄 Refresh", help="Refresh the current view", key=f"tb_refresh_{inst_id}", use_container_width=True, on_click=lambda: st.rerun())
        with ic2:
             st.button("⊟", help="Toggle sidebar / view", key=f"tb_sidebar_icon_{inst_id}", use_container_width=True)
        with ic3:
             st.button("🔍", help="Expanded search", key=f"tb_search_icon_{inst_id}", use_container_width=True)
             
    st.markdown("<div style='margin-bottom: 8px'></div>", unsafe_allow_html=True)

    # Apply Optional CSS for Alternating Colors & Gridlines
    grid_css = ""
    if not st.session_state.get(f"gs_gridlines_{inst_id}", True): 
        grid_css += 'div[data-testid="stDataEditor"] td { border: none !important; } '
    if st.session_state.get(f"gs_altcolors_{inst_id}", False): 
        grid_css += 'div[data-testid="stDataEditor"] tr:nth-child(even) { background-color: #f8fafc !important; } '

    # --- 5. GOOGLE SHEETS STYLE GRID ---
    st.markdown(f"""
    <style>
    div[data-testid="stDataEditor"] th {{ 
        background-color: #f8f9fa !important; color: #3c4043 !important; font-weight: 600 !important;
        border-bottom: 2px solid #dadce0 !important;
    }}
    div[data-testid="stDataEditor"] td {{ border: 1px solid #e0e0e0 !important; }}
    {grid_css}
    </style>
    """, unsafe_allow_html=True)

    # 1. Filter Row
    c1, c2, c3, c4, c5 = st.columns([2, 1.2, 1.2, 1.2, 1.5], vertical_alignment="center")
    
    # Filter Logic (Applying the filters from Row 2)
    # Note: Variable names (q, s_fil, p_fil, d_fil) must match what was defined above.
    pass

    # Filter Logic
    df_filtered = df.copy()
    if q: df_filtered = df_filtered[df_filtered.astype(str).apply(lambda x: x.str.contains(q, case=False, na=False)).any(axis=1)]
    if s_fil != "All" and "Status" in df_filtered.columns: df_filtered = df_filtered[df_filtered['Status'] == s_fil]
    if p_fil != "All" and "Priority" in df_filtered.columns: df_filtered = df_filtered[df_filtered['Priority'] == p_fil]
    
    # Hide ID Column (User request: "not have id")
    cols_to_drop = [c for c in df_filtered.columns if c.lower() == "id"]
    if cols_to_drop:
        df_filtered = df_filtered.drop(columns=cols_to_drop)

    # Ensure dates are datetime objects for Editor compatibility
    date_columns = []
    for col in df_filtered.columns:
        if "date" in col.lower() or col in ["created_at", "nextFollowUpDate", "lastFollowUpDate"]:
            date_columns.append(col)
            try:
                df_filtered[col] = pd.to_datetime(df_filtered[col], errors='coerce').dt.date
            except Exception:
                pass
        
        # FIX: Ensure map_url is string to avoid float incompatibility with LinkColumn
        if col.lower() == "map_url":
             df_filtered[col] = df_filtered[col].astype(str).replace("nan", "")

    # Clean None values to show as empty cells in the Grid
    for col in df_filtered.columns:
        is_bool = pd.api.types.is_bool_dtype(df_filtered[col]) or pd.api.types.infer_dtype(df_filtered[col]).startswith("bool")
        is_num = pd.api.types.is_numeric_dtype(df_filtered[col])
        if col not in date_columns and not is_bool and not is_num:
            if pd.api.types.is_object_dtype(df_filtered[col]):
                df_filtered[col] = df_filtered[col].fillna("").astype(str).replace(["None", "nan", "NaN"], "")

    # Pad with empty rows to simulate a large spreadsheet grid (Google Sheets style)
    target_rows = 100
    if len(df_filtered) < target_rows:
        pad_len = target_rows - len(df_filtered)
        empty_data = {}
        for c in df_filtered.columns:
            is_bool = pd.api.types.is_bool_dtype(df_filtered[c]) or pd.api.types.infer_dtype(df_filtered[c]).startswith("bool")
            is_num = pd.api.types.is_numeric_dtype(df_filtered[c])
            if c in date_columns or is_num:
                empty_data[c] = [None] * pad_len
            elif is_bool:
                empty_data[c] = [False] * pad_len
            else:
                empty_data[c] = [""] * pad_len
                
        pad_df = pd.DataFrame(empty_data)
        # Ensure index continues properly
        start_idx = max(df_filtered.index.max() + 1, len(df_filtered)) if not df_filtered.empty else 0
        pad_df.index = range(start_idx, start_idx + pad_len)
        df_filtered = pd.concat([df_filtered, pad_df])

    # --- GRID LOGIC & STYLER ---
    status_options = ["Interested", "Not picking", "Asked to call later", "Meeting set", "Meeting Done", "Proposal sent", "Follow-up scheduled", "Not interested", "Closed - Won", "Closed - Lost"]
    priority_options = ["HOT", "WARM", "COLD"]
    user_options = [""] + [u.name for u in auth_manager.users.values() if u.role != Role.CEO]

    col_cfg = {}
    for col in df_filtered.columns:
        if "date" in col.lower() or col in ["created_at"]:
            col_cfg[col] = st.column_config.DateColumn(col, format="DD/MM/YYYY")
        elif col == "Status":
             col_cfg[col] = st.column_config.SelectboxColumn(col, options=status_options)
        elif col == "Priority":
             col_cfg[col] = st.column_config.SelectboxColumn(col, options=priority_options)
        elif col in ["Called By", "Meeting By", "Closed By", "calledBy", "meetingBy", "closedBy"]:
             col_cfg[col] = st.column_config.SelectboxColumn(col, options=user_options)
        elif col.lower() == "map_url" or col.lower() == "map":
             col_cfg[col] = st.column_config.LinkColumn(col, display_text="Open 🗺️")
        elif pd.api.types.is_bool_dtype(df_filtered[col]) or pd.api.types.infer_dtype(df_filtered[col]).startswith("bool"):
             col_cfg[col] = st.column_config.CheckboxColumn(col)
        elif pd.api.types.is_numeric_dtype(df_filtered[col]):
             col_cfg[col] = st.column_config.NumberColumn(col)
        else:
             col_cfg[col] = st.column_config.TextColumn(col)

    # Force minimum widths 
    if "Name" in col_cfg: col_cfg["Name"] = st.column_config.TextColumn("Name", width="medium")
    if "Company Name" in col_cfg: col_cfg["Company Name"] = st.column_config.TextColumn("Company Name", width="medium")
    if "Phone Number" in col_cfg: col_cfg["Phone Number"] = st.column_config.TextColumn("Phone Number", width="medium")
    if "Address" in col_cfg: col_cfg["Address"] = st.column_config.TextColumn("Address", width="large")
    if "Notes" in col_cfg: col_cfg["Notes"] = st.column_config.TextColumn("Notes", width="large")
             
    # Add '➕' Fake Column for creation
    df_filtered["➕"] = None
    col_cfg["➕"] = st.column_config.TextColumn("Add Col", width="small", help="Type here to add a new column")

    def crm_style(val, col):
        if not val or pd.isna(val): return ""
        val_str = str(val).strip()

        status_colors = {
            "Interested": ("#D6E9FF", "#1A73E8"),
            "Not picking": ("#FF00B8", "#FFFFFF"),
            "Asked to call later": ("#FFD9B3", "#8A4B00"),
            "Meeting set": ("#E8D6FF", "#5E35B1"),
            "Meeting Done": ("#D1C4E9", "#4527A0"),
            "Proposal sent": ("#E2F0D9", "#2E7D32"),
            "Follow-up scheduled": ("#FFE082", "#8A6D00"),
            "Not interested": ("#D7CCC8", "#5D4037"),
            "Closed - Won": ("#C8E6C9", "#1B5E20"),
            "Closed - Lost": ("#FFCDD2", "#B71C1C")
        }
        
        prio_colors = {
            "HOT": ("#F4B183", "#7A2E00"),
            "WARM": ("#FFE599", "#7A5B00"),
            "COLD": ("#9FC5E8", "#0B5394")
        }

        user_colors = [
            ("#E1BEE7", "#4A148C"), ("#C8E6C9", "#1B5E20"), 
            ("#BBDEFB", "#0D47A1"), ("#FFCCBC", "#BF360C"), ("#B2DFDB", "#004D40")
        ]

        # Status
        if col == "Status" and val_str in status_colors:
            bg, fg = status_colors[val_str]
            return f"background-color: {bg} !important; color: {fg} !important; font-weight: 600;"
        
        # Priority
        if col == "Priority" and val_str in prio_colors:
            bg, fg = prio_colors[val_str]
            return f"background-color: {bg} !important; color: {fg} !important; font-weight: 600;"
        
        # Users
        if col in ["Called By", "Meeting By", "Closed By", "calledBy", "meetingBy", "closedBy"] and val_str in user_options and val_str:
            h = sum(ord(c) for c in val_str) % len(user_colors)
            bg, fg = user_colors[h]
            return f"background-color: {bg} !important; color: {fg} !important; font-weight: 600;"

        return ""

    # Row Numbering (Start from 1)
    # Insert visual "No." column at the start
    df_filtered.insert(0, "No.", range(1, len(df_filtered) + 1))
    col_cfg["No."] = st.column_config.NumberColumn("No.", width="small", disabled=True)

    if not df_filtered.empty:
        styled_df = df_filtered.style.apply(lambda x: [crm_style(x[c], c) for c in x.index], axis=1).hide(axis="index")
    else:
        styled_df = df_filtered

    # Grid Height
    # Default to a larger height for 'End to End' experience
    default_h = 750 
    grid_h = 900 if st.session_state.get("view_full_height", False) else default_h
    # SIMPLIFIED KEY FOR STABILITY
    # editor_key = f"editor_{crm_ctx}_{st.session_state.active_sheet}"
    # Using a static key allows Streamlit to handle data updates more gracefully without unmounting
    editor_key = "universal_data_editor" 
    
    # Hide index explicitly in column config to fix Styler index visibility issue
    col_cfg["_index"] = None

    edited_df = st.data_editor(
        styled_df,
        column_config=col_cfg,
        width="stretch",
        height=grid_h,
        num_rows="fixed",
        key=editor_key,
        hide_index=True
    )
    
    # --- EVENTS ---
    # 1. New Column Creation (via '➕' column)
    if "➕" in edited_df.columns:
        added_cols = edited_df[edited_df["➕"].notna() & (edited_df["➕"] != "")]
        if not added_cols.empty:
            new_name = str(added_cols["➕"].iloc[0])
            if new_name and new_name not in st.session_state.crm_sheets[st.session_state.active_sheet].columns:
                st.session_state.crm_sheets[st.session_state.active_sheet][new_name] = ""
                st.success(f"Column '{new_name}' added!")
                st.rerun()

    # 2. Auto Save
    # Drop "➕" AND the new "No." column before saving
    edited_clean = edited_df.drop(columns=["➕", "No."], errors='ignore')
    # Filter original to compare (align columns)
    df_filtered_clean = df_filtered.drop(columns=["➕", "No."], errors='ignore')
    
    # Detect changes
    # Detect changes & Update Session State
    if not edited_clean.equals(df_filtered_clean):
         st.toast("DEBUG: Change Detected!")
         print("DEBUG: Change Detected in Data Editor")
         try:
             # Update existing rows (aligns by index)
             st.session_state.crm_sheets[st.session_state.active_sheet].update(edited_clean)
             
             # Handle Added Rows (Indices not in current)
             current_idx = st.session_state.crm_sheets[st.session_state.active_sheet].index
             # Find rows in edited_clean that are not in original index
             # Note: Streamlit's data_editor adds numeric indices for new rows usually.
             new_rows = edited_clean[~edited_clean.index.isin(current_idx)]
             
             if not new_rows.empty:
                 # Filter out padding rows that were not actually edited (still perfectly blank)
                 new_rows_check = new_rows.copy()
                 for c in new_rows_check.columns:
                     if pd.api.types.is_object_dtype(new_rows_check[c]):
                         new_rows_check[c] = new_rows_check[c].replace(["", "None", "nan"], pd.NA)
                 # Drop if all columns are NA
                 valid_new_rows = new_rows.loc[new_rows_check.dropna(how="all").index]
                 
                 if not valid_new_rows.empty:
                     # Append new rows
                     st.session_state.crm_sheets[st.session_state.active_sheet] = pd.concat(
                         [st.session_state.crm_sheets[st.session_state.active_sheet], valid_new_rows]
                     )
             
             
             # --- AUTO SAVE TO DISK ---
             ctx_path = st.session_state.get("current_crm_path", current_user.username)
             full_ctx_dir = os.path.join("crm_data", ctx_path)
             if not os.path.exists(full_ctx_dir): os.makedirs(full_ctx_dir, exist_ok=True)
             
             safe_name = st.session_state.active_sheet.replace("/", "_")
             f_path = os.path.join(full_ctx_dir, f"{safe_name}.json")
             
             # DEBUGGING
             print(f"DEBUG: Attempting to save to: {f_path}")
             print(f"DEBUG: Context: {ctx_path}")
             
             # Save the FULL sheet (not just filtered view)
             st.session_state.crm_sheets[st.session_state.active_sheet].to_json(f_path)
             print(f"DEBUG: Successfully wrote to {f_path}")
             st.toast(f"Saved to {ctx_path}", icon="✅")
         except Exception as e:
             st.error(f"Save failed: {e}")
             print(f"ERROR: Auto-save crash: {e}")
             
    
    st.markdown("<br>", unsafe_allow_html=True) 

    # --- BOTTOM SHEET TABS (Google Sheets Style) ---
    st.markdown("---")
    sheet_names = list(st.session_state.crm_sheets.keys())
    
    # Layout: [Add] [Tab 1] ... [Spacer] [Options]
    # Ratios: Add(0.5) | Tabs(1.5 each) | Spacer(6) | Options(1)
    col_ratios = [0.5] + [1.5] * len(sheet_names) + [6, 1]
    cols = st.columns(col_ratios)
    
    # 1. Add Button (Leftmost)
    with cols[0]:
         if st.button("➕", key="btn_add_sheet_google", help="Add Sheet"):
             new_name = f"Sheet{len(sheet_names) + 1}"
             if new_name not in st.session_state.crm_sheets:
                 st.session_state.crm_sheets[new_name] = pd.DataFrame(columns=required_cols)
                 st.session_state.active_sheet = new_name
                 save_active_to_disk()
                 st.rerun()

    # 2. Sheet Tabs
    for i, s_name in enumerate(sheet_names):
        with cols[i+1]:
            is_active = (s_name == st.session_state.active_sheet)
            if st.button(s_name, key=f"google_tab_{s_name}", type="primary" if is_active else "secondary", use_container_width=True):
                st.session_state.active_sheet = s_name
                st.rerun()

    # 3. Sheet Options (Rightmost)
    with cols[-1]:
         with st.popover("⚙️", use_container_width=True, help="Sheet Options"):
             st.markdown("#### Sheet Settings")
             # Rename
             new_name_input = st.text_input("Rename Sheet", value=st.session_state.active_sheet, key="sheet_rename_popover")
             if st.button("Update Name", key="btn_update_name_popover"):
                 if new_name_input and new_name_input != st.session_state.active_sheet:
                     if new_name_input not in st.session_state.crm_sheets:
                        old_name = st.session_state.active_sheet
                        st.session_state.crm_sheets[new_name_input] = st.session_state.crm_sheets.pop(old_name)
                        delete_sheet_from_disk(old_name)
                        st.session_state.active_sheet = new_name_input
                        save_active_to_disk()
                        st.rerun()
                     else:
                        st.toast("Name exists!", icon="⚠️")
             
             st.divider()
             
             # Delete
             if st.button("Delete Sheet", key="btn_del_sheet_popover", type="primary"):
                 if len(st.session_state.crm_sheets) > 1:
                     old_name = st.session_state.active_sheet
                     delete_sheet_from_disk(old_name)
                     st.session_state.crm_sheets.pop(old_name)
                     st.session_state.active_sheet = list(st.session_state.crm_sheets.keys())[0]
                     st.rerun()
                 else:
                     st.toast("Cannot delete last sheet", icon="⚠️")
         


    # --- 8. ADD LEAD FORM (Moved to Dialog) ---
if "Spreadsheet" in page:
    # Enhanced Header with Premium UI
    # Embedded Refined SVG with 3D Effect
    icon_svg = """<svg width="74" height="74" viewBox="0 0 64 64" fill="none" xmlns="http://www.w3.org/2000/svg"><filter id="shadow" x="-50%" y="-50%" width="200%" height="200%"><feDropShadow dx="0" dy="4" stdDeviation="4" flood-color="rgba(16, 124, 65, 0.3)"/></filter><rect x="8" y="8" width="48" height="48" rx="8" fill="#107C41" style="filter: url(#shadow);"/><linearGradient id="grad1" x1="8" y1="8" x2="56" y="56" gradientUnits="userSpaceOnUse"><stop offset="0%" stop-color="#21A366"/><stop offset="100%" stop-color="#107C41"/></linearGradient><rect x="8" y="8" width="48" height="48" rx="8" fill="url(#grad1)"/><path d="M22 8V56" stroke="white" stroke-width="2" stroke-opacity="0.4"/><path d="M40 8V56" stroke="white" stroke-width="2" stroke-opacity="0.4"/><path d="M8 24H56" stroke="white" stroke-width="2" stroke-opacity="0.4"/><path d="M8 40H56" stroke="white" stroke-width="2" stroke-opacity="0.4"/><rect x="25" y="27" width="12" height="10" rx="2" fill="white" fill-opacity="0.9" style="animation: pulseCell 3s infinite;"/></svg>"""
    
    st.markdown(f"""<style>@keyframes floatIcon {{ 0%, 100% {{ transform: perspective(500px) rotateY(10deg) rotateX(5deg) translateY(0px); }} 50% {{ transform: perspective(500px) rotateY(10deg) rotateX(5deg) translateY(-8px); }} }} @keyframes pulseCell {{ 0%, 100% {{ fill-opacity: 0.9; }} 50% {{ fill-opacity: 0.4; }} }}</style><div style="background: linear-gradient(145deg, #ffffff 0%, #f1f5f9 100%); border-radius: 20px; padding: 28px; box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.05), 0 4px 6px -2px rgba(0, 0, 0, 0.025); border: 1px solid rgba(255, 255, 255, 0.8); margin-bottom: 24px; display: flex; align-items: center; gap: 24px; position: relative; overflow: hidden;">
<div style="position: absolute; top: -20px; right: -20px; width: 150px; height: 150px; background: radial-gradient(circle, rgba(33, 163, 102, 0.08) 0%, rgba(255, 255, 255, 0) 70%); border-radius: 50%; pointer-events: none;"></div>
<div style="flex-shrink: 0; animation: floatIcon 6s ease-in-out infinite;">
{icon_svg}
</div>
<div style="z-index: 1;">
<h1 style="margin: 0; font-size: 2.5rem; color: #0f172a; font-weight: 800; letter-spacing: -0.03em; line-height: 1.1; font-family: 'Inter', sans-serif;">Spreadsheet Intelligence</h1>
<p style="margin: 8px 0 0 0; color: #475569; font-size: 1.1rem; font-weight: 500; line-height: 1.5;">Unlock insights from your Excel files. Deduplicate, compare, and analyze data with precision.</p>
</div>
</div>""", unsafe_allow_html=True)

    import io
    import re
    # Ensure openpyxl is installed for Excel writing (standard in streamlit envs)

    # --- HELPERS ---
    def normalize_text(text):
        """Strict normalization: lower, trim, remove special chars. Returns None for noise."""
        if pd.isna(text):
            return None
        s = str(text).strip().lower()
        if s in ["", "none", "nan", "null"]:
            return None
        # Remove special characters (keep alphanumeric and spaces)
        s = re.sub(r'[^a-zA-Z0-9\s]', '', s)
        return s if s.strip() != "" else None

    # Helper for styled sub-sections
    def styled_header(title, icon="🚀"):
        st.markdown(f"""
        <div style="
            padding: 10px 15px;
            background: #f8fafc;
            border-left: 4px solid #3b82f6;
            margin: 20px 0 10px 0;
            border-radius: 0 8px 8px 0;
        ">
            <h3 style="margin:0; font-size:1.1rem; color:#334155; display:flex; align-items:center; gap:8px;">
                <span>{icon}</span> {title}
            </h3>
        </div>
        """, unsafe_allow_html=True)

    # Custom CSS for Tabs
    st.markdown("""
<style>
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background-color: transparent;
        border-bottom: 2px solid #e2e8f0;
        padding-bottom: 0px;
    }

    
    .stTabs [data-baseweb="tab"] {
        height: 48px;
        background-color: transparent;
        border: none;
        color: #64748b;
        font-weight: 600;
        font-size: 14px;
        padding: 0 20px;
        border-radius: 8px 8px 0 0;
        transition: all 0.2s ease;
        margin-bottom: -2px; 
    }

    
    .stTabs [data-baseweb="tab"]:hover {
        color: #1e293b;
        background-color: #f1f5f9;
        border-bottom: 2px solid #cbd5e1;
    }

    
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        color: #2563eb; 
        background-color: #ffffff;
        border-bottom: 2px solid #2563eb;
    }
    
    
    .stTabs [data-baseweb="tab"][aria-selected="true"] p {
        font-weight: 700;
    }
</style>
""", unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["Single File Analysis", "1-vs-1 Comparison", "Multi-File Clean"])

    # ==========================
    # TAB 1: SINGLE FILE (ALL SHEETS)
    # ==========================
    with tab1:
        styled_header("Single File Duplicate Analysis")
        st.markdown("<p style='color:#64748b; margin-bottom:15px; font-size:0.95rem;'>Process a workbook to separate Duplicates and Unique rows into distinct sheets.</p>", unsafe_allow_html=True)

        # Define Helper
        def load_excel_safe_single(file_obj, respect_filters=True):
            if not respect_filters:
                file_obj.seek(0)
                return pd.read_excel(file_obj, sheet_name=None), 0

            import openpyxl
            file_obj.seek(0)
            wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=False)
            xls = {}
            total_hidden_skipped = 0
            
            for s_name in wb.sheetnames:
                ws = wb[s_name]
                if ws.sheet_state != 'visible': continue
                    
                data = []
                sheet_hidden_count = 0
                
                rows_iter = ws.iter_rows()
                try:
                    header_row = next(rows_iter)
                    raw_headers = [cell.value for cell in header_row]
                    headers = []
                    counts = {}
                    for i, h in enumerate(raw_headers):
                        h_str = str(h).strip() if h is not None else f"Unnamed: {i}"
                        if h_str == "": h_str = f"Unnamed: {i}"
                        if h_str in counts:
                            counts[h_str] += 1
                            headers.append(f"{h_str}.{counts[h_str]}")
                        else:
                            counts[h_str] = 0
                            headers.append(h_str)
                except StopIteration:
                    continue # Empty sheet

                for row in rows_iter:
                    current_row_idx = row[0].row
                    is_hidden = False
                    if current_row_idx in ws.row_dimensions:
                        dim = ws.row_dimensions[current_row_idx]
                        if dim.hidden: is_hidden = True
                        
                    if is_hidden:
                        sheet_hidden_count += 1
                    else:
                        row_values = [cell.value for cell in row]
                        if len(row_values) < len(headers):
                            row_values += [None] * (len(headers) - len(row_values))
                        elif len(row_values) > len(headers):
                            row_values = row_values[:len(headers)]
                        data.append(row_values)
                
                total_hidden_skipped += sheet_hidden_count
                if data:
                    xls[s_name] = pd.DataFrame(data, columns=headers)
                else:
                    xls[s_name] = pd.DataFrame(columns=headers)
            
            return xls, total_hidden_skipped

        # STEP 1: UPLOAD
        st.markdown("""
        <div style="margin-top: 20px; margin-bottom: 10px; display: flex; align-items: center; gap: 10px;">
            <div style="background: #eff6ff; color: #1d4ed8; font-weight: 700; padding: 5px 12px; border-radius: 20px; font-size: 0.9rem;">Step 1</div>
            <span style="font-weight: 600; color: #334155; font-size: 1.05rem;">Upload Workbook</span>
        </div>
        """, unsafe_allow_html=True)
        
        with st.container(border=True):
            st.markdown("#### 📁 Input Workbook")
            st.caption("Upload the Excel file you want to analyze for duplicates.")
            f_single = st.file_uploader("Upload Excel", type=['xlsx'], key="sit_single", label_visibility="collapsed")

        if f_single:
            try:
                # STEP 2: CONFIGURE
                st.markdown("""
                <div style="margin-top: 20px; margin-bottom: 10px; display: flex; align-items: center; gap: 10px;">
                    <div style="background: #eff6ff; color: #1d4ed8; font-weight: 700; padding: 5px 12px; border-radius: 20px; font-size: 0.9rem;">Step 2</div>
                    <span style="font-weight: 600; color: #334155; font-size: 1.05rem;">Configure Analysis</span>
                </div>
                """, unsafe_allow_html=True)

                with st.container(border=True):
                    use_excel_filters = st.checkbox("Respect Excel Filters (Exclude Hidden Rows)", value=True, help="Make sure to SAVE your Excel file with the filters active before uploading.")
                    
                    xls, total_hidden_skipped = load_excel_safe_single(f_single, use_excel_filters)
                    
                    if use_excel_filters:
                        if total_hidden_skipped > 0:
                            st.success(f"✅ Successfully excluded {total_hidden_skipped} hidden rows from analysis.")
                        else:
                            st.warning("⚠️ No hidden rows were detected. Please ensure you SAVED the Excel file with filters applied.")
                    
                    sheet_names = list(xls.keys())
                    
                    # Check for empty sheets
                    valid_sheets = {k: v for k, v in xls.items() if not v.empty}
                    
                    if not valid_sheets:
                        st.error("Workbook contains no data.")
                    else:
                        st.success(f"Loaded {len(valid_sheets)} sheets: {', '.join(valid_sheets.keys())}")
                        
                        # 2. COLUMN SELECTION (Global)
                        all_cols = set()
                        for df in valid_sheets.values():
                            all_cols.update(df.columns.astype(str))
                        
                        target_cols = st.multiselect("Select Duplicate Key Columns (e.g. Email, Phone)", sorted(list(all_cols)))

                if target_cols:
                     if st.button("🚀 Process Workbook", type="primary", use_container_width=True):

                    

                            




                        with st.spinner("Analyzing across all sheets..."):
                            
                            # A. PREPARE GLOBAL DATASET FOR DETECTION
                            # We need to track: (SheetName, OriginalIndex) -> NormalizedKey
                            global_rows = []
                            
                            for s_name, df in valid_sheets.items():
                                # Check if target cols exist in this sheet
                                missing = [c for c in target_cols if c not in df.columns]
                                if missing:
                                    # Skip or handle? Let's treat missing cols as None
                                    pass
                                
                                # Normalize
                                temp_df = df.copy()
                                temp_df['_sheet'] = s_name
                                temp_df['_idx'] = temp_df.index
                                
                                # normalization
                                norm_vals = []
                                for c in target_cols:
                                    if c in df.columns:
                                        # Apply norm
                                        temp_df[f'__norm_{c}'] = temp_df[c].apply(normalize_text)
                                    else:
                                        temp_df[f'__norm_{c}'] = None
                                
                                global_rows.append(temp_df)
                            
                            if not global_rows:
                                st.error("No data found to process.")
                            else:
                                master_df = pd.concat(global_rows, ignore_index=True)
                                
                                # Norm Cols Key
                                norm_keys = [f'__norm_{c}' for c in target_cols]
                                
                                # Do NOT drop rows with empty keys, preserve them as Unique
                                master_valid = master_df.copy()
                                
                                # 1. Identify ALL duplicate instances (for location reporting)
                                all_dups_mask = master_valid.duplicated(subset=norm_keys, keep=False)
                                
                                # Handle empty keys 
                                empty_mask = master_valid[norm_keys].isna().all(axis=1)
                                all_dups_mask = all_dups_mask & (~empty_mask)

                                # 2. Identify ONLY 2nd+ instances (for removal/splitting)
                                # keep='first' marks duplicates as True, except the first one (Unique).
                                split_mask = master_valid.duplicated(subset=norm_keys, keep='first')
                                split_mask = split_mask & (~empty_mask)
                                
                                master_valid['is_duplicate'] = split_mask

                                # 3. Generate Location Map using ALL instances
                                if not master_valid.empty and all_dups_mask.any():
                                    dup_rows_all = master_valid[all_dups_mask].copy()
                                    dup_rows_all['__loc_label'] = dup_rows_all['_sheet'] + " (Row " + (dup_rows_all['_idx'] + 2).astype(str) + ")"
                                    dup_rows_all['__group_key'] = dup_rows_all[norm_keys].apply(tuple, axis=1)
                                    
                                    loc_map = dup_rows_all.groupby('__group_key')['__loc_label'].apply(lambda x: ", ".join(x)).to_dict()
                                    
                                    master_valid['__group_key'] = master_valid[norm_keys].apply(tuple, axis=1)
                                    master_valid['duplicate_locations'] = master_valid['__group_key'].map(loc_map)
                                else:
                                    master_valid['duplicate_locations'] = None
                                
                                # Now we map back to splitting Dups vs Uniques per sheet
                                # Create Output Excel
                                output = io.BytesIO()
                                writer = pd.ExcelWriter(output, engine='openpyxl')
                                
                                total_dups = 0
                                total_uniques = 0
                                
                                results_summary = {} # {sheet: {'dups': count, 'uniques': count}}
                                
                                for s_name in valid_sheets.keys():
                                    # Get rows belonging to this sheet from valid set
                                    sheet_subset = master_valid[master_valid['_sheet'] == s_name]
                                    
                                    # Split
                                    dups_df = sheet_subset[sheet_subset['is_duplicate'] == True]
                                    uniques_df = sheet_subset[sheet_subset['is_duplicate'] == False]
                                    
                                    # Clean helper cols
                                    # Retrieve original rows using index is tricky after concat/drop.
                                    # Better: Use the stored original index or just columns.
                                    # We have original columns in sheet_subset. Just drop helpers.
                                    drop_cols = ['_sheet', '_idx', 'is_duplicate'] + norm_keys
                                    
                                    # For original DF cols
                                    orig_cols = list(valid_sheets[s_name].columns)
                                    
                                    # Add 'duplicate_locations' to dups output
                                    final_dups = dups_df[orig_cols + ['duplicate_locations']]
                                    final_uniques = uniques_df[orig_cols]
                                    
                                    total_dups += len(final_dups)
                                    total_uniques += len(final_uniques)
                                    results_summary[s_name] = {'dups': len(final_dups), 'uniques': len(final_uniques)}
                                    
                                    # Write to Excel
                                    if not final_dups.empty:
                                        final_dups.to_excel(writer, sheet_name=f"{s_name[:20]}_Dups", index=False)
                                    if not final_uniques.empty:
                                        final_uniques.to_excel(writer, sheet_name=f"{s_name[:20]}_Uniques", index=False)
                                
                                writer.close()
                                
                                # UI METRICS
                                m1, m2, m3 = st.columns(3)
                                m1.metric("Rows Scanned", len(master_valid))
                                m2.metric("Total Duplicates", total_dups, delta_color="inverse")
                                m3.metric("Total Unique", total_uniques)
                                
                                st.markdown("### 📋 Sheet Breakdown")
                                for s in valid_sheets.keys():
                                    # Filter again for display logic
                                    sub = master_valid[master_valid['_sheet'] == s]
                                    d_rows = sub[sub['is_duplicate'] == True]
                                    u_count = len(sub) - len(d_rows)
                                    
                                    # Restore original cols for display
                                    orig_cols = list(valid_sheets[s].columns)
                                    d_disp = d_rows[orig_cols + ['duplicate_locations']]
                                    
                                    with st.expander(f"{s} (Dups: {len(d_disp)} | Uniques: {u_count})"):
                                        if not d_disp.empty:
                                            st.warning(f"Found {len(d_disp)} duplicates.")
                                            st.dataframe(d_disp, use_container_width=True)
                                        else:
                                            st.success("No duplicates in this sheet.")
                                            
                                        if u_count > 0:
                                            st.info(f"Sheet contains {u_count} unique rows.")

                                # DOWNLOAD
                                processed_data = output.getvalue()
                                st.download_button(
                                    label="📥 Download Split Workbook (XLSX)",
                                    data=processed_data,
                                    file_name="processed_workbook.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )

            except Exception as e:
                st.error(f"Error: {e}")

    # ==========================
    # TAB 2: CROSS FILE (A vs B)
    # ==========================
    with tab2:
        # Define helper first (cleaner scope)
        def load_excel_safe(file_obj, respect_filters=True):
            """Helper to load excel, respecting hidden rows if requested."""
            if not respect_filters:
                file_obj.seek(0)
                return pd.read_excel(file_obj, sheet_name=None)
            
            import openpyxl
            file_obj.seek(0)
            wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=False)
            res = {}
            for s_name in wb.sheetnames:
                ws = wb[s_name]
                if ws.sheet_state != 'visible': continue
                
                data = []
                rows_iter = ws.iter_rows()
                try:
                    header_row = next(rows_iter)
                    raw_headers = [cell.value for cell in header_row]
                    headers = []
                    counts = {}
                    for i, h in enumerate(raw_headers):
                        h_str = str(h).strip() if h is not None else f"Unnamed: {i}"
                        if h_str == "": h_str = f"Unnamed: {i}"
                        if h_str in counts:
                            counts[h_str] += 1
                            headers.append(f"{h_str}.{counts[h_str]}")
                        else:
                            counts[h_str] = 0
                            headers.append(h_str)
                except StopIteration:
                    continue

                for row in rows_iter:
                    current_row_idx = row[0].row
                    is_hidden = False
                    if current_row_idx in ws.row_dimensions:
                        dim = ws.row_dimensions[current_row_idx]
                        if dim.hidden:
                            is_hidden = True
                    
                    if not is_hidden:
                        row_values = [cell.value for cell in row]
                        if len(row_values) < len(headers):
                            row_values += [None] * (len(headers) - len(row_values))
                        elif len(row_values) > len(headers):
                            row_values = row_values[:len(headers)]
                        data.append(row_values)
                
                if data:
                    res[s_name] = pd.DataFrame(data, columns=headers)
                else:
                    res[s_name] = pd.DataFrame(columns=headers)
            return res

        # UI STRUTURE
        # STEP 1: UPLOADS
        st.markdown("""
        <div style="margin-top: 20px; margin-bottom: 10px; display: flex; align-items: center; gap: 10px;">
            <div style="background: #eff6ff; color: #1d4ed8; font-weight: 700; padding: 5px 12px; border-radius: 20px; font-size: 0.9rem;">Step 1</div>
            <span style="font-weight: 600; color: #334155; font-size: 1.05rem;">Upload Files</span>
        </div>
        """, unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            with st.container(border=True):
                st.markdown("#### 📁 File A (Baseline)")
                st.caption("The original file to compare against.")
                f_a = st.file_uploader("File A", type=['xlsx'], key="sit_a_1", label_visibility="collapsed")
        with c2:
            with st.container(border=True):
                st.markdown("#### 🆕 File B (New)")
                st.caption("The new file containing potential updates.")
                f_b = st.file_uploader("File B", type=['xlsx'], key="sit_b_1", label_visibility="collapsed")

        if f_a and f_b:
            try:
                # 1. READ BOTH
                st.markdown("""
                <div style="margin-top: 20px; margin-bottom: 10px; display: flex; align-items: center; gap: 10px;">
                    <div style="background: #eff6ff; color: #1d4ed8; font-weight: 700; padding: 5px 12px; border-radius: 20px; font-size: 0.9rem;">Step 2</div>
                    <span style="font-weight: 600; color: #334155; font-size: 1.05rem;">Configure Data</span>
                </div>
                """, unsafe_allow_html=True)

                with st.container(border=True):
                    use_filters_cross = st.checkbox("Respect Excel Filters (Exclude Hidden Rows)", value=True, help="Make sure to SAVE your Excel files with filters active.", key="cross_filter_check_1")

                    xls_a = load_excel_safe(f_a, use_filters_cross)
                    xls_b = load_excel_safe(f_b, use_filters_cross)
                    
                    if not xls_a or not xls_b:
                        st.error("One or both workbooks contain no visible data.")
                    else:
                        all_cols_b = set()
                        for df in xls_b.values():
                            # Clean Unnamed cols for selection
                            clean_cols = [c for c in df.columns if not str(c).startswith("Unnamed")]
                            all_cols_b.update(clean_cols)
                        
                        match_cols = st.multiselect("Select Unique Identifier Columns (e.g. Email, ID)", sorted(list(all_cols_b)), key="ms_1")

                if match_cols and st.button("🚀 Find New Rows in B", type="primary", use_container_width=True, key="btn_1"):
                    
                    # Helper for styling
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    def format_and_autofit_sheet(ws, df):
                        # 1. Header Style
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Green like screenshot
                        center_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        
                        # Apply to Row 1
                        for cell in ws[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                        
                        # 2. Auto-Filter & Freeze Panes
                        ws.auto_filter.ref = ws.dimensions
                        ws.freeze_panes = "A2"
                        
                        # 3. Auto-fit Columns
                        for i, col in enumerate(df.columns, 1):
                            max_len = len(str(col)) + 4
                            col_letter = get_column_letter(i)
                            
                            # Sample data for width
                            for row_x in range(min(len(df), 20)):
                                val = df.iloc[row_x, i-1]
                                if val:
                                    max_len = max(max_len, len(str(val)))
                            
                            # Cap width
                            adjusted_width = min(max_len + 2, 50) 
                            ws.column_dimensions[col_letter].width = adjusted_width
                            
                            # Apply wrap text to all data cells
                            for cell in ws[col_letter]:
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                                # Reset header align
                                if cell.row == 1:
                                    cell.alignment = center_align

                    with st.spinner("Building baseline from File A..."):
                        baseline_tuples = set()
                        count_a = 0
                        for df in xls_a.values():
                            valid_cols = [c for c in match_cols if c in df.columns]
                            if len(valid_cols) != len(match_cols): continue
                            temp = df[valid_cols].copy()
                            for c in valid_cols:
                                temp[c] = temp[c].apply(normalize_text)
                            temp = temp.dropna(how='all')
                            count_a += len(temp)
                            baseline_tuples.update(list(temp.itertuples(index=False, name=None)))
                        
                    with st.spinner(f"Comparing File B ({len(xls_b)} sheets) against {len(baseline_tuples)} unique baseline records..."):
                        output_b = io.BytesIO()
                        writer_b = pd.ExcelWriter(output_b, engine='openpyxl')
                        total_new = 0
                        results_b = {}
                        
                        for s_name, df in xls_b.items():
                            # Remove 'Unnamed' columns from processing/output
                            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                            
                            valid_cols = [c for c in match_cols if c in df.columns]
                            if len(valid_cols) != len(match_cols): continue
                            temp_b = df.copy()
                            norm_keys = []
                            for c in valid_cols:
                                norm_n = f"__norm_{c}"
                                temp_b[norm_n] = temp_b[c].apply(normalize_text)
                                norm_keys.append(norm_n)
                            temp_b_valid = temp_b.dropna(subset=norm_keys, how='all')
                            temp_b_valid['__tuple'] = temp_b_valid[norm_keys].apply(tuple, axis=1)
                            new_rows_mask = ~temp_b_valid['__tuple'].isin(baseline_tuples)
                            new_rows_df = temp_b_valid[new_rows_mask]
                            
                            # Use cleaned columns for output
                            final_new = new_rows_df[[c for c in df.columns if c in new_rows_df.columns]]
                            
                            cnt = len(final_new)
                            total_new += cnt
                            results_b[s_name] = final_new
                            if cnt > 0:
                                sheet_n = f"{s_name[:20]}_New"
                                final_new.to_excel(writer_b, sheet_name=sheet_n, index=False)
                                # Apply Styling
                                format_and_autofit_sheet(writer_b.sheets[sheet_n], final_new)
                        
                        writer_b.close()
                        
                        c_m1, c_m2, c_m3 = st.columns(3)
                        c_m1.metric("Baseline Rows (A)", count_a)
                        c_m2.metric("Total Rows Processed (B)", sum([len(d) for d in xls_b.values()]))
                        c_m3.metric("New Unique Rows Found", total_new, delta_color="normal")
                        
                        st.markdown("### 🆕 New Rows per Sheet")
                        for s, df_new in results_b.items():
                            cnt = len(df_new)
                            if cnt > 0:
                                with st.expander(f"{s} ({cnt} New Rows)"):
                                    st.warning(f"Found {cnt} rows unique to File B.")
                                    st.dataframe(df_new, use_container_width=True)
                            else:
                                with st.expander(f"{s} (No New Rows)"):
                                    st.success("All rows in this sheet already exist in File A.")
                        
                        # --- FULL PREVIEW SECTION ---
                        st.markdown("---")
                        st.subheader("👁️ Full File Content Preview")
                        st.caption("Browse all generated new rows across all sheets.")
                        
                        valid_sheets_res = [s for s, d in results_b.items() if not d.empty]
                        if valid_sheets_res:
                            sheet_tabs = st.tabs(valid_sheets_res)
                            for i, s_name in enumerate(valid_sheets_res):
                                with sheet_tabs[i]:
                                    st.dataframe(results_b[s_name], height=600, use_container_width=True)
                        else:
                            st.info("No new rows found to preview.")

                        # --- ACTIONS ---
                        # --- ACTIONS ---
                        st.markdown("### 📤 Actions")
                        
                        # Custom Filename Input
                        custom_filename = st.text_input("Output File Name (Optional)", value="comparison_unique_to_B", help="Enter a name for the output Excel file.")
                        final_file_name = custom_filename.strip()
                        if not final_file_name.lower().endswith(".xlsx"):
                            final_file_name += ".xlsx"
                        
                        col_actions = st.columns([1, 1, 2])
                        
                        if total_new > 0:
                            processed_b = output_b.getvalue()
                            
                            # 1. Download Button
                            with col_actions[0]:
                                st.download_button(
                                    label="📥 Download Result",
                                    data=processed_b,
                                    file_name=final_file_name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )
                            
                            # 2. Upload to Drive Button (Optional Section)
                            with col_actions[1]:
                                with st.expander("☁️ Save to Google Drive (Optional)"):
                                    import os
                                    has_creds = os.path.exists("service-account.json")
                                    
                                    if not has_creds:
                                        st.warning("Setup Required")
                                        st.caption("To enable one-click uploads, you need a **Google Service Account Key**.")
                                        st.caption("Don't want to set this up? Just use the **Download** button on the left instead!")
                                        
                                        uploaded_key = st.file_uploader("Upload service-account.json", type=['json'], key="sa_uploader")
                                        if uploaded_key:
                                            with open("service-account.json", "wb") as f:
                                                f.write(uploaded_key.getbuffer())
                                            st.success("✅ Credentials saved! refreshing...")
                                            st.rerun()
                                    else:
                                        # Creds exist, try to upload
                                        try:
                                            # Check usage dependencies
                                            try:
                                                from googleapiclient.discovery import build
                                                from google.oauth2 import service_account
                                                from googleapiclient.http import MediaIoBaseUpload
                                            except ImportError:
                                                st.error("Missing Google Libraries.")
                                                if st.button("🔧 Install Google Dependencies"):
                                                    import subprocess, sys
                                                    subprocess.check_call([sys.executable, "-m", "pip", "install", "google-api-python-client", "google-auth-httplib2", "google-auth-oauthlib"])
                                                    st.rerun()
                                                st.stop()

                                            if st.button("☁️ Upload Now"):
                                                SCOPES = ['https://www.googleapis.com/auth/drive.file']
                                                creds = service_account.Credentials.from_service_account_file(
                                                    'service-account.json', scopes=SCOPES)
                                                drive_service = build('drive', 'v3', credentials=creds)
                                                
                                                file_metadata = {'name': final_file_name}
                                                
                                                media = MediaIoBaseUpload(io.BytesIO(processed_b), 
                                                                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                                                        resumable=True)
                                                
                                                with st.spinner(f"Uploading '{final_file_name}' to Google Drive..."):
                                                    file = drive_service.files().create(body=file_metadata,
                                                                                        media_body=media,
                                                                                        fields='id, webViewLink').execute()
                                                    
                                                st.success(f"✅ Uploaded! [Open in Drive]({file.get('webViewLink')})")
                                                
                                        except Exception as e:
                                            st.error(f"Upload failed: {e}")

            except Exception as e:
                st.error(f" Comparison Error: {e}")

    # ==========================
    # TAB 3: MULTI FILE (Multiple vs One)
    # ==========================
    with tab3:
        # Define Helper First
        def load_excel_safe_multi(file_obj, respect_filters=True):
            # We need to always use openpyxl to get row indices, even if filters are off
            # because we need to map back to the original file for deletion.
            import openpyxl
            file_obj.seek(0)
            wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=False)
            res = {}
            for s_name in wb.sheetnames:
                ws = wb[s_name]
                if ws.sheet_state != 'visible': continue
                
                data = []
                rows_iter = ws.iter_rows()
                try:
                    header_row = next(rows_iter)
                    raw_headers = [cell.value for cell in header_row]
                    headers = []
                    counts = {}
                    for i, h in enumerate(raw_headers):
                        h_str = str(h).strip() if h is not None else f"Unnamed: {i}"
                        if h_str == "": h_str = f"Unnamed: {i}"
                        if h_str in counts:
                            counts[h_str] += 1
                            headers.append(f"{h_str}.{counts[h_str]}")
                        else:
                            counts[h_str] = 0
                            headers.append(h_str)
                except StopIteration:
                        continue

                # Add a special column for Row Index
                headers.append("__row_idx")

                for row in rows_iter:
                    current_row_idx = row[0].row
                    is_hidden = False
                    if respect_filters and current_row_idx in ws.row_dimensions:
                        dim = ws.row_dimensions[current_row_idx]
                        if dim.hidden: is_hidden = True
                    
                    if not is_hidden:
                        row_values = [cell.value for cell in row]
                        # Pad or trim
                        # The headers list now includes __row_idx at the end, so we compare against len(headers)-1
                        expected_cols = len(headers) - 1
                        if len(row_values) < expected_cols:
                            row_values += [None] * (expected_cols - len(row_values))
                        elif len(row_values) > expected_cols:
                            row_values = row_values[:expected_cols]
                        
                        # Append the row index
                        row_values.append(current_row_idx)
                        data.append(row_values)
                
                if data:
                    res[s_name] = pd.DataFrame(data, columns=headers)
                else:
                    res[s_name] = pd.DataFrame(columns=headers)
            return res

        # UI
        # STEP 1: UPLOAD
        st.markdown("""
        <div style="margin-top: 20px; margin-bottom: 10px; display: flex; align-items: center; gap: 10px;">
            <div style="background: #eff6ff; color: #1d4ed8; font-weight: 700; padding: 5px 12px; border-radius: 20px; font-size: 0.9rem;">Step 1</div>
            <span style="font-weight: 600; color: #334155; font-size: 1.05rem;">Upload Files</span>
        </div>
        """, unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            with st.container(border=True):
                st.markdown("#### 📁 Baseline Files (A)")
                st.caption("Upload one or more reference files.")
                files_a_multi = st.file_uploader("Files A", type=['xlsx'], accept_multiple_files=True, key="sit_a_multi", label_visibility="collapsed")
        with c2:
            with st.container(border=True):
                st.markdown("#### 🆕 Target Files (B)")
                st.caption("Upload files to clean duplicates from.")
                files_b_multi = st.file_uploader("Files B", type=['xlsx'], accept_multiple_files=True, key="sit_b_multi", label_visibility="collapsed")

        if files_a_multi and files_b_multi:
            try:
                # STEP 2: CONFIGURE
                st.markdown("""
                <div style="margin-top: 20px; margin-bottom: 10px; display: flex; align-items: center; gap: 10px;">
                    <div style="background: #eff6ff; color: #1d4ed8; font-weight: 700; padding: 5px 12px; border-radius: 20px; font-size: 0.9rem;">Step 2</div>
                    <span style="font-weight: 600; color: #334155; font-size: 1.05rem;">Configure & Process</span>
                </div>
                """, unsafe_allow_html=True)

                with st.container(border=True):
                    use_filters_multi = st.checkbox("Respect Excel Filters (Exclude Hidden Rows)", value=True, help="Make sure to SAVE your Excel files with filters active.", key="cross_filter_check_multi")

                    # Load A (Baseline)
                    xls_a_map = {}
                    for f in files_a_multi:
                        xls_a_map[f.name] = load_excel_safe_multi(f, use_filters_multi)
                    
                    # Load B (Target)
                    xls_b_map = {} 
                    for f in files_b_multi:
                        xls_b_map[f.name] = load_excel_safe_multi(f, use_filters_multi)

                    if not xls_a_map or not xls_b_map:
                        st.error("Baseline or Comparison files contain no visible data.")
                    else:
                        # Collect all columns from B (excluding __row_idx) to choose match cols
                        all_cols_b = set()
                        for f_name, sheets in xls_b_map.items():
                            for df in sheets.values():
                                cols = [c for c in df.columns if c != "__row_idx"]
                                all_cols_b.update(cols)
                        
                        match_cols = st.multiselect("Select Unique Identifier Columns", sorted(list(all_cols_b)), key="multi_col_select")

                    if match_cols:
                        if st.button("🚀 Find New Rows (Clean Files)", type="primary", use_container_width=True, key="btn_multi"):
                            
                            # 1. BUILD BASELINE
                            with st.spinner("Building combined baseline from Files A..."):
                                baseline_tuples = set()
                                count_a = 0
                                for f_name, sheets in xls_a_map.items():
                                    for df in sheets.values():
                                        valid_cols = [c for c in match_cols if c in df.columns]
                                        if len(valid_cols) != len(match_cols): continue
                                        temp = df[valid_cols].copy()
                                        for c in valid_cols:
                                            temp[c] = temp[c].apply(normalize_text)
                                        temp = temp.dropna(how='all')
                                        count_a += len(temp)
                                        baseline_tuples.update(list(temp.itertuples(index=False, name=None)))
                            
                            # 2. PROCESS FILES
                            import zipfile
                            import io
                            import openpyxl

                            # We will create a Zip file containing the "Cleaned" versions of File B
                            zip_buffer = io.BytesIO()
                            
                            with st.spinner("Processing files... Removing duplicates while preserving styles."):
                                processed_count = 0
                                total_removed = 0
                                
                                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                                    
                                    for f_obj in files_b_multi:
                                        f_name = f_obj.name
                                        sheets_data = xls_b_map.get(f_name, {})
                                        
                                        # We need to modify the ORIGINAL file
                                        f_obj.seek(0)
                                        wb = openpyxl.load_workbook(f_obj)
                                        
                                        file_removed_count = 0
                                        
                                        for s_name, df in sheets_data.items():
                                            if s_name not in wb.sheetnames: continue
                                            ws = wb[s_name]

                                            valid_cols = [c for c in match_cols if c in df.columns]
                                            if len(valid_cols) != len(match_cols): continue
                                            
                                            # Normalize and match
                                            temp_b = df.copy()
                                            norm_keys = []
                                            for c in valid_cols:
                                                norm_n = f"__norm_{c}"
                                                temp_b[norm_n] = temp_b[c].apply(normalize_text)
                                                norm_keys.append(norm_n)
                                            
                                            # Identify rows to DELETE (Duplicates)
                                            # Row must NOT be empty in keys to be a duplicate
                                            temp_b_valid = temp_b.dropna(subset=norm_keys, how='all')
                                            temp_b_valid['__tuple'] = temp_b_valid[norm_keys].apply(tuple, axis=1)
                                            
                                            # Mask: True if it IS in baseline (Duplicate)
                                            duplicates_mask = temp_b_valid['__tuple'].isin(baseline_tuples)
                                            
                                            # Get the __row_idx of these duplicates
                                            rows_to_delete = temp_b_valid.loc[duplicates_mask, '__row_idx'].tolist()
                                            
                                            if rows_to_delete:
                                                file_removed_count += len(rows_to_delete)
                                                # Delete rows in descending order to avoid shifting indices
                                                for r_idx in sorted(rows_to_delete, reverse=True):
                                                    ws.delete_rows(r_idx)
                                        
                                        total_removed += file_removed_count
                                        processed_count += 1
                                        
                                        # Save the modified workbook to the zip
                                        out = io.BytesIO()
                                        wb.save(out)
                                        clean_name = f"Cleaned_{f_name}"
                                        zf.writestr(clean_name, out.getvalue())

                                # 3. METRICS & DOWNLOAD
                                c_m1, c_m2 = st.columns(2)
                                c_m1.metric("Baseline Rows (A)", count_a)
                                c_m2.metric("Duplicate Rows Removed from B", total_removed)
                                
                                if processed_count > 0:
                                    zip_buffer.seek(0)
                                    st.success(f"Successfully processed {processed_count} files.")
                                    
                                    st.download_button(
                                        label="📥 Download Cleaned Files (ZIP)",
                                        data=zip_buffer,
                                        file_name="cleaned_files.zip",
                                        mime="application/zip",
                                        type="primary"
                                    )
            except Exception as e:
                st.error(f" Comparison Error: {e}")


# ================== POWER DIALER ==================
if "Power Dialer" in page:

    
    leads = fetch_data(LEADS_API)
    if not leads:
        st.info("No leads found.")
        st.stop()
        
    df = pd.DataFrame(leads)
    
    # Filter Logic
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    # Ensure nextFollowUpDate is handled
    if "nextFollowUpDate" not in df.columns:
        df["nextFollowUpDate"] = None
        
    # Convert properly to string
    df["nextFollowUpDate"] = df["nextFollowUpDate"].astype(str).replace("None", "").replace("nan", "")
    
    
    # Header Layout with Filter on Right
    c_header, c_filter = st.columns([3.5, 2])
    
    with c_header:
        # Import Outfit Font
        st.markdown("""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;700&display=swap');
        </style>
        """, unsafe_allow_html=True)
        
        st.markdown(f"""
        <h1 style='font-family: "Outfit", sans-serif; font-weight: 700; display: flex; align-items: center; gap: 10px; margin: 0; padding: 0; line-height: 1;'>
            <img src="https://raw.githubusercontent.com/Tarikul-Islam-Anik/Animated-Fluent-Emojis/master/Emojis/Objects/Telephone%20Receiver.png" width="50" height="50" alt="📞"> 
            Power Dialer
        </h1>
        """, unsafe_allow_html=True)
        
    with c_filter:
        # Custom Toggle Switch UI
        # --- HIDDEN NATIVE RADIO FOR STATE MANAGEMENT ---
        # We keep this in the DOM but invisible so our custom JS can click it.
        # This ensures Streamlit state is handled natively without full reloads.
        st.markdown("""
        <style>
        
        section[data-testid="stMain"] div[data-testid="stRadio"] > label { display: none; }
        section[data-testid="stMain"] div[key="pd_filter_choice"], 
        section[data-testid="stMain"] div[data-testid="stRadio"] {
            height: 0px; opacity: 0; overflow: hidden; margin: 0; padding: 0;
        }
        </style>
        """, unsafe_allow_html=True)
        
        filter_choice = st.radio(
            "Filter Leads:", 
            ["Today's Follow-ups", "All Leads"],
            index=0,
            horizontal=True,
            key="pd_filter_choice"
        )

        # --- CUSTOM 3D TOGGLE COMPONENT ---
        # User-provided source code adapted for Streamlit bridge
        
        # Determine initial state for the visual toggle
        init_checked_1 = "checked" if filter_choice == "Today's Follow-ups" else ""
        init_checked_2 = "checked" if filter_choice == "All Leads" else ""
        init_flap_text = "Today's Follow-ups" if filter_choice == "Today's Follow-ups" else "All Leads"
        init_rotation = "-15deg" if filter_choice == "Today's Follow-ups" else "15deg" # Initial animation state helper
        
        # HTML/CSS/JS Block
        toggle_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
        <style>
        :root {{
            --accent: #04da97;
            --border-width: 3px; 
            --border-radius: 40px;
            --font-size: 13px; 
        }}

        body {{
            display: flex;
            justify-content: center;
            align-items: center;
            margin: 0;
            background-color: transparent; 
            font-family: 'Inter', sans-serif;
            overflow: hidden;
        }}

        .container {{
            perspective: 800px;
            transform: scale(0.9); 
            transform-origin: center right;
        }}

        .toggle {{
            position: relative;
            border: solid var(--border-width) var(--accent);
            border-radius: var(--border-radius);
            transition: transform cubic-bezier(0, 0, 0.30, 2) .4s;
            transform-style: preserve-3d;
            perspective: 800px;
            display: flex;
            flex-direction: row;
        }}

        .toggle>input[type="radio"] {{
            display: none;
        }}

        .toggle>#choice1:checked~#flap {{
            transform: rotateY(-180deg);
        }}

        .toggle>#choice1:checked~#flap>.content {{
            transform: rotateY(-180deg);
        }}

        .toggle>#choice2:checked~#flap {{
            transform: rotateY(0deg);
        }}

        .toggle>label {{
            display: inline-block;
            min-width: 120px; 
            padding: 8px 10px;
            font-size: var(--font-size);
            text-align: center;
            color: var(--accent);
            cursor: pointer;
            white-space: nowrap;
             user-select: none;
        }}

        .toggle>label,
        .toggle>#flap {{
            font-weight: bold;
            text-transform: capitalize;
        }}

        .toggle>#flap {{
            position: absolute;
            top: calc( 0px - var(--border-width));
            left: 50%;
            height: calc(100% + var(--border-width) * 2);
            width: 50%;
            display: flex;
            justify-content: center;
            align-items: center;
            font-size: var(--font-size);
            background-color: var(--accent);
            border-top-right-radius: var(--border-radius);
            border-bottom-right-radius: var(--border-radius);
            transform-style: preserve-3d;
            transform-origin: left;
            transition: transform cubic-bezier(0.4, 0, 0.2, 1) .5s;
        }}

        .toggle>#flap>.content {{
            color: #333;
            transition: transform 0s linear .25s;
            transform-style: preserve-3d;
        }}
        </style>
        </head>
        <body>

        <div class="container">
            <form class="toggle">
                <input type="radio" id="choice1" name="choice" value="today" {init_checked_1} onclick="selectOption('today')">
                <label for="choice1">Today's Follow-ups</label>

                <input type="radio" id="choice2" name="choice" value="all" {init_checked_2} onclick="selectOption('all')">
                <label for="choice2">All Leads</label>

                <div id="flap"><span class="content">{init_flap_text}</span></div>
            </form>
        </div>

        <script>
        const st = {{}};
        st.flap = document.querySelector('#flap');
        st.toggle = document.querySelector('.toggle');
        st.choice1 = document.querySelector('#choice1');
        st.choice2 = document.querySelector('#choice2');

        st.flap.addEventListener('transitionend', () => {{
            if (st.choice1.checked) {{
                st.toggle.style.transform = 'rotateY(-15deg)';
                setTimeout(() => st.toggle.style.transform = '', 400);
            }} else {{
                st.toggle.style.transform = 'rotateY(15deg)';
                setTimeout(() => st.toggle.style.transform = '', 400);
            }}
        }})

        st.clickHandler = (e) => {{
            if (e.target.tagName === 'LABEL') {{
                setTimeout(() => {{
                    st.flap.children[0].textContent = e.target.textContent;
                }}, 250);
            }}
        }}

        document.addEventListener('click', (e) => st.clickHandler(e));

        // --- SMART BRIDGE TO STREAMLIT ---
        function selectOption(val) {{
            try {{
                // 1. Target labels specifically in the MAIN section (avoid sidebar)
                const doc = window.parent.document;
                const main = doc.querySelector('section[data-testid="stMain"]');
                
                if (main) {{
                    const labels = main.querySelectorAll('div[role="radiogroup"] label');
                    
                    // 2. Find by TEXT CONTENT instead of index to be safe
                    for (let label of labels) {{
                        const text = label.textContent.trim();
                        if (val === 'today' && text.includes("Today")) {{
                            label.click();
                            return;
                        }}
                        if (val === 'all' && text.includes("All")) {{
                            label.click();
                            return;
                        }}
                    }}
                }}
            }} catch(e) {{
                console.error("Streamlit Bridge Error:", e);
            }}
        }}
        </script>
        </body>
        </html>
        """
        
        components.html(toggle_html, height=55, scrolling=False)
    
    st.markdown("---") # Visual separator
    
    # Apply Filter
    if filter_choice == "Today's Follow-ups":
        today_str = datetime.now().strftime("%Y-%m-%d")
        mask_active = ~df["status"].str.contains("Closed", case=False, na=False)
        # Include today AND overdue (past dates)
        mask_due = df["nextFollowUpDate"] <= today_str
        
        # EXCLUDE leads already contacted today
        # Ensure lastFollowUpDate is handled safely
        if "lastFollowUpDate" not in df.columns:
            df["lastFollowUpDate"] = None
        df["lastFollowUpDate"] = df["lastFollowUpDate"].astype(str).replace("None", "").replace("nan", "")
        
        mask_not_contacted_today = df["lastFollowUpDate"] != today_str
        
        # EXCLUDE Empty Dates (String comparison "" <= "2024-.." is True, so we must explicitly filter them out)
        mask_has_date = df["nextFollowUpDate"] != ""

        filtered_df = df[mask_active & mask_due & mask_not_contacted_today & mask_has_date]
        
        # SORT: Oldest First (Overdue -> Today)
        filtered_df = filtered_df.sort_values(by="nextFollowUpDate", ascending=True)
        
        # --- NEW: QUEUE VIEW vs DIALER VIEW LOGIC ---
        
        # 1. Check if we are in "List Mode" or "Active Call Mode"
        # We use a session state variable to track the actively selected lead from the list
        if "pd_active_lead_id" not in st.session_state:
            st.session_state.pd_active_lead_id = None
            
        # If filtered list is empty, show empty state
        if filtered_df.empty:
            # Styled Empty State (Existing Code)
            is_dark = st.session_state.get("theme", "light") == "dark"
            bg_col = "#1E1E1E" if is_dark else "#F8F9FA"
            border_col = "#333" if is_dark else "#DDD"
            text_col = "#EEE" if is_dark else "#333"
            
            st.markdown(f"""
            <div style="
                text-align: center;
                padding: 40px;
                background-color: {bg_col};
                border: 1px dashed {border_col};
                border-radius: 12px;
                margin: 20px 0;
            ">
                <div style="font-size: 3rem; margin-bottom: 10px;">🎉</div>
                <h2 style="color: {text_col}; font-size: 1.5rem; margin-bottom: 8px;">You're all caught up!</h2>
                <p style="color: #888; margin-bottom: 24px;">No remaining follow-ups scheduled for today.</p>
            </div>
            """, unsafe_allow_html=True)
            
            c1, c2, c3 = st.columns([1, 2, 1])
            with c2:
                def switch_to_all_leads():
                    st.session_state.pd_filter_choice = "All Leads"
                st.button("📂 Browse All Leads", type="primary", use_container_width=True, on_click=switch_to_all_leads)
            st.stop()
            
        # If we have leads, decide what to show



        if st.session_state.pd_active_lead_id is None:
            # --- SHOW THE LIST (QUEUE VIEW) ---
            st.markdown("### 📋 Today's Call Queue")
            
            # Custom CSS for the Cards Grid & Clickable Overlay
            # IMPORTANT: We use unindented string content to prevent Markdown from treating it as a code block.
            st.markdown("""<style>
@import url('https://fonts.googleapis.com/css?family=Inter:400,600,700&display=swap');
.queue-grid-wrapper {
    display: grid;
    grid-template-columns: repeat(3, 1fr) !important;
    gap: 20px;
    padding: 10px 0;
}
.queue-card-container {
    position: relative;
    border: 1px solid rgba(0,0,0,0.05);
    border-radius: 12px;
    padding: 18px;
    color: #333;
    background: white;
    box-shadow: 0 4px 6px rgba(0,0,0,0.04);
    transition: transform 0.1s ease, box-shadow 0.2s ease;
    height: 190px !important;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    overflow: hidden;
    pointer-events: none; 
}


.qc-header { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 12px; }
.qc-date-badge {
    background: rgba(255,255,255,0.5);
    border-radius: 8px;
    padding: 6px 10px;
    text-align: center;
    border: 1px solid rgba(255,255,255,0.4);
    backdrop-filter: blur(4px);
}
.qc-day-num { font-size: 1.3rem; font-weight: 800; line-height: 1; color: #333; }
.qc-day-name { font-size: 0.75rem; font-weight: 700; text-transform: uppercase; color: #555; margin-top: 2px; }
.qc-title {
    font-family: 'Inter', sans-serif;
    font-size: 1.1rem;
    font-weight: 700;
    color: #2c3e50;
    margin-bottom: 8px;
    line-height: 1.35;
    word-wrap: break-word;
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
}
.qc-details {
    font-family: 'Inter', sans-serif;
    font-size: 0.9rem;
    color: #4a5568;
    display: flex;
    flex-direction: column;
    gap: 6px;
}
.qc-row { display: flex; align-items: center; gap: 8px; }








div[data-testid="stElementContainer"]:has(.queue-card-container) + div[data-testid="stElementContainer"]:has(.stButton),
div[data-testid="element-container"]:has(.queue-card-container) + div[data-testid="element-container"]:has(.stButton) {
    margin-top: -206px !important;
    height: 190px !important;
    z-index: 10 !important;
    position: relative !important;
    pointer-events: auto !important;
    margin-bottom: 24px !important; 
}


div[data-testid="stElementContainer"]:has(.queue-card-container) + div[data-testid="stElementContainer"]:has(.stButton) .stButton,
div[data-testid="element-container"]:has(.queue-card-container) + div[data-testid="element-container"]:has(.stButton) .stButton {
    width: 100% !important;
    height: 100% !important;
    min-height: 190px !important;
}

div[data-testid="stElementContainer"]:has(.queue-card-container) + div[data-testid="stElementContainer"]:has(.stButton) .stButton button,
div[data-testid="element-container"]:has(.queue-card-container) + div[data-testid="element-container"]:has(.stButton) .stButton button {
    width: 100% !important;
    height: 100% !important;
    border: none !important;
    background: transparent !important;
    color: transparent !important;
    cursor: pointer !important;
    padding: 0 !important;
    margin: 0 !important;
    display: block !important;
}


div[data-testid="stElementContainer"]:has(.queue-card-container) + div[data-testid="stElementContainer"]:has(.stButton) .stButton button:hover,
div[data-testid="stElementContainer"]:has(.queue-card-container) + div[data-testid="stElementContainer"]:has(.stButton) .stButton button:focus,
div[data-testid="stElementContainer"]:has(.queue-card-container) + div[data-testid="stElementContainer"]:has(.stButton) .stButton button:active,
div[data-testid="element-container"]:has(.queue-card-container) + div[data-testid="element-container"]:has(.stButton) .stButton button:hover,
div[data-testid="element-container"]:has(.queue-card-container) + div[data-testid="element-container"]:has(.stButton) .stButton button:focus,
div[data-testid="element-container"]:has(.queue-card-container) + div[data-testid="element-container"]:has(.stButton) .stButton button:active {
    background: transparent !important;
    color: transparent !important;
    border: none !important;
    box-shadow: none !important;
}


.queue-card-container {
    pointer-events: none;
    height: 190px !important; 
    margin-bottom: 0px !important; 
}



div[data-testid="stElementContainer"]:has(.queue-card-container):has(+ div[data-testid="stElementContainer"] .stButton button:active) .queue-card-container,
div[data-testid="element-container"]:has(.queue-card-container):has(+ div[data-testid="element-container"] .stButton button:active) .queue-card-container {
    transform: scale(0.98);
    transition: transform 0.1s ease;
}
</style>""", unsafe_allow_html=True)

            # Gradients
            gradients = [
                "linear-gradient(135deg, #E3F2FD 0%, #90CAF9 100%)", # Blue
                "linear-gradient(135deg, #FFF3E0 0%, #FFCC80 100%)", # Orange
                "linear-gradient(135deg, #E8F5E9 0%, #A5D6A7 100%)", # Green
                "linear-gradient(135deg, #F3E5F5 0%, #CE93D8 100%)", # Purple
                "linear-gradient(135deg, #FAFAFA 0%, #E0E0E0 100%)"  # Grey
            ]

            # Icons
            import base64
            def get_b64_icon(path):
                try:
                    with open(path, "rb") as f:
                        return base64.b64encode(f.read()).decode()
                except: return ""

            icon_phone_path = '/Users/satyajeetsinhrathod/.gemini/antigravity/brain/dc0185d0-ed71-439f-bab7-e65957b15690/uploaded_image_2_1768732438815.png'
            phone_icon_b64 = get_b64_icon(icon_phone_path)
            phone_icon_html = f'<img src="data:image/png;base64,{phone_icon_b64}" width="16" style="vertical-align:middle;">' if phone_icon_b64 else '📞'
            
            # --- RENDER GRID ---
            # We iterate in chunks of 3 to create rows
            cols = st.columns(3)
            
            for i, (idx, row) in enumerate(filtered_df.iterrows()):
                col_idx = i % 3
                with cols[col_idx]:
                    # USE CONTAINER TO ISOLATE EACH CARD CONTEXT
                    with st.container():
                        bg_gradient = gradients[i % len(gradients)]
                        
                        # Date Logic
                        try:
                            d_obj = pd.to_datetime(row.get('nextFollowUpDate'))
                            day_num = d_obj.strftime("%d")
                            day_name = d_obj.strftime("%b").upper()
                        except:
                            day_num = "--"
                            day_name = "TODO"
                        
                        # 1. Render the Visual Card (Markdown)
                        # We ensure the markdown takes up the fixed height
                        card_html = f"""
                        <div class="queue-card-container" style="background: {bg_gradient};">
                            <div class="qc-header">
                                <div class="qc-date-badge">
                                    <div class="qc-day-num">{day_num}</div>
                                    <div class="qc-day-name">{day_name}</div>
                                </div>
                                <!-- Index removed as requested -->
                            </div>
                            <div class="qc-title">{row.get('businessName', 'Unknown')}</div>
                            <div class="qc-details">
                                <div class="qc-row">👤 {row.get('contactName', 'No Name')}</div>
                                <div class="qc-row">{phone_icon_html} {row.get('phone', 'No Phone')}</div>
                            </div>
                        </div>
                        """
                        st.markdown(card_html, unsafe_allow_html=True)
                        
                        # 2. Render the Button OVERLAY
                        # The CSS now targets this button using the Sibling Selector (+ div:has(.stButton))
                        # It will be pulled UP by 206px to cover the card.
                        if st.button(" ", key=f"btn_card_{row['id']}", use_container_width=True):
                             st.session_state.pd_active_lead_id = row['id']
                             st.session_state.pd_jump_to_id = row['id'] # Signal to force-update the index
                             st.rerun()

            st.stop()

        else:
            # --- ACTIVE CALL MODE (QUEUE NAV) ---
            # We want to iterate through the filtered queue, not just see one record.
            
            # 1. Use the Queue subset (filtered_df) as our working dataframe
            if not filtered_df.empty:
                df = filtered_df.reset_index(drop=True)
            else:
                # Fallback if queue empty
                st.session_state.pd_active_lead_id = None
                st.rerun()

            # 2. Handle "Jump To" from Card Click
            if "pd_jump_to_id" in st.session_state and st.session_state.pd_jump_to_id is not None:
                # Find the index of the requested lead in our new 'df'
                jump_id = st.session_state.pd_jump_to_id
                idx_list = df.index[df['id'] == jump_id].tolist()
                
                if idx_list:
                    st.session_state.dialer_index = idx_list[0]
                else:
                    # Lead might have disappeared from filter (e.g. status changed elsewhere)
                    st.toast("Lead not found in current queue.")
                    st.session_state.dialer_index = 0
                
                # Clear the trigger so we don't keep forcing index on next run
                st.session_state.pd_jump_to_id = None

            # 3. Back Button to return to Grid View
            if st.button("⬅ Back to Queue"):
                st.session_state.pd_active_lead_id = None
                st.rerun()

        # The existing code continues below, rendering the 'Main Lead Card' using 'df' and 'dialer_index'
        # which we have now correctly set up.
        
        # We assign to df, so the logic falls through
        pass # proceed to rendering
        
    else:
        # For "All Leads", we reset the active lead state to ensure standard navigation
        st.session_state.pd_active_lead_id = None
        df = df
        st.success(f"Showing all {len(df)} leads.")
    
    # ... Existing pagination logic below ...

    # Session State for Current Index
    if "dialer_index" not in st.session_state:
        st.session_state.dialer_index = 0
        
    # Validation
    if st.session_state.dialer_index >= len(df):
        st.session_state.dialer_index = 0
        
    lead = df.iloc[st.session_state.dialer_index]
    
    # --- PREMIUM LAYOUT RESTRUCTURE ---
    
    # Custom CSS for Cards (Dynamic Theme Support)
    st.markdown("""
    <style>
    .style_card {
        background-color: white;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.05);
        margin-bottom: 20px;
        border: 1px solid #f0f0f0;
    }
    .dark-mode-card {
        background-color: #1E1E1E !important;
        border: 1px solid #333 !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.2) !important;
        color: #e0e0e0 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # 1. Progress Bar Section (Full Width)
    current_theme = st.session_state.get("theme", "light") # 'light' or 'dark'
    # Check boolean toggle for Google Mode as well
    is_google_mode = st.session_state.get("google_ui_mode", False)
    
    # Determine CSS classes based on effective theme
    card_bg = "#1E1E1E" if (current_theme == "dark" and not is_google_mode) else "white"
    card_border = "#333" if (current_theme == "dark" and not is_google_mode) else "#f0f0f0"
    card_shadow = "0 4px 12px rgba(0,0,0,0.2)" if (current_theme == "dark" and not is_google_mode) else "0 4px 12px rgba(0,0,0,0.05)"
    
    def get_card_style(key):
        return f"""<style>
div[data-testid="stVerticalBlock"]:has(div#{key}) {{
    background-color: {card_bg};
    border-radius: 12px;
    padding: 20px;
    box-shadow: {card_shadow};
    border: 1px solid {card_border};
    gap: 10px;
}}
</style>
<div id="{key}" style="display:none;"></div>"""

    card_class = "style_card"
    if current_theme == "dark" and not is_google_mode:
        card_class += " dark-mode-card"
    
    progress_val = st.session_state.dialer_index + 1
    total_leads = len(df)
    prog_percent = int((progress_val / total_leads) * 100) if total_leads > 0 else 0
    
    # Colors
    prog_color = "#1976d2" if not (current_theme == "dark" and not is_google_mode) else "#90caf9"
    text_main = "#333" if not (current_theme == "dark" and not is_google_mode) else "#fff"
    text_color = text_main # Alias for use in f-strings
    text_sub = "#888" if not (current_theme == "dark" and not is_google_mode) else "#aaa"

    with st.container():
        st.markdown(f"""
        <div class="{card_class}">
            <div style="display: flex; justify-content: space-between; align-items: flex-end; margin-bottom: 8px;">
                <div style="font-weight: 600; font-size: 1.1rem; display: flex; align-items: center; gap: 8px; color: {text_main};">
                    <span style="color: {prog_color};">📊</span> Lead {progress_val} of {total_leads}
                </div>
                <div style="font-size: 0.85rem; color: {text_sub};">
                    <span style="margin-right: 15px;">Completed <b style="color: {prog_color}; font-size:1.1rem;">{st.session_state.dialer_index}</b></span>
                    <span>Remaining <b style="color: {text_main}; font-size:1.1rem;">{total_leads - st.session_state.dialer_index}</b></span>
                </div>
            </div>
            <div style="font-size: 0.8rem; color: {text_sub}; margin-bottom: 8px;">{prog_percent}% Complete</div>
            <div style="width: 100%; background-color: {'#f0f0f0' if not (current_theme == 'dark' and not is_google_mode) else '#333'}; height: 6px; border-radius: 4px; overflow: hidden;">
                <div style="width: {prog_percent}%; background-color: {prog_color}; height: 100%;"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # 2. Main 2-Column Layout
    main_col1, main_col2 = st.columns([2.2, 1])

    # --- LEFT COLUMN ---
    with main_col1:
        # A. Lead Header Card
        st.markdown(f"""
        <div class="{card_class}">
            <div style="display: flex; align-items: center; gap: 15px;">
                <div style="font-size: 2.5rem;">🏢</div>
                <div>
                    <h2 style="margin: 0; font-size: 1.6rem; font-weight: 700; color: {text_main};">{lead.get('businessName', 'Unknown')}</h2>
                    <div style="color: {text_sub}; margin-top: 4px;">{lead.get('contactName', 'None')}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
        # B. Lead Information Card
        with st.container():
            st.markdown(get_card_style("lead_info_card"), unsafe_allow_html=True)
            st.markdown(f"<h3 style='color:{text_main}'>Lead Information</h3>", unsafe_allow_html=True)
            
            # Status & Priority Row
            c_info1, c_info2 = st.columns(2)
            with c_info1:
                st.caption("STATUS")
                # Custom Badge Style
                st.markdown(f"""
                <div style="
                    background-color: #dff5e1; 
                    color: #1b5e20; 
                    padding: 4px 12px; 
                    border-radius: 16px; 
                    display: inline-block; 
                    font-weight: 600; 
                    font-size: 0.9rem;">
                    {lead.get('status', 'New')}
                </div>
                """, unsafe_allow_html=True)

            with c_info2:
                # Priority Key Logic
                prio_key = f"prio_edit_{lead['id']}"
                current_prio = lead.get('priority', 'WARM')
                
                # Sync state if needed
                if prio_key not in st.session_state:
                    st.session_state[prio_key] = current_prio
                
                # Priority Dropdown Options
                prio_opts = ["HOT", "WARM", "COLD"]
                if current_prio not in prio_opts: prio_opts.append(current_prio)
                
                # Display Mapping (Emojis for colored dots)
                prio_map = {
                    "HOT": "🔴 HOT",
                    "WARM": "🟠 WARM",
                    "COLD": "🔵 COLD"
                }

                # Custom Styled Label (Matches Image 2)
                st.markdown("""
                    <div style="font-size: 0.75rem; color: #9CA3AF; text-transform: uppercase; letter-spacing: 0.05em; font-weight: 600; margin-bottom: 4px;">
                        PRIORITY
                    </div>
                """, unsafe_allow_html=True)
                
                # Selectbox with collapsed label and custom formatting
                new_prio = st.selectbox(
                    "Priority_Select", 
                    prio_opts, 
                    format_func=lambda x: prio_map.get(x, x),
                    key=prio_key,
                    label_visibility="collapsed"
                )
                
                # Update if changed
                if new_prio != current_prio:
                    update_lead(lead['id'], {"priority": new_prio})
                    st.toast(f"Priority updated to {new_prio}")
                    time.sleep(0.5)
                    st.rerun()

            st.markdown("---")
        
            # Contact Details List (HTML representation)
            def icon_row(icon, label, value):
                theme_text = "#eee" if (current_theme == "dark" and not is_google_mode) else "#333"
                theme_sub = "#aaa" if (current_theme == "dark" and not is_google_mode) else "#999"
                return f"""
                <div style="display: flex; gap: 15px; margin-bottom: 15px; align-items: flex-start;">
                    <div style="min-width: 24px; text-align: center; font-size: 1.2rem; color: #666;">{icon}</div>
                    <div>
                        <div style="font-size: 0.75rem; color: {theme_sub}; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom:2px;">{label}</div>
                        <div style="font-size: 1rem; color: {theme_text}; font-weight: 500;">{value}</div>
                    </div>
                </div>
                """
            
            contact_html = ""
            contact_html += icon_row("📍", "Address", lead.get('address', 'N/A'))
            
            # Phone with copy style
            # Phone with copy style
            phone_val = str(lead.get('phone', 'N/A'))
            if phone_val and phone_val not in ['N/A', 'nan', 'None', '']:
                phone_content = f"<a href='tel:{phone_val}' style='color: inherit; text-decoration: none;'>{phone_val}</a>"
            else:
                phone_content = phone_val
                
            phone_box = f"<span style='background: {'#333' if (current_theme == 'dark' and not is_google_mode) else '#f5f5f5'}; padding: 4px 8px; border-radius: 4px; font-family: monospace;'>{phone_content}</span>"
            contact_html += icon_row("📞", "Phone", phone_box)
            
            contact_html += icon_row("✉️", "Email", lead.get('email', 'N/A'))
            
            st.markdown(contact_html, unsafe_allow_html=True)
            # End Card

        # C. Call Notes Card
        with st.container():
            st.markdown(get_card_style("call_notes_card"), unsafe_allow_html=True)
            st.markdown(f"<h3 style='color:{text_main}'>📝 Call Notes</h3>", unsafe_allow_html=True)
            
            notes_key = f"notes_{lead['id']}"
            if notes_key not in st.session_state:
                st.session_state[notes_key] = lead.get('callNotes', '')
                
            new_notes = st.text_area(
                "Add notes about this call",
                value=st.session_state[notes_key],
                height=150,
                key=f"widget_{notes_key}",
                placeholder="Type call details here..."
            )
            # Sync back to session state
            st.session_state[notes_key] = new_notes

    # --- RIGHT COLUMN (Action Panel) ---
    with main_col2:
        # D. Quick Actions Card
        with st.container():
            # Inject Card Style + Special Button Style for Not Interested
            st.markdown(get_card_style("actions_card") + """
<style>

div[data-testid="stVerticalBlock"]:has(div#actions_card) div.stButton:nth-of-type(5) button {
    border-color: #ef4444 !important;
    color: #ef4444 !important;
}
div[data-testid="stVerticalBlock"]:has(div#actions_card) div.stButton:nth-of-type(5) button:hover {
    background-color: #fef2f2 !important;
    border-color: #dc2626 !important;
}
</style>
""", unsafe_allow_html=True)
            
            st.markdown(f"<h3 style='color:{text_main}'>⚡ Quick Actions</h3>", unsafe_allow_html=True)
            
            today_date = datetime.now().strftime("%Y-%m-%d")
            
            # Action Helper
            def next_lead():
                if filter_choice == "All Leads":
                    st.session_state.dialer_index += 1
                # If 'Today', rerun handles list update
            
            # 1. Interested (Primary)
            if st.button("✅ Interested", type="primary", use_container_width=True):
                update_lead(lead['id'], {
                    "status": "Interested", "priority": "HOT", 
                    "callNotes": st.session_state[notes_key], "lastFollowUpDate": today_date
                })
                st.toast("Marked Interested! 🚀")
                next_lead()
                time.sleep(0.5)
                st.rerun()
                
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

            # 2. Meeting Set (Primary)
            if st.button("📅 Meeting Set", type="primary", use_container_width=True):
                st.session_state[f"open_meet_{lead['id']}"] = True
            
            if st.session_state.get(f"open_meet_{lead['id']}", False):
                @st.dialog("📅 Schedule Meeting")
                def show_meeting_modal():
                    st.write(f"Book meeting with {lead.get('businessName')}")
                    c1, c2 = st.columns(2)
                    d = c1.date_input("Date", value=datetime.now().date() + pd.Timedelta(days=1))
                    t = c2.time_input("Time", value=datetime.now().time())
                    if st.button("Confirm", type="primary", use_container_width=True):
                        ts = pd.Timestamp(datetime.combine(d, t))
                        update_lead(lead['id'], {
                            "status": "Meeting set", "priority": "HOT",
                            "callNotes": st.session_state[notes_key], "lastFollowUpDate": today_date,
                            "meetingDate": str(ts), "nextFollowUpDate": str(d)
                        })
                        st.success("Meeting Scheduled!")
                        del st.session_state[f"open_meet_{lead['id']}"]
                        next_lead()
                        st.rerun()
                show_meeting_modal()

            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

            # 3. Not Picking (Standard)
            if st.button("🚫 Not Picking", use_container_width=True):
                next_d_val = st.session_state.get(f"next_action_date_{lead['id']}", datetime.now().date() + pd.Timedelta(days=1))
                update_lead(lead['id'], {
                    "status": "Not picking", "callNotes": st.session_state[notes_key],
                    "lastFollowUpDate": today_date, "nextFollowUpDate": str(next_d_val)
                })
                st.toast(f"Marked Not Picking (Next: {next_d_val})")
                next_lead()
                time.sleep(0.3)
                st.rerun()
            
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

            # 4. Call Later
            if st.button("⏰ Call Later", use_container_width=True):
                next_d_val = st.session_state.get(f"next_action_date_{lead['id']}", datetime.now().date() + pd.Timedelta(days=1))
                update_lead(lead['id'], {
                    "status": "Call Later", "callNotes": st.session_state[notes_key],
                    "lastFollowUpDate": today_date, "nextFollowUpDate": str(next_d_val)
                })
                st.toast(f"Snoozed until {next_d_val}")
                next_lead()
                time.sleep(0.3)
                st.rerun()

            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            
            # 5. Not Interested
            if st.button("❌ Not Interested", use_container_width=True):
                update_lead(lead['id'], {
                    "status": "Not Interested", "priority": "COLD",
                    "callNotes": st.session_state[notes_key], "lastFollowUpDate": today_date
                })
                st.toast("Marked Not Interested")
                next_lead()
                time.sleep(0.3)
                st.rerun()



        # E. Next Follow-Up (Nav) Card
        with st.container():
            st.markdown(get_card_style("next_step_card"), unsafe_allow_html=True)
            st.markdown(f"<h3 style='color:{text_main}'>🗓 Next Follow-Up</h3>", unsafe_allow_html=True)
            
            next_date_key = f"next_action_date_{lead['id']}"
            
            # Default to tomorrow
            default_d = datetime.now().date() + pd.Timedelta(days=1)
            f_next = lead.get('nextFollowUpDate')
            if f_next:
                try:
                    parsed_d = pd.to_datetime(f_next).date()
                    default_d = max(parsed_d, datetime.now().date())
                except: pass

            def on_next_date_change():
                new_val = st.session_state[next_date_key]
                if new_val:
                    update_lead(lead['id'], {"nextFollowUpDate": str(new_val)})
                    st.toast(f"Updated Next Follow-Up: {new_val}")

            st.date_input(
                "Select Date",
                value=default_d,
                min_value=datetime.now().date(),
                key=next_date_key,
                label_visibility="collapsed",
                on_change=on_next_date_change
            )

            st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
            
            # Navigation Buttons
            c_nav1, c_nav2 = st.columns(2)
            with c_nav1:
                # Previous Button (Secondary/Outline style if possible via CSS, using standard here)
                if st.button("Previous", use_container_width=True, disabled=(st.session_state.dialer_index == 0)):
                    st.session_state.dialer_index = max(0, st.session_state.dialer_index - 1)
                    st.rerun()
            with c_nav2:
                # Skip Button
                if st.button("Skip", use_container_width=True):
                    st.session_state.dialer_index += 1
                    st.rerun()
                    




        



# ================== LEAD GENERATOR (COMMENTED OUT) ==================
# if "Lead Generator" in page:
#     st.markdown(f"""<h1 style='display: flex; align-items: center;'>{get_icon_html('rocket.png', 70)} Lead Generator</h1>""", unsafe_allow_html=True)
#     
#     with st.form("gen_form"):
#         c1, c2 = st.columns(2)
#         query = c1.text_input("Business Type", "Dentist")
#         loc = c2.text_input("Location", "New York")
#         submitted = st.form_submit_button("Start Mining Leads", type="primary")
#         
#     if submitted:
#         with st.spinner("Disconnecting from matrix... (Scraping)"):
#             try:
#                 r = requests.post(LEAD_GEN_API, json={"query": query, "location": loc}, timeout=120)
#                 if r.status_code == 200:
#                     st.success("✅ Leads Generated & Saved to CRM!")
#                     st.download_button("Download CSV", r.content, f"{query}.csv", "text/csv")
#                 else:
#                     st.error("backend error")
#             except Exception as e:
#                 st.error(f"Error: {e}")
#                 
#     # History moved to Sidebar

# ================== GOOGLE MAPS SCRAPER ==================
if "Google Maps Scraper" in page:

    
    # INIT SESSION STATE
    if 'scraper_running' not in st.session_state:
        st.session_state.scraper_running = False
        
    # --- STATE MANAGEMENT BRIDGE REMOVED TO FIX CHECKBOX VISIBILITY ---
    # The previous CSS here was hiding ALL checkboxes globally.
    
    if 'google_ui_mode' not in st.session_state:
        st.session_state.google_ui_mode = True

    # Hidden Bridge Checkbox
    # The HTML/JS toggle below clicks this invisible checkbox to trigger Python state changes
    # Use a span with a specific ID inside the label to make it easily targetable by JS
    # Hidden Bridge Checkbox - Uses a unique ALT text marker for robust JS targeting
    google_mode = st.checkbox("![bridge_marker](http://#bridge)", key="google_ui_mode")
    
    # Hide the bridge checkbox
    st.markdown("""
        <style>
            
            div[data-testid="stCheckbox"]:has(img[alt="bridge_marker"]) {
                display: none !important;
            }
        </style>
        <script>
            function hideCheckbox() {
                try {
                    // Try both current frame and parent frame
                    const contexts = [document, window.parent.document];
                    
                    contexts.forEach(doc => {
                        try {
                            const marker = doc.querySelector('img[alt="bridge_marker"]');
                            if (marker) {
                                const widget = marker.closest('div[data-testid="stCheckbox"]');
                                if (widget) {
                                    widget.style.display = 'none';
                                }
                            }
                        } catch(err) {}
                    });
                } catch(e) {
                    console.error("Failed to hide Google UI Mode checkbox:", e);
                }
            }
            // Run immediately and periodically
            hideCheckbox();
            setInterval(hideCheckbox, 500); // Check every 500ms to ensure it stays hidden
        </script>
    """, unsafe_allow_html=True)

    # --- SHARED TRANSITION CSS (applies to both themes for smooth switching) ---
    st.markdown("""
    <style>
    /* Smooth theme transitions */
    .stApp,
    [data-testid="stAppViewContainer"],
    div[data-testid="stTextInput"] input,
    div[data-testid="stTextInput"] label,
    div.stButton > button,
    div[data-testid="stDataFrame"],
    div[data-testid="stForm"],
    div[data-testid="stAlert"],
    div[data-testid="stSuccess"],
    div[data-testid="stInfo"],
    div[data-testid="stWarning"],
    div[data-testid="stError"],
    .helper-text,
    .loading-container,
    .loading-text,
    .loading-subtext,
    thead tr th,
    tbody tr td,
    div[data-testid="stSelectbox"] > div > div,
    div[data-testid="stMarkdownContainer"] p,
    div[data-testid="stMarkdownContainer"] h1,
    div[data-testid="stMarkdownContainer"] h2,
    div[data-testid="stMarkdownContainer"] h3,
    a,
    .g-card-header,
    .g-card-sub,
    .g-section-header {
        transition: background-color 0.4s cubic-bezier(0.4, 0, 0.2, 1),
                    color 0.4s cubic-bezier(0.4, 0, 0.2, 1),
                    border-color 0.4s cubic-bezier(0.4, 0, 0.2, 1),
                    box-shadow 0.4s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # --- CONDITIONAL STYLES ---
    if google_mode:
        # GOOGLE THEME (Blue/Colorful + Global Overrides)
        st.markdown("""
        <style>
        
        @import url('https://fonts.googleapis.com/css?family=Product+Sans:400,500,700&display=swap');
        
        html, body, [class*="css"] {
            font-family: 'Product Sans', 'Roboto', Arial, sans-serif !important;
        }
        
        .stApp {
            background-color: #ffffff !important; 
        }

        [data-testid="stAppViewContainer"] {
            background: #ffffff !important;
        }
        
        div[data-testid="stDataFrame"] {
            border: 1px solid #e0e0e0 !important;
            border-radius: 8px !important;
        }
        div[data-testid="stDataFrame"] table {
            font-family: 'Roboto', sans-serif !important;
        }
        thead tr th {
            background-color: #f1f3f4 !important; 
            color: #202124 !important;
            font-weight: 600 !important;
            border-bottom: 2px solid #1a73e8 !important;
        }
        tbody tr td {
            color: #3c4043 !important;
        }
        tbody tr:hover {
            background-color: #f8f9fa !important;
        }
        
        div[data-testid="stAlert"] {
            padding: 0.5rem 1rem !important;
            border-radius: 8px !important;
            border: 1px solid transparent !important;
            background-color: #fce8e6 !important; 
        }
        
        div[data-testid="stMain"] div.stButton > button,
        .main .block-container div.stButton > button[type="primary"],
        div[data-testid="column"] div.stButton > button[type="primary"],
        div.stButton > button[type="primary"] {
            background: #1a73e8 !important;
            background-color: #1a73e8 !important;
            background-image: none !important;
            color: white !important;
            border-radius: 24px !important;
            border: none !important;
            box-shadow: 0 1px 2px rgba(60,64,67,0.3), 0 2px 6px 2px rgba(60,64,67,0.15) !important;
            font-family: 'Product Sans', sans-serif !important;
            padding: 0 24px !important;
            font-size: 14px !important;
            font-weight: 500 !important;
            height: 48px !important;
            text-transform: none !important;
            letter-spacing: 0.25px !important;
            transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1) !important;
        }
        div[data-testid="stMain"] div.stButton > button:hover,
        .main .block-container div.stButton > button[type="primary"]:hover,
        div[data-testid="column"] div.stButton > button[type="primary"]:hover,
        div.stButton > button[type="primary"]:hover {
            background: #1765cc !important;
            background-color: #1765cc !important;
            box-shadow: 0 1px 3px 1px rgba(60,64,67,0.15), 0 1px 2px 0 rgba(60,64,67,0.3) !important;
            transform: translateY(-1px);
        }
        div.stButton > button[type="primary"]:active {
            box-shadow: 0 1px 2px 0 rgba(60,64,67,0.3) !important;
            transform: translateY(0);
        }
        
        div[data-testid="stTextInput"] input {
            border-radius: 8px !important;
            border: 1px solid #dadce0 !important;
            background: #fff !important;
            color: #202124 !important;
            padding: 12px 16px !important;
            font-size: 14px !important;
            font-family: 'Roboto', sans-serif !important;
            transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1) !important;
        }
        div[data-testid="stTextInput"] input:focus {
            border: 2px solid #1a73e8 !important;
            padding: 11px 15px !important;
            box-shadow: 0 0 0 4px rgba(26, 115, 232, 0.15) !important;
            transform: translateY(-1px);
        }
        div[data-testid="stTextInput"] label {
            color: #202124 !important;
            font-weight: 500 !important;
            font-size: 13px !important;
            font-family: 'Google Sans', Roboto, sans-serif !important;
        }

        /* Material Card - Google colored top border */
        div[data-testid="stForm"] {
            background: #ffffff;
            border-radius: 12px;
            padding: 24px;
            box-shadow: 0 1px 2px 0 rgba(60,64,67,0.3), 0 1px 3px 1px rgba(60,64,67,0.15);
            margin-bottom: 24px;
            position: relative;
            transition: all 0.4s cubic-bezier(0.165, 0.84, 0.44, 1);
            border: none !important;
        }
        div[data-testid="stForm"]::before {
            content: "";
            position: absolute;
            top: 0;
            left: 0;
            width: 100%;
            height: 4px;
            border-top-left-radius: 12px;
            border-top-right-radius: 12px;
            background: linear-gradient(90deg, #1a73e8 0%, #ea4335 25%, #fbbc05 50%, #34a853 75%, #1a73e8 100%);
            background-size: 200% auto;
            animation: googleBorderMove 3s linear infinite;
        }
        @keyframes googleBorderMove {
            0% { background-position: 0% 0; }
            100% { background-position: 200% 0; }
        }
        div[data-testid="stForm"]:hover {
            box-shadow: 0 4px 12px 0 rgba(60,64,67,0.2), 0 2px 6px 1px rgba(60,64,67,0.1);
            transform: translateY(-3px) scale(1.005);
        }

        /* Links */
        a { color: #1a73e8 !important; text-decoration: none !important; }
        a:hover { text-decoration: underline !important; }
        div[data-testid="stDataFrame"] a { color: #1a73e8 !important; font-weight: 500 !important; }

        /* Alerts */
        div[data-testid="stSuccess"] { background-color: #e6f4ea !important; border-left: 4px solid #34a853 !important; color: #137333 !important; }
        div[data-testid="stInfo"] { background-color: #e8f0fe !important; border-left: 4px solid #1a73e8 !important; color: #174ea6 !important; }
        div[data-testid="stWarning"] { background-color: #fef7e0 !important; border-left: 4px solid #fbbc05 !important; color: #b06000 !important; }
        div[data-testid="stError"] { background-color: #fce8e6 !important; border-left: 4px solid #ea4335 !important; color: #c5221f !important; }

        /* Progress */
        div[data-testid="stProgress"] > div > div { background-color: #1a73e8 !important; }

        /* Helper text */
        .helper-text { color: #5f6368; font-size: 13px; line-height: 1.6; margin: 16px 0; font-family: 'Roboto', sans-serif; }

        /* Loading */
        .loading-container { background: #ffffff; border-radius: 8px; padding: 24px; text-align: center; margin: 24px 0; box-shadow: 0 1px 2px 0 rgba(60,64,67,.3), 0 1px 3px 1px rgba(60,64,67,.15); border-left: 4px solid #1a73e8; }
        .loading-text { font-size: 16px; font-weight: 500; color: #202124; margin-top: 12px; font-family: 'Google Sans', 'Roboto', sans-serif; }
        .loading-subtext { color: #5f6368; font-size: 13px; margin-top: 8px; }

        /* Selectbox */
        div[data-testid="stSelectbox"] > div > div:focus-within { border-color: #1a73e8 !important; box-shadow: 0 0 0 2px rgba(26, 115, 232, 0.2) !important; }

        /* Animations */
        @keyframes mining { 0% { transform: rotate(0deg) translateY(0); } 25% { transform: rotate(-45deg) translateY(-2px); } 50% { transform: rotate(0deg) translateY(0); } 75% { transform: rotate(-45deg) translateY(-2px); } 100% { transform: rotate(0deg) translateY(0); } }
        .miner-icon { font-size: 48px; display: inline-block; animation: mining 1.2s infinite ease-in-out; margin-bottom: 16px; }
        .tech-lines { height: 4px; width: 100%; background: linear-gradient(to right, #e8f0fe 0%, #1a73e8 50%, #e8f0fe 100%); background-size: 200% 100%; animation: shimmer 1.5s infinite linear; border-radius: 2px; margin-top: 16px; }
        @keyframes shimmer { 0% { background-position: 200% 0; } 100% { background-position: -200% 0; } }

        /* Typography classes */
        .g-card-header { font-family: 'Google Sans', Roboto, sans-serif; font-size: 20px; color: #202124; font-weight: 400; line-height: 1.3; }
        .g-card-sub { font-family: Roboto, sans-serif; font-size: 14px; color: #5f6368; margin-top: 4px; }
        .g-section-header { font-family: 'Google Sans', Roboto, sans-serif; font-size: 11px; color: #5f6368; font-weight: 500; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 0.5px; }

        /* Max width */
        .main .block-container { max-width: 1000px !important; padding-top: 1rem !important; }

        </style>
        """, unsafe_allow_html=True)
        
        gradient_bar = "linear-gradient(to right, #4285F4 25%, #EA4335 25%, #EA4335 50%, #FBBC05 50%, #FBBC05 75%, #34A853 75%)"
        
    else:
        # ═══════════════════════════════════════════════════════════════
        # NORMAL THEME (Toggle OFF) — Basic light gray
        # ═══════════════════════════════════════════════════════════════
        st.markdown("""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
        
        html, body, [class*="css"] {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
        }

        /* ── Basic light background ── */
        .stApp {
            background-color: #f8f9fa !important;
        }

        div[data-testid="stForm"] {
            background: #ffffff !important;
            border-radius: 12px;
            padding: 24px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            margin-bottom: 24px;
            border-top: 4px solid #e0e0e0 !important;
            border-left: 1px solid #e0e0e0 !important;
            border-right: 1px solid #e0e0e0 !important;
            border-bottom: 1px solid #e0e0e0 !important;
            transition: all 0.4s cubic-bezier(0.165, 0.84, 0.44, 1);
        }
        div[data-testid="stForm"]:hover {
            box-shadow: 0 8px 16px rgba(0,0,0,0.1);
            transform: translateY(-3px);
        }

        div[data-testid="stMain"] div.stButton > button,
        .main .block-container div.stButton > button[type="primary"],
        div[data-testid="column"] div.stButton > button[type="primary"],
        div.stButton > button[type="primary"] {
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%) !important;
            color: #495057 !important;
            border-radius: 8px !important;
            border: 1px solid #ced4da !important;
            font-weight: 500 !important;
            padding: 0 24px !important;
            transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1) !important;
            height: 48px !important;
        }
        div[data-testid="stMain"] div.stButton > button:hover,
        .main .block-container div.stButton > button[type="primary"]:hover,
        div[data-testid="column"] div.stButton > button[type="primary"]:hover,
        div.stButton > button[type="primary"]:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(0,0,0,0.1) !important;
            background: linear-gradient(135deg, #e9ecef 0%, #dee2e6 100%) !important;
            border: 1px solid #adb5bd !important;
        }

        div[data-testid="stTextInput"] input {
            border-radius: 8px !important;
            border: 1px solid #e0e0e0 !important;
            background: #ffffff !important;
            color: #212529 !important;
            padding: 12px 16px !important;
            font-size: 14px !important;
            transition: all 0.3s ease;
        }
        div[data-testid="stTextInput"] input:focus {
            box-shadow: 0 0 0 3px rgba(173, 181, 189, 0.2) !important;
            border-color: #adb5bd !important;
            transform: translateY(-1px);
        }

        div[data-testid="stDataFrame"] {
            transition: all 0.3s ease;
            border-radius: 8px !important;
            border: 1px solid #e0e0e0 !important;
            overflow: hidden !important;
        }
        div[data-testid="stDataFrame"]:hover {
            box-shadow: 0 4px 12px rgba(0,0,0,0.05);
        }
        thead tr th {
            background-color: #f8f9fa !important;
            color: #495057 !important;
            font-weight: 600 !important;
            border-bottom: 2px solid #dee2e6 !important;
        }
        tbody tr td {
            color: #212529 !important;
            border-bottom: 1px solid #f8f9fa !important;
            transition: background-color 0.2s ease;
        }
        tbody tr:hover td {
            background-color: #e9ecef !important;
        }

        /* ── Max width ── */
        .main .block-container { max-width: 1000px !important; padding-top: 1rem !important; }

        </style>
        """, unsafe_allow_html=True)
        
        gradient_bar = "linear-gradient(to right, #e0e0e0, #f8f9fa)"

    # --- REDESIGNED HEADER WIDGET (With Toggle Switch + localStorage) ---
    is_checked = "checked" if google_mode else ""
    
    components.html("""
    <!DOCTYPE html>
    <html>
    <head>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Product+Sans:wght@400;500;700&display=swap');
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
    
    :root {
        --blue: #4285F4;
        --yellow: #FBBC05;
        --green: #34A853;
        --red: #EA4335;
    }
    
    * { margin: 0; padding: 0; box-sizing: border-box; }
    
    body {
        font-family: 'Product Sans', 'Inter', 'Arial', sans-serif;
        background: transparent;
        overflow: hidden;
    }
    
    .header-container {
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding: 12px 4px;
        height: 80px;
    }
    
    /* LEFT: Logo Section */
    .logo-section {
        display: flex;
        align-items: center;
        gap: 0;
    }
    
    .logo-widget {
        display: flex;
        align-items: center;
        font-size: 52px;
        font-weight: 600;
        line-height: 1;
    }
    
    .logo-widget span.letter {
        display: inline-block;
        transition: color 0.4s cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    /* Default: normal appearance */
    body { background: #f8f9fa; }
    .header-container { background: #f8f9fa; }
    .logo-widget span.letter { color: #5f6368; }
    .maps-text { color: #5f6368; }
    
    /* Toggle section - normal mode default */
    .toggle-section {
        background: #ffffff;
        border: 1px solid #e0e0e0;
    }
    .toggle-label { color: #5f6368; }
    
    /* Accent bar - normal mode default */
    .accent-bar {
        background: linear-gradient(to right, #e0e0e0, #f8f9fa) !important;
    }
    
    /* Toggle track - normal mode default */
    .toggle-track { background: #dadce0; border: 1px solid #e0e0e0; }
    .toggle-thumb { background: #ffffff; color: #5f6368; }

    /* Google mode: restore light/colored appearance */
    .header-container.google-active { background: transparent; }
    body:has(.header-container.google-active) { background: transparent; }

    .header-container.google-active .logo-widget .letter-G { color: var(--blue); }
    .header-container.google-active .logo-widget .letter-o1 { color: var(--red); }
    .header-container.google-active .logo-widget .letter-o2 { color: var(--yellow); }
    .header-container.google-active .logo-widget .letter-g { color: var(--blue); }
    .header-container.google-active .logo-widget .letter-l { color: var(--green); }
    .header-container.google-active .logo-widget .letter-e { color: var(--red); }
    .header-container.google-active .maps-text { color: #202124; }

    .header-container.google-active .toggle-section {
        background: #e8f0fe;
        border-color: #c6dafc;
    }
    .header-container.google-active .toggle-label { color: #1a73e8; }
    .header-container.google-active .toggle-track { background: #dadce0; border: none; }
    .header-container.google-active .toggle-thumb { background: #ffffff; color: inherit; }

    .letter-e { display: inline-block; transform: rotate(-15deg); }

    .maps-text {
        font-size: 46px;
        font-weight: 700;
        margin-left: 14px;
        white-space: nowrap;
        letter-spacing: -0.5px;
        transition: color 0.4s cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    /* RIGHT: Toggle Section */
    .toggle-section {
        display: flex;
        align-items: center;
        gap: 14px;
        border-radius: 100px;
        padding: 6px 8px 6px 18px;
        transition: all 0.3s ease;
    }
    
    .toggle-label {
        font-family: 'Inter', 'Product Sans', sans-serif;
        font-size: 13px;
        font-weight: 600;
        letter-spacing: 0.2px;
        white-space: nowrap;
        user-select: none;
        transition: color 0.3s ease;
    }
    
    /* Toggle Switch */
    input#theme-toggle { display: none; }
    
    .toggle-track {
        position: relative;
        width: 52px;
        height: 28px;
        background: #dadce0;
        border-radius: 14px;
        cursor: pointer;
        transition: background 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        flex-shrink: 0;
    }
    
    .toggle-thumb {
        position: absolute;
        top: 2px;
        left: 2px;
        width: 24px;
        height: 24px;
        border-radius: 50%;
        background: #ffffff;
        box-shadow: 0 1px 3px rgba(0,0,0,0.3), 0 1px 2px rgba(0,0,0,0.15);
        transition: transform 0.3s cubic-bezier(0.4, 0, 0.2, 1), background 0.3s ease;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 12px;
    }
    
    /* Toggle ON state (Google mode) */
    input#theme-toggle:checked + .toggle-track {
        background: linear-gradient(135deg, var(--blue), #1a73e8);
    }
    
    input#theme-toggle:checked + .toggle-track .toggle-thumb {
        transform: translateX(24px);
        background: #fff;
    }
    
    /* Hover effects */
    .toggle-track:hover {
        box-shadow: 0 0 0 4px rgba(0,0,0,0.05);
    }
    
    input#theme-toggle:checked + .toggle-track:hover {
        box-shadow: 0 0 0 4px rgba(26, 115, 232, 0.15);
    }
    
    /* Colored accent bar under header */
    .accent-bar {
        height: 3px;
        border-radius: 2px;
        background: linear-gradient(to right, #dadce0, #e8eaed);
        transition: background 0.5s cubic-bezier(0.4, 0, 0.2, 1);
        margin-top: 2px;
    }
    
    .header-container.google-active ~ .accent-bar {
        background: linear-gradient(to right, var(--blue) 25%, var(--red) 25%, var(--red) 50%, var(--yellow) 50%, var(--yellow) 75%, var(--green) 75%);
    }
    
    /* Responsive */
    @media only screen and (max-width: 600px) {
        .logo-widget { font-size: 32px; }
        .maps-text { font-size: 26px; margin-left: 8px; }
        .toggle-section { padding: 4px 6px 4px 12px; }
        .toggle-label { font-size: 11px; }
        .toggle-track { width: 44px; height: 24px; }
        .toggle-thumb { width: 20px; height: 20px; }
        input#theme-toggle:checked + .toggle-track .toggle-thumb { transform: translateX(20px); }
    }
    </style>
    </head>
    <body>
    
    <div class="header-container {google_active_class}" id="headerContainer">
        <!-- LEFT: Google Logo -->
        <div class="logo-section">
            <div class="logo-widget">
                <span class="letter letter-G">G</span>
                <span class="letter letter-o1">o</span>
                <span class="letter letter-o2">o</span>
                <span class="letter letter-g">g</span>
                <span class="letter letter-l">l</span>
                <span class="letter letter-e">e</span>
            </div>
            <div class="maps-text">Maps Scraper</div>
        </div>
        
        <!-- RIGHT: Theme Toggle -->
        <div class="toggle-section">
            <span class="toggle-label" id="toggleLabel">{toggle_label}</span>
            <input id="theme-toggle" type="checkbox" {is_checked}>
            <label for="theme-toggle" class="toggle-track">
                <div class="toggle-thumb" id="toggleThumb">{toggle_icon}</div>
            </label>
        </div>
    </div>
    
    <div class="accent-bar"></div>
    
    <script>
    const toggle = document.getElementById('theme-toggle');
    const header = document.getElementById('headerContainer');
    const label = document.getElementById('toggleLabel');
    const thumb = document.getElementById('toggleThumb');
    
    function updateUI(isChecked) {
        if (isChecked) {
            header.classList.add('google-active');
            label.textContent = 'Google Theme';
            thumb.textContent = '🎨';
        } else {
            header.classList.remove('google-active');
            label.textContent = 'Normal Mode';
            thumb.textContent = '✨';
        }
        // Persist to localStorage
        try {
            localStorage.setItem('gmaps_google_theme', isChecked ? '1' : '0');
        } catch(e) {}
    }
    
    // Init from current state
    updateUI(toggle.checked);
    
    toggle.addEventListener('change', function() {
        updateUI(this.checked);
        
        // BRIDGE: Trigger Streamlit hidden checkbox
        try {
            const doc = window.parent.document;
            const marker = doc.querySelector('img[alt="bridge_marker"]');
            if (marker) {
                const lbl = marker.closest('label');
                if (lbl) lbl.click();
            }
        } catch (e) {
            console.error("Bridge Error:", e);
        }
    });
    
    // On load, check localStorage and sync if needed
    try {
        const stored = localStorage.getItem('gmaps_google_theme');
        if (stored !== null) {
            const shouldBeChecked = stored === '1';
            if (toggle.checked !== shouldBeChecked) {
                // State mismatch - let Python state win on first load
                // but persist the current Python state
                localStorage.setItem('gmaps_google_theme', toggle.checked ? '1' : '0');
            }
        } else {
            // First visit: save current state
            localStorage.setItem('gmaps_google_theme', toggle.checked ? '1' : '0');
        }
    } catch(e) {}
    </script>
    </body>
    </html>
    """.replace("{is_checked}", is_checked)
       .replace("{google_active_class}", "google-active" if google_mode else "")
       .replace("{toggle_label}", "Google Theme" if google_mode else "Normal Mode")
       .replace("{toggle_icon}", "🎨" if google_mode else "✨"), height=90)
    
    # INIT SESSION STATE
    if 'scraper_running' not in st.session_state:
        st.session_state.scraper_running = False

    # Use a clean container for the form
    with st.form("scraper_form"):
        
        # Main Inputs in 2 columns
        c1, c2 = st.columns([1, 1], gap="medium")
        with c1:
            target_business = st.text_input(
                "Business Category", 
                value="Dentist", 
                placeholder="e.g. Dentist, Restaurant, Gym"
            )
        with c2:
            target_location = st.text_input(
                "Target Location", 
                value="Gotri, Vadodara", 
                placeholder="e.g. Gotri, Vadodara"
            )
        
        # Helper Text
        st.markdown("""
        <div class="helper-text">
            💡 <strong>Tip:</strong> Use specific locations like "Dentist in Gotri, Vadodara" for better results.
        </div>
        """, unsafe_allow_html=True)
        
        st.write("")  # Spacing
        
        # Action Button (Material Style)
        start_scrape = st.form_submit_button("Launch Scraper", type="primary", use_container_width=True)

        # Hardcoded Settings for Simple Mode (High Performance Default)
        filter_phone = False
        filter_website = False
        perf_deep = True
        perf_turbo = False
    
    # --- SHOW EXISTING SCRAPED RESULTS ---
    output_file = "scraped_results.csv"
    if os.path.exists(output_file) and not start_scrape:
        try:
            if os.stat(output_file).st_size == 0:
                df_existing = pd.DataFrame()
            else:
                df_existing = pd.read_csv(output_file, on_bad_lines='skip')
            
            if not df_existing.empty:
                st.markdown("---")
                st.subheader(f"📊 Previous Scrape Results ({len(df_existing)} leads)")
                
                # Reorder and Rename Columns
                # Reorder and Rename Columns
                # Format: Business Name, Phone Number, Address, Website, Email, Map Link, Status, Notes
                col_map = {
                    "clinic_name": "Business Name",
                    "phone_number": "Phone Number",
                    "address": "Address",
                    "website_url": "Website",
                    "email": "Email",
                    "place_url": "Map Link", # "Map Link" as per user request
                    "google_maps_url": "Map Link",
                    "url": "Map Link",
                    "Maps Link": "Map Link" # Handle legacy/user-edit consistency
                }
                
                df_existing.rename(columns=col_map, inplace=True)

                # Ensure missing columns exist
                if "Status" not in df_existing.columns:
                    df_existing["Status"] = "Generated"
                if "Priority" not in df_existing.columns:
                    df_existing["Priority"] = "WARM"
                if "Notes" not in df_existing.columns:
                    df_existing["Notes"] = ""
                
                # Define desired order
                desired_order = ["Business Name", "Phone Number", "Address", "Website", "Email", "Map Link", "Status", "Priority", "Notes"]
                
                # Filter / Arrange columns
                final_cols = [c for c in desired_order if c in df_existing.columns]
                # Add any other remaining columns at the end just in case
                remaining_cols = [c for c in df_existing.columns if c not in final_cols and c not in ["query", "location", "rating", "reviews"]]
                final_cols.extend(remaining_cols)
                
                df_display_existing = df_existing[final_cols]
                
                # Force string type for phone to avoid decimals causing issues later
                if "Phone Number" in df_display_existing.columns:
                     df_display_existing["Phone Number"] = df_display_existing["Phone Number"].astype(str).replace(r'\.0$', '', regex=True).replace("nan", "")

                # --- APPLY FILTERS ---
                if filter_phone:
                    # Filter: Must have phone (not empty, not "nan")
                    df_display_existing = df_display_existing[
                        (df_display_existing["Phone Number"].str.len() > 3) 
                    ]
                
                if filter_website:
                    # Filter: Must NOT have website (is empty or "nan")
                    # Check if Website col exists first
                    if "Website" in df_display_existing.columns:
                        df_display_existing = df_display_existing[
                            (df_display_existing["Website"].isna()) | 
                            (df_display_existing["Website"] == "") |
                            (df_display_existing["Website"].astype(str).str.lower() == "nan")
                        ]
                
                # Update header count to reflect filtered results
                if filter_phone or filter_website:
                    st.toast(f"Filters Active: Showing {len(df_display_existing)} leads")
                
                # --- Interactive Editor Config ---
                scrape_config = {
                    "Status": st.column_config.SelectboxColumn(
                        "Status",
                        options=["Generated", "Interested", "Not picking", "Asked to call later", "Meeting set", "Meeting Done", "Proposal sent", "Follow-up scheduled", "Not interested", "Closed - Won", "Closed - Lost"],
                        required=True,
                        width="medium"
                    ),
                    "Priority": st.column_config.SelectboxColumn(
                         "Priority",
                         options=["HOT", "WARM", "COLD"],
                         required=True,
                         width="small"
                    ),
                    "Map Link": st.column_config.LinkColumn(
                        "Map Link", display_text="View on Map"
                    ),
                    "Website": st.column_config.LinkColumn(
                        "Website", display_text="Visit"
                    ),
                    "Notes": st.column_config.TextColumn(
                        "Notes", width="large"
                    )
                }
                
                scrape_config["_index"] = None

                edited_scrape_df = st.data_editor(
                    df_display_existing, 
                    column_config=scrape_config,
                    use_container_width=True, 
                    height=400,
                    key="scraped_results_editor",
                    num_rows="fixed",
                    hide_index=True
                )
                
                # Update df_display_existing with edits for download/saving
                if not edited_scrape_df.equals(df_display_existing):
                    df_display_existing = edited_scrape_df
                
                # Download and Import buttons
                # Download and Import buttons
                col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
                
                import io
                # Robust filename sanitizer
                def sanitize_name(s):
                    if not s: return "data"
                    # Keep alphanumeric, underscores, hyphens. Replace spaces with underscores.
                    s = str(s).strip().replace(" ", "_")
                    s = "".join(c for c in s if c.isalnum() or c in ('_', '-'))
                    return s if s else "data"
                
                # Use current inputs for filename, default to "Scraped_Leads" if empty
                safe_query = sanitize_name(target_business) if target_business else "Search"
                safe_loc = sanitize_name(target_location) if target_location else "Location"
                export_filename = f"{safe_query}_{safe_loc}_Leads"

                # Prepare Export DataFrame with Renamed Columns
                export_df = df_display_existing.copy()
                
                # Normalize columns to ensure matching works
                export_df.columns = [str(c).strip() for c in export_df.columns]
                
                # Direct mapping - force rename
                # We used "Business Name" in the UI, we want "Name" in export
                rename_map = {
                    "Business Name": "Name",
                    "Phone Number": "Contact No",
                    "Priority": "Response",
                    "Map Link": "Map Link",
                    "Status": "Status", # Keep same
                    "Address": "Address", # Keep same
                    "Website": "Website", # Keep same
                    "Email": "Email" # Keep same
                }
                export_df.rename(columns=rename_map, inplace=True)
                
                # Enforce exact column order for the primary columns
                target_order = ["Name", "Contact No", "Address", "Website", "Email", "Status", "Response", "Map Link", "Notes"]
                
                # Reorder columns: Keep those that exist in target_order, append the rest
                existing_cols = list(export_df.columns)
                ordered_cols = [c for c in target_order if c in existing_cols]
                other_cols = [c for c in existing_cols if c not in ordered_cols]
                
                export_df = export_df[ordered_cols + other_cols]
                
                # Explanation for the user constraint - Removed per user request

                # Removed Plain CSV button to avoid confusion
                
                with col_btn2:
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        export_df.to_excel(writer, index=False, sheet_name='Leads')
                        
                        # --- Excel Styling & Validation ---
                        workbook = writer.book
                        worksheet = writer.sheets['Leads']
                        
                        # 1. Validation (Hidden Sheet)
                        ref_sheet = workbook.create_sheet("DataValidation")
                        ref_sheet.sheet_state = 'hidden'
                        
                        status_options = ["Generated", "Interested", "Not picking", "Asked to call later", "Meeting set", "Meeting Done", "Proposal sent", "Follow-up scheduled", "Not interested", "Closed - Won", "Closed - Lost"]
                        response_options = ["HOT", "WARM", "COLD"]
                        
                        for idx, val in enumerate(status_options, start=1): ref_sheet.cell(row=idx, column=1).value = val
                        for idx, val in enumerate(response_options, start=1): ref_sheet.cell(row=idx, column=2).value = val
                            
                        from openpyxl.worksheet.datavalidation import DataValidation
                        from openpyxl.styles import Font, PatternFill, Alignment
                        from openpyxl.utils import get_column_letter

                        max_row = len(export_df) + 1
                        
                        # Apply Validation Logic
                        def get_col_letter_map(col_name):
                            try: return get_column_letter(export_df.columns.get_loc(col_name) + 1)
                            except: return None

                        # Status Validation
                        status_col = get_col_letter_map("Status")
                        if status_col:
                            dv_status = DataValidation(type="list", formula1=f"'DataValidation'!$A$1:$A${len(status_options)}", allow_blank=True)
                            worksheet.add_data_validation(dv_status)
                            dv_status.add(f'{status_col}2:{status_col}{max_row}')

                        # Response (Priority) Validation
                        resp_col = get_col_letter_map("Response")
                        if resp_col:
                            dv_resp = DataValidation(type="list", formula1=f"'DataValidation'!$B$1:$B${len(response_options)}", allow_blank=True)
                            worksheet.add_data_validation(dv_resp)
                            dv_resp.add(f'{resp_col}2:{resp_col}{max_row}')

                        # 2. Header Styling (Dark Grey Background, Yellow Bold Text)
                        header_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid") # Dark Grey
                        header_font = Font(color="FFFF00", bold=True, size=11, name='Calibri') # Yellow, Bold
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 3. AutoFilter
                        worksheet.auto_filter.ref = worksheet.dimensions
                        
                        # 4. Make Links Clickable (Website & Map)
                        # Helper to apply hyperlinks
                        link_font = Font(color="0563C1", underline="single")
                        
                        # Website Column
                        try:
                            web_idx = export_df.columns.get_loc("Website") + 1
                            web_col = get_column_letter(web_idx)
                            for row in range(2, max_row + 1):
                                cell = worksheet[f"{web_col}{row}"]
                                if cell.value and isinstance(cell.value, str) and (cell.value.startswith('http') or cell.value.startswith('www')):
                                    cell.hyperlink = cell.value
                                    cell.font = link_font
                        except: pass

                        # Google Maps Column
                        try:
                            map_idx = export_df.columns.get_loc("Map Link") + 1
                            map_col = get_column_letter(map_idx)
                            for row in range(2, max_row + 1):
                                cell = worksheet[f"{map_col}{row}"]
                                if cell.value:
                                    cell.hyperlink = cell.value
                                    cell.value = "View on Map"
                                    cell.font = link_font
                        except Exception as e:
                            print(f"DEBUG: Map Link formatting failed: {e}")
                        
                        # 5. Conditional Formatting with DXF (Differential Formatting)
                        # This ensures styles apply even if cells have default styles
                        from openpyxl.styles import PatternFill, Font, Color
                        from openpyxl.styles.differential import DifferentialStyle
                        from openpyxl.formatting.rule import Rule

                        def create_dxf_rule(col_letter, value, fill_hex, font_hex):
                            # Strip hash
                            bg_color = Color(rgb="FF" + fill_hex.lstrip('#')) 
                            font_color = Color(rgb="FF" + font_hex.lstrip('#'))
                            
                            dxf = DifferentialStyle(
                                font=Font(color=font_color), 
                                fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                            )
                            rule = Rule(type="expression", dxf=dxf, stopIfTrue=True)
                            # Formula: check if cell equals value (wrapped in quotes)
                            rule.formula = [f'${col_letter}2="{value}"'] 
                            return rule

                        # Status Colors
                        status_styles = {
                            "Interested": ("#DFF5E1", "#1B5E20"),
                            "Not picking": ("#F0F0F0", "#616161"),
                            "Asked to call later": ("#FFF8E1", "#8D6E00"),
                            "Meeting set": ("#E3F2FD", "#0D47A1"),
                            "Meeting Done": ("#E0F2F1", "#004D40"),
                            "Proposal sent": ("#F3E5F5", "#4A148C"),
                            "Follow-up scheduled": ("#FFE0B2", "#E65100"),
                            "Not interested": ("#FDECEA", "#B71C1C"),
                            "Closed - Won": ("#C8E6C9", "#1B5E20"),
                            "Closed - Lost": ("#ECEFF1", "#37474F"),
                            "Generated": ("#FFFFFF", "#000000")
                        }
                        # Response (Priority) Colors
                        response_styles = {
                            "HOT": ("#F25C54", "#FFFFFF"),
                            "WARM": ("#FFE5B4", "#5A3E00"),
                            "COLD": ("#E3F2FD", "#1E3A8A")
                        }

                        # Status Rules
                        if status_col:
                            ws_range = f'{status_col}2:{status_col}{max_row}'
                            for status_val, (bg, txt) in status_styles.items():
                                rule = create_dxf_rule(status_col, status_val, bg, txt)
                                worksheet.conditional_formatting.add(ws_range, rule)

                        # Response Rules
                        if resp_col:
                            ws_range_resp = f'{resp_col}2:{resp_col}{max_row}'
                            for resp_val, (bg, txt) in response_styles.items():
                                rule = create_dxf_rule(resp_col, resp_val, bg, txt)
                                worksheet.conditional_formatting.add(ws_range_resp, rule)

                        # 6. Auto-width columns adjustments
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except: pass
                            adjusted_width = (max_length + 2) * 1.05
                            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

                    st.download_button(
                        label="📥 Excel (Formatted)",
                        data=buffer.getvalue(),
                        file_name=f"{export_filename}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key='dl_xlsx_formatted_v1',
                        use_container_width=True
                    )

                with col_btn3:
                    if st.button("💾 Import to CRM", use_container_width=True, type="primary"):
                        count = 0
                        prog_bar = st.progress(0, text="Importing leads to CRM...")
                        for idx, row in df_display_existing.iterrows():
                            payload = {
                                "businessName": str(row.get('Business Name') or row.get('clinic_name') or "Unknown"),
                                "phone": str(row.get('Phone Number') or row.get('phone_number') or ""),
                                "address": str(row.get('Address') or row.get('address') or ""),
                                "email": str(row.get('Email') or row.get('email') or ""),
                                "status": "Generated",
                                "priority": "WARM"
                            }
                            # Clean 'nan' strings
                            for key in payload:
                                if payload[key] == 'nan':
                                    payload[key] = ""
                            
                            if create_lead(payload):
                                count += 1
                            prog_bar.progress(min((idx+1)/len(df_display_existing), 1.0))
                        
                        st.success(f"✅ Successfully imported {count} leads to CRM!")
                        time.sleep(1.5)
                        st.rerun()

                # Add explicit Save to History button
                if st.button("📂 Save to Scraped Leads History", use_container_width=True):
                    try:
                        csv_for_history = df_display_existing.to_csv(index=False)
                        # Use actual input values if available
                        ts = datetime.now().strftime("%d %b %H:%M")
                        
                        # Get values directly from the input widgets above if possible
                        # Since button is pressed, input widgets 'target_business' and 'target_location' from lines 3459/3460 are in scope
                        t_biz = target_business.strip() if target_business else "Search"
                        t_loc = target_location.strip() if target_location else "Location"
                        
                        # Format: Dentist_Gotri, Vadodara
                        q_name = f"{t_biz}_{t_loc}"
                        
                        # POST to backend
                        requests.post(
                            EXECUTIONS_API, 
                            json={
                                "query": t_biz, 
                                "location": t_loc, 
                                "name": q_name,
                                "leadsGenerated": len(df_display_existing),
                                "status": "Success",
                                "fileContent": csv_for_history
                            },
                            timeout=10
                        )
                        st.toast("✅ Saved to Scraped Leads History!")
                        time.sleep(1)
                    except Exception as h_err:
                        st.error(f"Save Error: {h_err}")
        except Exception as e:
            if "No columns to parse" not in str(e):
                st.info(f"Could not load previous results: {e}")

    if start_scrape:
        if not target_business or not target_location:
            st.error("Please provide both Business Type and Location.")
        else:
            search_query = f"{target_business} in {target_location}"
            
            # Persist params for autosave
            st.session_state['last_scrape_business'] = target_business
            st.session_state['last_scrape_location'] = target_location
            
            output_file = "scraped_results.csv"
            
            # Remove previous file if exists
            if os.path.exists(output_file):
                try:
                    os.remove(output_file)
                except: pass
            
            scraper_dir = os.path.join(os.getcwd(), "dental_scraper")
            # Output in root
            # Determine Settings
            # Determine Settings
            # Determine Settings
            # Determine Settings
            # User Request: "300-350 agents deep, 100 turbo, 180 standard"
            
            if perf_turbo:
                s_limit = 5000 if perf_deep else 2000
            else:
                s_limit = 5000 if perf_deep else 600 # Base limit covering "at least 180"
            
            cmd = [
                sys.executable, "-m", "scrapy", "crawl", "dental_spider",
                "-a", f"search_query={search_query}",
                "-a", f"must_have_phone={str(filter_phone).lower()}",
                "-a", f"no_website={str(filter_website).lower()}",
                "-a", f"scroll_limit={s_limit}",
                "-O", f"../{output_file}"
            ]
            
            # Turbo / Deep Mode Logic: High-Fidelity Headless
            # To fix "low leads" in headless, we MUST simulate a real screen resolution (1920x1080).
            # Otherwise Google Maps assumes a tiny viewport and stops lazy-loading results.
            
            if perf_deep:
                # Deep Mining: 100 Agents, Headless with 1080p Viewport
                cmd.extend(["-s", "CONCURRENT_REQUESTS=100"]) 
                cmd.extend(["-s", "CONCURRENT_REQUESTS_DOMAIN=100"])
                cmd.extend(["-s", "DOWNLOAD_DELAY=0.03"]) 
                cmd.extend(["-s", 'PLAYWRIGHT_LAUNCH_OPTIONS={"headless": true, "timeout": 120000, "args": ["--window-size=1920,1080", "--start-maximized", "--disable-dev-shm-usage", "--no-sandbox", "--disable-gpu", "--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"]}'])
            elif perf_turbo:
                # Turbo: 50 Agents, Headless with 1080p Viewport
                cmd.extend(["-s", "CONCURRENT_REQUESTS=50"]) 
                cmd.extend(["-s", "CONCURRENT_REQUESTS_DOMAIN=50"])
                cmd.extend(["-s", "DOWNLOAD_DELAY=0.05"]) 
                cmd.extend(["-s", 'PLAYWRIGHT_LAUNCH_OPTIONS={"headless": true, "timeout": 120000, "args": ["--window-size=1920,1080", "--start-maximized", "--disable-dev-shm-usage", "--no-sandbox", "--disable-gpu", "--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"]}'])
            else:
                # Standard: Default
                cmd.extend(["-s", 'PLAYWRIGHT_LAUNCH_OPTIONS={"headless": true, "timeout": 120000, "args": ["--window-size=1920,1080", "--disable-dev-shm-usage", "--no-sandbox", "--disable-gpu"]}'])
            
            # Status container placeholder with modern UI
            status_container = st.empty()
            status_container.markdown("""
            <div class="loading-container" style="height: 200px; display: flex; flex-direction: column; justify-content: center; align-items: center;">
                <div class="miner-icon">⛏️</div>
                <div class="loading-text">Mining Google Maps for business data...</div>
                <div class="tech-lines"></div>
                <div class="loading-subtext">
                    This may take a few moments. Please don't refresh the page.
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            import subprocess
            import time
            
            # Clean up old progress file
            progress_file = "scraper_progress.txt"
            if os.path.exists(progress_file):
                try:
                    os.remove(progress_file)
                except:
                    pass
            
            try:
                # Redirect output to file for debugging
                log_path = os.path.join(scraper_dir, "run_log.txt")
                log_file = open(log_path, "w")
                
                # Run async with Popen to monitor progress
                process = subprocess.Popen(
                    cmd, 
                    cwd=scraper_dir, 
                    stdout=log_file,
                    stderr=subprocess.STDOUT,
                    text=True
                )
                
                # Save file handle to close later? 
                # Actually, subprocess keeps it open. We can let Python close it on GC or explicit cleanup.
                # Ideally we should store it to close it, but for now let's just let it run.
                
                # Monitor loop removed for background logic
                pass
                
                # Result processing logic removed
                st.session_state.scraper_process = process
                st.session_state.scraper_running = True
                st.rerun()
                        
            except Exception as e:
                st.error(f"Execution Error: {e}")

    # --- BACKGROUND MONITOR ---
    if st.session_state.scraper_running:
        status_container = st.empty()
        
        # Get Process
        process = st.session_state.get('scraper_process')

        # Robust check if process is actually running
        if process:
            # Poll status
            poll = process.poll()
            
            if poll is None:
                # --- RUNNING ---
                # Read progress
                progress_file = "scraper_progress.txt"
                count_info = "Starting..."
                current_query = "Initializing..."
                
                if os.path.exists(progress_file):
                    try:
                        with open(progress_file, "r") as f:
                            progress = f.read().strip()
                            parts = progress.split("|")
                            if len(parts) == 2:
                                count_info = parts[0]
                                current_query = parts[1]
                    except: pass
                
                # STYLE DISPATCHER FOR STATUS CARD
                if st.session_state.get("google_ui_mode", False):
                    # Google Material Style
                    s_bg = "#ffffff"
                    s_border = "1px solid #dadce0"
                    s_shadow = "0 1px 3px 0 rgba(60,64,67,0.3), 0 4px 8px 3px rgba(60,64,67,0.15)"
                    s_text_main = "#202124"
                    s_text_sub = "#5f6368"
                    s_icon_bg = "#e8f0fe"
                    s_rocket_filter = "none" 
                    s_scale = "1.005"
                else:
                    # Stealth/Cyber Gradient Style
                    s_bg = "linear-gradient(135deg, #3b82f6 0%, #8b5cf6 100%)"
                    s_border = "1px solid rgba(255, 255, 255, 0.2)"
                    s_shadow = "0 8px 32px rgba(59, 130, 246, 0.3)"
                    s_text_main = "white"
                    s_text_sub = "rgba(255, 255, 255, 0.9)"
                    s_icon_bg = "rgba(255, 255, 255, 0.2)"
                    s_rocket_filter = "none"
                    s_scale = "1.02"

                status_container.markdown(f"""
                <div style="
                    background: {s_bg};
                    border-radius: 16px;
                    padding: 24px 32px;
                    margin: 20px 0;
                    box-shadow: {s_shadow};
                    border: {s_border};
                    animation: pulse 2s ease-in-out infinite;
                ">
                    <div style="display: flex; align-items: center; gap: 16px;">
                        <div style="
                            width: 48px;
                            height: 48px;
                            background: {s_icon_bg};
                            border-radius: 50%;
                            display: flex;
                            align-items: center;
                            justify-content: center;
                            font-size: 24px;
                            animation: pulse 1s ease-in-out infinite;
                            filter: {s_rocket_filter};
                        ">
                            🚀
                        </div>
                        <div style="flex: 1;">
                            <div style="
                                color: {s_text_main};
                                font-size: 18px;
                                font-weight: 700;
                                margin-bottom: 4px;
                                font-family: 'Product Sans', sans-serif;
                            ">
                                Agents Active / Leads Found ({count_info})
                            </div>
                            <div style="
                                color: {s_text_sub};
                                font-size: 14px;
                                font-weight: 500;
                            ">
                                Now searching: <strong>'{current_query}'</strong>
                            </div>
                        </div>
                    </div>
                </div>
                <style>
                    @keyframes pulse {{
                        0%, 100% {{ transform: scale(1); box-shadow: {s_shadow}; }}
                        50% {{ transform: scale({s_scale}); box-shadow: {s_shadow}; }}
                    }}
                    @keyframes spin {{
                        from {{ transform: rotate(0deg); }}
                        to {{ transform: rotate(360deg); }}
                    }}
                </style>
                """, unsafe_allow_html=True)

                if st.button("🛑 Stop Scraper"):
                    process.terminate()
                    st.session_state.scraper_running = False
                    st.rerun()

                time.sleep(1)
                st.rerun()
            
            else:
                # --- FINISHED ---
                st.session_state.scraper_running = False
                
                if poll == 0:
                    status_container.success("✅ Scraping Complete!")
                    st.balloons()
                    
                    output_file = "scraped_results.csv"
                    # --- POST PROCESSING LOGIC ---
                    if os.path.exists(output_file):
                        df_res = pd.DataFrame()
                        if os.stat(output_file).st_size == 0:
                             st.warning("⚠️ Scraper finished but no data was found. Please try a different location or business type.")
                        else:
                            try:
                                df_res = pd.read_csv(output_file, on_bad_lines='skip')
                            except pd.errors.EmptyDataError:
                                st.warning("⚠️ Scraper finished but the results file was empty.")
                        
                        if not df_res.empty:
                            # Show Data
                            col_map = {
                                "clinic_name": "Business Name",
                                "phone_number": "Phone Number",
                                "address": "Address",
                                "website_url": "Website",
                                "email": "Email",
                                "place_url": "Map Link"
                            }
                            df_res.rename(columns=col_map, inplace=True)
                            
                            # Add CRM Columns
                            if "Status" not in df_res.columns: df_res["Status"] = "Generated"
                            if "Priority" not in df_res.columns: df_res["Priority"] = "WARM"
                            
                            st.dataframe(df_res, use_container_width=True)
                            
                            # Save to History (Automatic)
                            try:
                                csv_for_history = df_res.to_csv(index=False)
                                
                                # Use persisted params or fallback
                                s_bus = st.session_state.get('last_scrape_business', target_business)
                                s_loc = st.session_state.get('last_scrape_location', target_location)
                                
                                q_name = f"{s_bus} - {s_loc}"
                                
                                requests.post(
                                    EXECUTIONS_API, 
                                    json={
                                        "query": f"{s_bus} in {s_loc}", 
                                        "location": s_loc, 
                                        "name": q_name,
                                        "leadsGenerated": len(df_res), # Save TOTAL scraped, not just new
                                        "status": "Success",
                                        "fileContent": csv_for_history
                                    },
                                    timeout=10
                                )
                                st.success(f"✅ Auto-saved results to Scraped Leads History ({len(df_res)} records)")
                            except Exception as h_err:
                                st.error(f"Auto-save to History Failed: {h_err}")
                            except Exception as e:
                                st.error(f"Failed to connect to backend: {e}")
                                
                            if st.button("🗑️ Clear Results"):
                                st.rerun()

                    else:
                        status_container.error(f"❌ Scraper Failed with Exit Code {poll}")
                        # Capture Stderr (Robust)
                        try:
                            # We might have already consumed pipe? communicate() handles this.
                            _, stderr_out = process.communicate(timeout=5)
                            if stderr_out:
                                with st.expander("📝 View Error Logs (Click Here)", expanded=True):
                                    st.code(stderr_out, language="text")
                                
                                if "playwright" in stderr_out.lower() or "browser" in stderr_out.lower() or "executable" in stderr_out.lower():
                                    st.warning("💡 To Fix: In Dashboard, click 'Manage App' -> 'Reboot App' to install browsers.")
                        except Exception as e:
                             st.error(f"Could not retrieve error logs: {e}")
        else:
            st.session_state.scraper_running = False
            st.rerun()

# ================== LEAD GEN HISTORY ==================
# ================== SCRAPED LEADS (History + Editing) ==================
if "Scraped Leads" in page:
    st.markdown("""
        <style>
        .staging-header {
            font-size: 28px;
            font-weight: 700;
            color: #1e293b;
            margin-bottom: 8px;
        }
        .staging-subtitle {
            color: #64748b;
            font-size: 15px;
            margin-bottom: 24px;
        }
        /* Custom metric card styling to match other elements */
        [data-testid="stMetricValue"] {
            font-size: 28px !important;
            font-weight: 700 !important;
            color: #3b82f6 !important;
        }
        [data-testid="stMetricLabel"] {
            font-size: 14px !important;
            color: #64748b !important;
        }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="staging-header">📥 Lead Staging & Import</div>', unsafe_allow_html=True)
    st.markdown('<div class="staging-subtitle">Review, refine, and edit your scraped leads before pulling them into your active CRM pipeline.</div>', unsafe_allow_html=True)

    # 1. Fetch Summary List
    executions = fetch_data(EXECUTIONS_API)
    
    if not executions:
        st.info("No scraped data found.")
    else:
        df_hist = pd.DataFrame(executions)
        
        # Prepare minimal list
        if "name" not in df_hist.columns:
            df_hist["name"] = "Untitled"
        
        # Formatting
        try:
            df_hist["date_fmt"] = pd.to_datetime(df_hist["date"]).dt.strftime("%d %b %H:%M")
        except:
            df_hist["date_fmt"] = df_hist["date"]
            
        df_hist["label"] = df_hist.apply(lambda x: f"{x['name']} ({x['leadsGenerated']}) - {x['date_fmt']}", axis=1)
        
        # --- CONTROL BAR (Toolbar Style) ---
        st.markdown("### 🗂️ Select Session")
        c1, c2, c3 = st.columns([3, 2, 2], vertical_alignment="bottom")
        
        with c1:
            # Session Selector
            if "selected_scrape_id" not in st.session_state:
                st.session_state.selected_scrape_id = None
                
            def_idx = 0
            if st.session_state.selected_scrape_id:
                try:
                    match = df_hist[df_hist["id"] == st.session_state.selected_scrape_id]
                    if not match.empty:
                        def_idx = df_hist.index.get_loc(match.index[0])
                except: pass

            sel_label = st.selectbox(
                "Choose a scraping session",
                options=df_hist["label"].tolist(),
                index=def_idx,
                key="scrape_selector_box",
                label_visibility="collapsed"
            )
            # Find ID
            sel_row = df_hist[df_hist["label"] == sel_label].iloc[0]
            st.session_state.selected_scrape_id = sel_row["id"]

        with c2:
            # Meta Info (Compact)
            st.markdown(
                f"<div style='color: #64748b; font-size: 14px; padding: 10px 0;'>📅 {sel_row['date_fmt']} &nbsp; • &nbsp; 📍 {sel_row['location']}</div>", 
                unsafe_allow_html=True
            )

        with c3:
            # Actions (Right Aligned)
            ac_sub1, ac_sub2 = st.columns([1, 1])
            with ac_sub1:
                with st.popover("✏️ Rename", use_container_width=True):
                    new_name = st.text_input("New Name", value=sel_row["name"])
                    if st.button("Save", type="primary", use_container_width=True):
                            try:
                                requests.put(f"{EXECUTIONS_API}/{sel_row['id']}", json={"name": new_name})
                                st.toast("Renamed successfully")
                                time.sleep(0.5)
                                st.rerun()
                            except: pass
            with ac_sub2:
                if st.button("🗑️ Delete", type="secondary", use_container_width=True, key=f"del_{sel_row['id']}"):
                    try:
                        st.toast("Delete not implemented in API")
                    except: pass
        
        st.markdown("<hr style='margin: 10px 0 20px 0;'/>", unsafe_allow_html=True)

        # REMOVE INDENTATION FOR EDITOR BLOCK
        with st.container():
            if st.session_state.selected_scrape_id:
                # FETCH FULL RECORD (with fileContent)
                try:
                    res_detail = requests.get(f"{EXECUTIONS_API}/{st.session_state.selected_scrape_id}")
                    if res_detail.status_code == 200:
                        full_data = res_detail.json()
                        csv_content = full_data.get("fileContent", "")
                        
                        if csv_content:
                            # Parse CSV string
                            from io import StringIO
                            df_file = pd.read_csv(StringIO(csv_content))
                            
                            # --- 1. STANDARDIZE COLUMNS TO MATCH CRM ---
                            col_map = {
                                "clinic_name": "Company Name",
                                "Business Name": "Company Name",
                                "phone_number": "Phone Number",
                                "address": "Address",
                                "email": "Email",
                                "website": "Website",
                                "Website": "Website",
                                "rating": "Rating",
                                "reviews": "Reviews",
                                "url": "Map", # Scraper usually dumps url here
                                "google_maps_url": "Map",
                                "place_url": "Map",
                                "Maps Link": "Map",  # Handle files exported with "Maps Link"
                                "Map Link": "Map"
                            }
                            df_file.rename(columns=col_map, inplace=True)
                            
                            # Ensure essential columns exist
                            desired_order = ["Company Name", "Phone Number", "Email", "Address", "Map", "Website", "Status", "Priority", "Notes"]
                            for col in desired_order:
                                if col not in df_file.columns:
                                    if col == "Status": df_file[col] = "Generated"
                                    elif col == "Priority": df_file[col] = "WARM"
                                    elif col == "Notes": df_file[col] = "" 
                                    else: df_file[col] = None
                            
                            # Reorder for UI
                            other_cols = [c for c in df_file.columns if c not in desired_order and c != "No."]
                            final_cols = ["No."] + desired_order + other_cols
                            df_file = df_file[[c for c in final_cols if c in df_file.columns]]

                            # FORCE TEXT TYPES for editable columns
                            text_cols = ["Company Name", "Phone Number", "Email", "Address", "Notes", "Map", "Website"]
                            for t_c in text_cols:
                                if t_c in df_file.columns:
                                    df_file[t_c] = df_file[t_c].astype(str).replace("nan", "")
                                    df_file[t_c] = df_file[t_c].replace("None", "")
                                    
                                    # Fix Phone Number formatting (remove .0 decimals)
                                    if t_c == "Phone Number":
                                        df_file[t_c] = df_file[t_c].str.replace(r'\.0$', '', regex=True)

                            

                            # --- CALCULATE METRICS ---
                            total_leads = len(df_file)
                            emails_found = df_file['Email'].replace("", None).count()
                            phones_found = df_file['Phone Number'].replace("", None).count()
                            
                            # --- EDITOR CONTAINER ---
                            with st.container():
                                st.markdown(f"#### 📝 {full_data.get('name')}")

                                # Metric Row inside a visually distinct container
                                m1, m2, m3, m4 = st.columns(4)
                                m1.metric("Total Leads", total_leads)
                                m2.metric("Emails Found", emails_found)
                                m3.metric("Phones Found", phones_found)
                                map_count = df_file['Map'].replace("", None).count()
                                m4.metric("Maps Linked", map_count)
                                
                                st.write("") # Spacer

                                # --- STYLER ---
                                def staging_style(val, col):
                                    if not val or pd.isna(val): return ""
                                    val_str = str(val).strip()

                                    status_colors = {
                                        "Interested": ("#D6E9FF", "#1A73E8"),
                                        "Not picking": ("#FF00B8", "#FFFFFF"),
                                        "Asked to call later": ("#FFD9B3", "#8A4B00"),
                                        "Meeting set": ("#E8D6FF", "#5E35B1"),
                                        "Meeting Done": ("#D1C4E9", "#4527A0"),
                                        "Proposal sent": ("#E2F0D9", "#2E7D32"),
                                        "Follow-up scheduled": ("#FFE082", "#8A6D00"),
                                        "Not interested": ("#D7CCC8", "#5D4037"),
                                        "Closed - Won": ("#C8E6C9", "#1B5E20"),
                                        "Closed - Lost": ("#FFCDD2", "#B71C1C"),
                                        "Generated": ("#F5F5F5", "#616161")
                                    }
                                    
                                    prio_colors = {
                                        "HOT": ("#F4B183", "#7A2E00"),
                                        "WARM": ("#FFE599", "#7A5B00"),
                                        "COLD": ("#9FC5E8", "#0B5394")
                                    }

                                    if col == "Status" and val_str in status_colors:
                                        bg, fg = status_colors[val_str]
                                        return f"background-color: {bg} !important; color: {fg} !important; font-weight: 600;"
                                    if col == "Priority" and val_str in prio_colors:
                                        bg, fg = prio_colors[val_str]
                                        return f"background-color: {bg} !important; color: {fg} !important; font-weight: 600;"
                                    return ""

                                # Apply styling
                                df_file.insert(0, "No.", range(1, len(df_file) + 1))
                                if not df_file.empty:
                                    styled_df = df_file.style.apply(lambda x: [staging_style(x[c], c) for c in x.index], axis=1).hide(axis="index")
                                else:
                                    styled_df = df_file

                                # The Grid
                                staging_config = {
                                    "No.": st.column_config.NumberColumn("No.", width="small", disabled=True),
                                    "Company Name": st.column_config.TextColumn("Company Name", required=True, width="medium"),
                                    "Phone Number": st.column_config.TextColumn("Phone Number", width="medium"),
                                    "Email": st.column_config.TextColumn("Email", width="medium"),
                                    "Address": st.column_config.TextColumn("Address", width="large"),
                                    "Map": st.column_config.LinkColumn("Map", display_text="View on Map"),
                                    "Website": st.column_config.LinkColumn("Website", display_text="Visit"),
                                    "Status": st.column_config.SelectboxColumn("Status", options=["Generated", "Interested", "Not picking", "Asked to call later", "Meeting set", "Meeting Done", "Proposal sent", "Follow-up scheduled", "Not interested", "Closed - Won", "Closed - Lost"], default="Generated", required=True),
                                    "Priority": st.column_config.SelectboxColumn("Priority", options=["HOT", "WARM", "COLD"], default="WARM", required=True),
                                    "Notes": st.column_config.TextColumn("Notes", width="large"),
                                    "Rating": st.column_config.NumberColumn("Rating", format="%.1f ⭐"),
                                    "Reviews": st.column_config.NumberColumn("Reviews"),
                                }
                                
                                staging_config["_index"] = None

                                edited_df = st.data_editor(
                                    styled_df,
                                    num_rows="fixed",
                                    hide_index=True,
                                    use_container_width=True,
                                    key=f"editor_{st.session_state.selected_scrape_id}",
                                    height=600,
                                    column_config=staging_config,
                                    column_order=final_cols
                                )
                                
                                # AUTO-SAVE LOGIC
                                if not edited_df.equals(df_file):
                                    # Convert to CSV string (in-memory)
                                    from io import StringIO
                                    csv_buffer = StringIO()
                                    edited_df.to_csv(csv_buffer, index=False)
                                    new_csv = csv_buffer.getvalue()
                                    
                                    try:
                                        # Optimistic update or silent save
                                        requests.put(
                                            f"{EXECUTIONS_API}/{st.session_state.selected_scrape_id}", 
                                            json={"fileContent": new_csv}
                                        )
                                        # Optional toast
                                        # st.toast("✅ Changes saved automatically!", icon="💾")
                                    except Exception as e:
                                        st.error(f"Auto-save failed: {e}")
                                
                                st.write("")
                                # Export Logic (Bottom)
                                import io
                                from openpyxl.worksheet.datavalidation import DataValidation
                                from openpyxl.styles import PatternFill, Font, Color, Alignment
                                from openpyxl.styles.differential import DifferentialStyle
                                from openpyxl.formatting.rule import Rule
                                from openpyxl.utils import get_column_letter

                                if st.button("📥 Export & Save to Excel", use_container_width=True, type="primary", key=f"btn_export_staging_main_{st.session_state.selected_scrape_id}"):
                                    # Prepare Data
                                    export_staging = edited_df.copy()
                                    
                                    # Create Excel
                                    buffer = io.BytesIO()
                                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                        export_staging.to_excel(writer, index=False, sheet_name='Leads')
                                        wb = writer.book
                                        ws = writer.sheets['Leads']
                                        
                                        # --- 1. SETUP DROPDOWNS (Data Validation) ---
                                        # Create hidden sheet for lists
                                        ref_sheet = wb.create_sheet("DataValidation")
                                        ref_sheet.sheet_state = 'hidden'
                                        
                                        status_options = ["Generated", "Interested", "Not picking", "Asked to call later", "Meeting set", "Meeting Done", "Proposal sent", "Follow-up scheduled", "Not interested", "Closed - Won", "Closed - Lost"]
                                        priority_options = ["HOT", "WARM", "COLD"]
                                        
                                        # Write options to hidden sheet
                                        for idx, val in enumerate(status_options, start=1):
                                            ref_sheet.cell(row=idx, column=1).value = val
                                        for idx, val in enumerate(priority_options, start=1):
                                            ref_sheet.cell(row=idx, column=2).value = val
                                            
                                        # Helper for max row
                                        max_row = len(export_staging) + 1
                                        
                                        # Apply Validation to Status Column
                                        if "Status" in export_staging.columns:
                                            s_idx = export_staging.columns.get_loc("Status") + 1
                                            s_let = get_column_letter(s_idx)
                                            dv_status = DataValidation(type="list", formula1=f"'DataValidation'!$A$1:$A${len(status_options)}", allow_blank=True)
                                            dv_status.error = 'Select valid status'
                                            ws.add_data_validation(dv_status)
                                            dv_status.add(f"{s_let}2:{s_let}{max_row}")

                                        # Apply Validation to Priority Column
                                        if "Priority" in export_staging.columns:
                                            p_idx = export_staging.columns.get_loc("Priority") + 1
                                            p_let = get_column_letter(p_idx)
                                            dv_prio = DataValidation(type="list", formula1=f"'DataValidation'!$B$1:$B${len(priority_options)}", allow_blank=True)
                                            dv_prio.error = 'Select valid priority'
                                            ws.add_data_validation(dv_prio)
                                            dv_prio.add(f"{p_let}2:{p_let}{max_row}")
                                        
                                        # --- 2. HEADER STYLING ---
                                        header_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
                                        header_font = Font(color="FFFF00", bold=True, size=11, name='Calibri')
                                        for cell in ws[1]:
                                            cell.fill = header_fill
                                            cell.font = header_font
                                            
                                        # 3. HYPERLINKS (Short text)
                                        link_font = Font(color="0563C1", underline="single")
                                        
                                        # Flexible Column Mapping
                                        # (Column Name in DF -> Link Label)
                                        link_map_targets = {
                                            "Map": "View on Map",
                                            "Map Link": "View on Map",
                                            "Google Maps Link": "View on Map",
                                            "Website": "Visit Website",
                                            "Place Url": "View on Map",
                                            "url": "View on Map"
                                        }

                                        for col_name in export_staging.columns:
                                            # Check if this column matches any target (case-insensitive)
                                            matched_label = None
                                            for target_key, label in link_map_targets.items():
                                                if col_name.strip().lower() == target_key.lower():
                                                    matched_label = label
                                                    break
                                            
                                            if matched_label:
                                                c_idx = export_staging.columns.get_loc(col_name) + 1
                                                c_let = get_column_letter(c_idx)
                                                for r in range(2, max_row + 1):
                                                    cell = ws[f"{c_let}{r}"]
                                                    val = cell.value
                                                    if val and isinstance(val, str) and len(val) > 5:
                                                        # Basic validity check
                                                        if val.startswith('http') or val.startswith('www') or 'google.com/maps' in val:
                                                            try:
                                                                cell.hyperlink = val
                                                                cell.value = matched_label
                                                                cell.font = link_font
                                                            except: pass

                                        # --- 4. CONDITIONAL FORMATTING (Colors) ---
                                        # Status Colors
                                        status_styles = {
                                            "Interested": ("#DFF5E1", "#1B5E20"),
                                            "Not picking": ("#F0F0F0", "#616161"),
                                            "Asked to call later": ("#FFF8E1", "#8D6E00"),
                                            "Meeting set": ("#E3F2FD", "#0D47A1"),
                                            "Meeting Done": ("#E0F2F1", "#004D40"),
                                            "Proposal sent": ("#F3E5F5", "#4A148C"),
                                            "Follow-up scheduled": ("#FFE0B2", "#E65100"),
                                            "Not interested": ("#FDECEA", "#B71C1C"),
                                            "Closed - Won": ("#C8E6C9", "#1B5E20"),
                                            "Closed - Lost": ("#ECEFF1", "#37474F"),
                                            "Generated": ("#FFFFFF", "#000000")
                                        }
                                        # Priority Colors
                                        prio_styles = {
                                            "HOT": ("#F25C54", "#FFFFFF"),
                                            "WARM": ("#FFE5B4", "#5A3E00"),
                                            "COLD": ("#E3F2FD", "#1E3A8A")
                                        }
                                        
                                        def apply_dxf(col_name, style_map):
                                            if col_name not in export_staging.columns: return
                                            c_idx = export_staging.columns.get_loc(col_name) + 1
                                            c_let = get_column_letter(c_idx)
                                            rng = f"{c_let}2:{c_let}{max_row}"
                                            
                                            for val, (bg, fg) in style_map.items():
                                                bg_c = Color(rgb="FF"+bg.lstrip('#'))
                                                fg_c = Color(rgb="FF"+fg.lstrip('#'))
                                                dxf = DifferentialStyle(font=Font(color=fg_c), fill=PatternFill(start_color=bg_c, end_color=bg_c, fill_type='solid'))
                                                rule = Rule(type="expression", dxf=dxf, stopIfTrue=True)
                                                rule.formula = [f'${c_let}2="{val}"']
                                                ws.conditional_formatting.add(rng, rule)
                                        
                                        apply_dxf("Status", status_styles)
                                        apply_dxf("Priority", prio_styles)
                                        
                                        # --- 5. AUTO WIDTH ---
                                        for col in ws.columns:
                                            try:
                                                max_len = 0
                                                col_let = col[0].column_letter
                                                for cell in col:
                                                    if cell.value: 
                                                        max_len = max(max_len, len(str(cell.value)))
                                                ws.column_dimensions[col_let].width = min(max_len + 2, 50)
                                            except: pass

                                    st.download_button(
                                        label="✅ Download Excel",
                                        data=buffer.getvalue(),
                                        file_name=f"{full_data.get('name')}_Export.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"dl_staging_{st.session_state.selected_scrape_id}"
                                    )
                            

                            

                        else:
                            st.warning("⚠️ This record has no CSV content attached.")
                    else:
                        st.error("Failed to load details.")
                except Exception as e:
                    st.error(f"Error loading data: {e}")
            else:
                st.info("👈 Select a scrape session from the left to view data.")

# ================== EMAIL VERIFIER ==================
if "Email Verifier" in page:
    render_email_verifier()

# ================== ANALYTICS (CEO) ==================
if "Analytics" in page:
    render_analytics_dashboard()
    
# ================== TASKS (Intern) ==================
if "Tasks" in page:
    render_tasks_module(current_user, auth_manager)

# ================== WEBSITES ==================
if "Websites" in page:
    import builtins
    html_path = "/Users/satyajeetsinhrathod/Desktop/n8n-backend/websites_showcase.html"
    try:
        with builtins.open(html_path, 'r', encoding='utf-8') as f:
            html_data = f.read()
            # Inject CSS to remove all Streamlit padding and make it absolute full screen
            st.markdown("""
                <style>
                .block-container {
                    padding-top: 0rem !important;
                    padding-bottom: 0rem !important;
                    padding-left: 0rem !important;
                    padding-right: 0rem !important;
                    max-width: 100% !important;
                    width: 100% !important;
                }
                
                div[data-testid="stVerticalBlock"] {
                    gap: 0rem !important;
                }

                div[data-testid="stVerticalBlock"] > div {
                    padding-top: 0rem !important;
                    padding-bottom: 0rem !important;
                    margin-top: 0rem !important;
                    margin-bottom: 0rem !important;
                }

                header[data-testid="stHeader"] {
                    display: none !important;
                }

                iframe {
                    border: none !important;
                    margin: 0 !important;
                    padding: 0 !important;
                }
                </style>
            """, unsafe_allow_html=True)
            
            # Set height to 100vh of the user's screen minus any sidebar, typically around 1000px works best for full immersion
            st.components.v1.html(html_data, height=1000, scrolling=True)
    except FileNotFoundError:
        st.error("Website Showcase HTML file not found.")
 